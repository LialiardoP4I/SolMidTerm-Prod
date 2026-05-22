# -*- coding: utf-8 -*-
"""
================================================================================
GESTIONE DIPENDENZE - VERSIONE V2 (CONFIGURAZIONE ESTERNA)
================================================================================

Evoluzione del modulo gestione_dipendenze_OPTIMIZED.py.

DIFFERENZE RISPETTO ALLA VERSIONE PRECEDENTE:
----------------------------------------------
Tutte le logiche di business precedentemente hardcoded nel codice sono state
esternalizzate nei fogli di configurazione del file Mappatura_v2.xlsx:

- Foglio "Regole_Colonne":   definizione colonne calcolate e regole di calcolo
- Foglio "Pattern_Negativi": pattern che forzano il valore a "Not"
- Foglio "Sostituzioni_Post": sostituzioni condizionali post-elaborazione

Il codice Python e' ora un MOTORE GENERICO che legge la configurazione e la
applica. Per aggiungere/modificare pack, sedili, pattern negativi o sostituzioni
basta modificare il file Excel, senza toccare il codice.

FLUSSO DI ELABORAZIONE:
------------------------
1. Caricamento Mappatura_v2.xlsx:
   - Foglio Schema -> Dizionario CHAR/VALUE -> Bundle/Descrizione/Readable
   - Foglio Regole_Colonne -> Lista colonne calcolate + regole
   - Foglio Pattern_Negativi -> Pattern di esclusione per colonna
   - Foglio Sostituzioni_Post -> Regole di sostituzione post-elaborazione
2. Lettura BOM150 -> DataFrame con condizioni per ogni componente
3. Parsing condizioni -> Estrazione operatori (eq, ne, IN) e valori
4. Espansione OR/IN -> Ogni condizione diventa righe multiple
5. Popolamento matrice -> Ogni riga indica quali opzioni sono attive
6. Calcolo colonne derivate -> Guidato da Regole_Colonne
7. Applicazione pattern negativi -> Guidato da Pattern_Negativi
8. Sostituzioni post-elaborazione -> Guidato da Sostituzioni_Post
9. Salvataggio -> Excel con formato tabella

@author: AntonioLiardo-SMOPS
"""

import pandas as pd
import re
from typing import Dict, Any, List, Tuple, Optional
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from collections import OrderedDict

# ============================================================================
# CONFIGURAZIONE
# ============================================================================

# Flag per abilitare/disabilitare file di debug
# In produzione: False -> evita scritture disco non necessarie
# In sviluppo: True -> genera file di debug per analisi
DEBUG_MODE = False


# ============================================================================
# CLASSE PRINCIPALE: ConditionParser
# ============================================================================

class ConditionParser:
    """
    Parser per condizioni logiche presenti nel BOM150.

    Le condizioni nel BOM150 seguono una sintassi specifica:
    - $root.CHAR_CODE eq 'VALUE'    -> uguaglianza
    - $root.CHAR_CODE ne 'VALUE'    -> disuguaglianza
    - $root.CHAR_CODE IN ('V1','V2') -> appartenenza a lista
    - condition1 OR condition2       -> disgiunzione logica

    Questa classe:
    1. Carica il mapping CHAR/VALUE -> descrizioni human-readable
    2. Parsa le condizioni estraendo operatori e valori
    3. Fornisce lookup veloce per tradurre codici in descrizioni

    OTTIMIZZAZIONE: Usa dizionario nested per lookup O(1)
    """

    def __init__(self, mapping_df=None):
        """
        Inizializza il parser con il DataFrame di mapping.

        Args:
            mapping_df: DataFrame con colonne CHAR, VALUE, Bundle, Descrizione, Readable
        """
        self.mapping_dict = self._build_mapping_dict(mapping_df) if mapping_df is not None else {}

    def _build_mapping_dict(self, mapping_df: pd.DataFrame) -> Dict[str, Dict[str, List[Tuple[str, str, str]]]]:
        """
        Costruisce dizionario di lookup: CHAR -> {VALUE -> [(Bundle, Descrizione, Readable)]}

        OTTIMIZZAZIONE: itertuples() e' 10-100x piu' veloce di iterrows()
        """
        mapping = {}

        for row in mapping_df.itertuples(index=False):
            char = str(getattr(row, 'CHAR', '')).strip()
            value = str(getattr(row, 'VALUE', '')).strip()
            bundle = str(getattr(row, 'Bundle', '')).strip()
            descrizione = str(getattr(row, 'Descrizione', '')).strip()
            readable = str(getattr(row, 'Readable', '')).strip()

            if not char or not value:
                continue

            if char not in mapping:
                mapping[char] = {}
            if value not in mapping[char]:
                mapping[char][value] = []

            mapping[char][value].append((bundle, descrizione, readable))

        print(f"\n[OK] Mapping: {len(mapping)} CHAR codes, {sum(len(v) for v in mapping.values())} totali")

        if DEBUG_MODE:
            debug_data = []
            for c, vals in mapping.items():
                for v, entries in vals.items():
                    for b, d, r in entries:
                        debug_data.append({'CHAR': c, 'VALUE': v, 'Bundle': b, 'Descrizione': d, 'Readable': r})
            pd.DataFrame(debug_data).to_excel('mapping_debug.xlsx', index=False)
            print(f"[OK] Debug: mapping_debug.xlsx")

        return mapping

    def get_column_info(self, char_code: str, value_code: str) -> List[Tuple[str, str, str]]:
        """
        Ottiene tutte le tuple (Bundle, Descrizione, Readable) dato CHAR e VALUE.
        Complessita': O(1) grazie al dizionario nested
        """
        return self.mapping_dict.get(char_code, {}).get(value_code, [])

    def split_or_conditions(self, condition: str) -> List[str]:
        """Divide una condizione contenente OR in parti separate."""
        if not condition or pd.isna(condition):
            return [condition]
        return [p.strip() for p in re.split(r'\s+OR\s+', condition, flags=re.IGNORECASE)]

    def parse_condition(self, condition: str) -> Dict[str, Any]:
        """
        Parsa una singola condizione ed estrae variabili, operatori e valori.

        Operatori supportati:
        - eq: uguaglianza
        - ne: disuguaglianza
        - IN: appartenenza a lista
        """
        if not condition or pd.isna(condition):
            return {}

        result = {}

        for match in re.finditer(r'\$root\.(\w+)\s+eq\s+["\']([^"\']+)["\']', condition):
            var, value = match.groups()
            result[var] = {'op': 'eq', 'value': value}

        for match in re.finditer(r'\$root\.(\w+)\s+ne\s+["\']([^"\']+)["\']', condition, re.IGNORECASE):
            var, value = match.groups()
            result[var] = {'op': 'ne', 'value': value}

        for match in re.finditer(r'\$root\.(\w+)\s+IN\s*\(([^)]+)\)', condition, re.IGNORECASE):
            var, values_str = match.groups()
            values = re.findall(r"['\"]([^'\"]+)['\"]", values_str)
            result[var] = {'op': 'IN', 'values': values}

        return result


# ============================================================================
# FUNZIONI DI CARICAMENTO E UTILITA'
# ============================================================================

def load_mapping_from_file(mapping_file: str = '.\\Input\\Mappatura_v2.xlsx') -> pd.DataFrame:
    """
    Carica il foglio Schema dal file di mappatura e crea DataFrame normalizzato.

    Args:
        mapping_file: Path al file Excel di mappatura

    Returns:
        DataFrame con colonne: Bundle, Descrizione, CHAR, VALUE, Readable
    """
    df = pd.read_excel(mapping_file, sheet_name='Schema', header=0)

    mapping_rows = []

    for row in df.itertuples(index=False):
        bundle = str(getattr(row, 'Bundle', '')).strip() if hasattr(row, 'Bundle') else ''
        char_code = str(row[1]).strip() if len(row) > 1 else ''
        char_desc = str(row[2]).strip() if len(row) > 2 else ''

        if not char_code or pd.isna(row[1]):
            continue

        col_idx = 3
        row_list = list(row)
        while col_idx < len(row_list):
            opz_col_val = row_list[col_idx] if col_idx < len(row_list) else None
            desc_col_val = row_list[col_idx + 1] if col_idx + 1 < len(row_list) else None

            value_code = str(opz_col_val).strip() if opz_col_val is not None else ''
            readable = str(desc_col_val).strip() if desc_col_val is not None else ''

            if value_code and value_code.lower() not in ['', 'nan', 'none']:
                if value_code.upper() in ['TRUE', 'FALSE', 'NE TRUE']:
                    if not readable or readable.upper() in ['TRUE', 'FALSE', 'NE TRUE', 'NAN']:
                        final_readable = value_code
                    else:
                        final_readable = readable
                else:
                    final_readable = readable if readable and readable.lower() not in ['', 'nan'] else value_code

                mapping_rows.append({
                    'Bundle': bundle if bundle and bundle.lower() not in ['nan', 'no'] else '',
                    'Descrizione': char_desc,
                    'CHAR': char_code,
                    'VALUE': value_code,
                    'Readable': final_readable
                })

            col_idx += 2

    result_df = pd.DataFrame(mapping_rows)
    print(f"[OK] Mappatura: {len(result_df)} entries da {len(df)} righe Schema")

    return result_df


def load_calculated_columns_config(mapping_file: str) -> Tuple[List[dict], Dict[str, List[str]], List[dict]]:
    """
    Carica la configurazione delle colonne calcolate dai 3 fogli di configurazione.

    Legge:
    - Foglio "Regole_Colonne": regole di calcolo per ogni colonna calcolata
    - Foglio "Pattern_Negativi": pattern che forzano il valore a "Not"
    - Foglio "Sostituzioni_Post": sostituzioni condizionali post-elaborazione

    Args:
        mapping_file: Path al file Mappatura_v2.xlsx

    Returns:
        Tuple di:
        - rules: Lista di dict con le regole, ordinate per (Colonna Calcolata, Ordine)
        - negative_patterns: Dict {nome_colonna: [lista_pattern]}
        - post_substitutions: Lista di dict con le sostituzioni post
    """

    # --- Regole Colonne ---
    df_rules = pd.read_excel(mapping_file, sheet_name='Regole_Colonne', header=0)
    rules = []
    for row in df_rules.itertuples(index=False):
        col_calcolata = str(getattr(row, 'Colonna_Calcolata', '')).strip() \
            if hasattr(row, 'Colonna_Calcolata') else str(row[0]).strip()
        cerca_in = str(getattr(row, 'Cerca_In', '')).strip() \
            if hasattr(row, 'Cerca_In') else str(row[1]).strip()
        tipo_ricerca = str(getattr(row, 'Tipo_Ricerca', '')).strip() \
            if hasattr(row, 'Tipo_Ricerca') else str(row[2]).strip()
        valore_cercato = str(getattr(row, 'Valore_Cercato', '')).strip() \
            if hasattr(row, 'Valore_Cercato') else str(row[3]).strip()
        risultato = str(getattr(row, 'Risultato', '')).strip() \
            if hasattr(row, 'Risultato') else str(row[4]).strip()
        risultato_negato = str(row[5]).strip() if len(row) > 5 and not pd.isna(row[5]) else ''
        risultato_finale = str(row[6]).strip() if len(row) > 6 and not pd.isna(row[6]) else 'come_trovato'
        ordine = int(row[7]) if len(row) > 7 and not pd.isna(row[7]) else 1

        if not col_calcolata or col_calcolata.lower() == 'nan':
            continue

        rules.append({
            'colonna_calcolata': col_calcolata,
            'cerca_in': cerca_in,
            'tipo_ricerca': tipo_ricerca.lower(),
            'valore_cercato': valore_cercato,
            'risultato': risultato,
            'risultato_negato': risultato_negato,
            'risultato_finale': risultato_finale.lower(),
            'ordine': ordine
        })

    # Ordina per ordine SOLO all'interno di ogni colonna calcolata,
    # preservando l'ordine di apparizione delle colonne nel foglio Excel.
    # Questo garantisce che l'ordine delle colonne calcolate nell'output
    # corrisponda all'ordine in cui appaiono nel foglio Regole_Colonne.
    from itertools import groupby as _groupby
    ordered_rules = []
    seen_cols = []
    col_groups = {}
    for rule in rules:
        col = rule['colonna_calcolata']
        if col not in col_groups:
            col_groups[col] = []
            seen_cols.append(col)
        col_groups[col].append(rule)
    for col in seen_cols:
        col_rules_sorted = sorted(col_groups[col], key=lambda r: r['ordine'])
        ordered_rules.extend(col_rules_sorted)
    rules = ordered_rules

    print(f"[OK] Regole Colonne: {len(rules)} regole caricate")

    # --- Pattern Negativi ---
    df_patterns = pd.read_excel(mapping_file, sheet_name='Pattern_Negativi', header=0)
    negative_patterns: Dict[str, List[str]] = {}
    for row in df_patterns.itertuples(index=False):
        col_name = str(row[0]).strip()
        pattern = str(row[1]).strip().lower()
        if col_name and pattern and col_name.lower() != 'nan':
            if col_name not in negative_patterns:
                negative_patterns[col_name] = []
            negative_patterns[col_name].append(pattern)

    print(f"[OK] Pattern Negativi: {sum(len(v) for v in negative_patterns.values())} pattern per {len(negative_patterns)} colonne")

    # --- Sostituzioni Post ---
    df_sost = pd.read_excel(mapping_file, sheet_name='Sostituzioni_Post', header=0)
    post_substitutions = []
    for row in df_sost.itertuples(index=False):
        col_target = str(row[0]).strip()
        val_attuale = str(row[1]).strip()
        col_condizione = str(row[2]).strip()
        condizione = str(row[3]).strip()
        val_nuovo = str(row[4]).strip()
        ordine = int(row[5]) if len(row) > 5 and not pd.isna(row[5]) else 1

        if col_target and col_target.lower() != 'nan':
            post_substitutions.append({
                'colonna_target': col_target,
                'valore_attuale': val_attuale,
                'colonna_condizione': col_condizione,
                'condizione': condizione.lower(),
                'valore_nuovo': val_nuovo,
                'ordine': ordine
            })

    post_substitutions.sort(key=lambda s: (s['colonna_target'], s['ordine']))
    print(f"[OK] Sostituzioni Post: {len(post_substitutions)} regole caricate")

    return rules, negative_patterns, post_substitutions


def _remove_parentheses_variant(text):
    """
    Rimuove la parte tra parentesi da un testo per normalizzazione.
    Es: "Forged(17\\"/17\\")" -> "Forged"
    """
    if not text:
        return text
    return re.sub(r'\s*\([^)]*\)', '', text).strip()


def _build_column_index(col_structure: List[Tuple]) -> Dict[Tuple, List[int]]:
    """
    Pre-calcola indice per ricerca veloce delle colonne nella matrice.
    OTTIMIZZAZIONE: lookup O(1) invece di ricerca lineare O(n).
    """
    index = {}

    for idx, col_data in enumerate(col_structure):
        if len(col_data) == 3:
            b, d, r = col_data

            key = (b, d, r)
            if key not in index:
                index[key] = []
            index[key].append(idx)

            r_norm = _remove_parentheses_variant(r)
            if r_norm != r:
                key_norm = (b, d, r_norm)
                if key_norm not in index:
                    index[key_norm] = []
                index[key_norm].append(idx)

    return index


def _build_calc_indices_from_rules(rules: List[dict], col_structure: List[Tuple],
                                    mapping_cols_start: int, mapping_cols_end: int) -> Dict[str, List[int]]:
    """
    Pre-calcola gli indici delle colonne sorgente per ogni colonna calcolata,
    basandosi sulle regole lette dal foglio Regole_Colonne.

    Per ogni regola, cerca le colonne nella matrice che corrispondono alla
    sorgente indicata (bundle o descrizione) e, se specificato, al valore cercato.

    L'ordine degli indici rispetta l'ordine delle regole (campo "Ordine"),
    garantendo che la priorita' di valutazione sia corretta.

    Args:
        rules: Lista regole dal foglio Regole_Colonne
        col_structure: Struttura colonne matrice
        mapping_cols_start: Indice prima colonna mapping (tipicamente 4)
        mapping_cols_end: Indice dopo ultima colonna mapping

    Returns:
        Dict {nome_colonna_calcolata: [lista_indici_colonne_sorgente]}
    """
    calc_indices: Dict[str, List[int]] = {}

    # Raggruppa regole per colonna calcolata (mantiene ordine)
    rules_by_col: Dict[str, List[dict]] = OrderedDict()
    for rule in rules:
        col_name = rule['colonna_calcolata']
        if col_name not in rules_by_col:
            rules_by_col[col_name] = []
        rules_by_col[col_name].append(rule)

    for col_name, col_rules in rules_by_col.items():
        indices = []

        for rule in col_rules:
            cerca_in = rule['cerca_in']
            tipo = rule['tipo_ricerca']
            valore_cercato = rule['valore_cercato']

            for idx in range(mapping_cols_start, mapping_cols_end):
                bundle, desc, readable = col_structure[idx]

                # Determina se la colonna matrice corrisponde alla regola
                if tipo == 'bundle':
                    source_match = (bundle == cerca_in)
                elif tipo == 'descrizione':
                    source_match = (desc == cerca_in)
                else:
                    source_match = False

                if not source_match:
                    continue

                # Se valore_cercato specificato, filtra per readable
                if valore_cercato != '*qualsiasi*':
                    readable_norm = _remove_parentheses_variant(readable)
                    if readable != valore_cercato and readable_norm != valore_cercato:
                        continue

                if idx not in indices:
                    indices.append(idx)

        calc_indices[col_name] = indices

    return calc_indices


def _get_ordered_calc_col_names(rules: List[dict]) -> List[str]:
    """
    Restituisce la lista ordinata (per prima apparizione) dei nomi colonne calcolate.
    Preserva l'ordine di inserimento come definito nel foglio Regole_Colonne.
    """
    seen = set()
    result = []
    for rule in rules:
        name = rule['colonna_calcolata']
        if name not in seen:
            seen.add(name)
            result.append(name)
    return result


def _get_rules_for_column(rules: List[dict], col_name: str) -> List[dict]:
    """Filtra le regole per una specifica colonna calcolata."""
    return [r for r in rules if r['colonna_calcolata'] == col_name]


# ============================================================================
# FUNZIONE DI POPOLAMENTO RIGA (GENERICA, GUIDATA DA CONFIGURAZIONE)
# ============================================================================

def _populate_row_from_condition(row, condition, parser, col_structure,
                                  calc_indices, calc_col_names, rules,
                                  negative_patterns, col_index):
    """
    Popola una riga della matrice in base alla condizione parsata.

    Questa funzione e' il cuore dell'elaborazione. Per ogni condizione del BOM150:
    1. Parsa la condizione per estrarre variabili e operatori
    2. Valorizza le colonne mapping corrispondenti
    3. Calcola le colonne derivate usando le regole dal foglio Regole_Colonne
    4. Applica i pattern negativi dal foglio Pattern_Negativi

    DIFFERENZA DALLA VERSIONE PRECEDENTE:
    Nessuna logica di business e' presente nel codice. Tutto e' guidato dai
    parametri rules e negative_patterns caricati dal file Excel.

    Args:
        row: Lista rappresentante la riga da popolare
        condition: Stringa condizione da parsare
        parser: Istanza ConditionParser
        col_structure: Lista struttura colonne [(bundle, desc, readable), ...]
        calc_indices: Dict pre-calcolato {nome_col: [indici_sorgente]}
        calc_col_names: Lista ordinata nomi colonne calcolate
        rules: Lista regole dal foglio Regole_Colonne
        negative_patterns: Dict {nome_colonna: [lista_pattern]}
        col_index: Indice pre-calcolato per lookup veloce
    """

    # === FASE 1: Popola colonne mapping dalla condizione parsata ===
    if condition and not pd.isna(condition):
        parsed = parser.parse_condition(condition)

        for var, info in parsed.items():
            if info['op'] == 'eq':
                all_matches = parser.get_column_info(var, info['value'])

                for bundle, desc, readable in all_matches:
                    target_indices = []

                    # OTTIMIZZAZIONE: cerca nell'indice O(1)
                    key = (bundle if bundle else '', desc, readable)
                    if key in col_index:
                        target_indices.extend(col_index[key])

                    readable_norm = _remove_parentheses_variant(readable)
                    if readable_norm != readable:
                        key_norm = (bundle if bundle else '', desc, readable_norm)
                        if key_norm in col_index:
                            target_indices.extend(col_index[key_norm])

                    # Fallback a ricerca lineare se necessario
                    if not target_indices:
                        for idx, col_data in enumerate(col_structure):
                            if len(col_data) == 3:
                                b, d, r = col_data
                                readable_match = (r == readable or _remove_parentheses_variant(r) == readable)
                                if bundle:
                                    context_match = (b == bundle)
                                else:
                                    context_match = (d == desc)
                                if readable_match and context_match:
                                    target_indices.append(idx)

                    for idx in target_indices:
                        row[idx] = f"eq '{info['value']}'"

            elif info['op'] == 'IN':
                for val in info['values']:
                    all_matches = parser.get_column_info(var, val)
                    for bundle, desc, readable in all_matches:
                        target_indices = []
                        for idx, col_data in enumerate(col_structure):
                            if len(col_data) == 3:
                                b, d, r = col_data
                                readable_match = (r == readable or _remove_parentheses_variant(r) == readable)
                                if bundle:
                                    context_match = (b == bundle)
                                else:
                                    context_match = (d == desc)
                                if readable_match and context_match:
                                    target_indices.append(idx)

                        for idx in target_indices:
                            row[idx] = f"eq '{val}'"

            elif info['op'] == 'ne':
                ne_value = f"NE {info['value']}"
                all_matches = parser.get_column_info(var, ne_value)

                if not all_matches:
                    all_matches = parser.get_column_info(var, info['value'])

                for bundle, desc, readable in all_matches:
                    target_indices = []
                    for idx, col_data in enumerate(col_structure):
                        if len(col_data) == 3:
                            b, d, r = col_data
                            readable_match = (r == readable or _remove_parentheses_variant(r) == readable)
                            if bundle:
                                context_match = (b == bundle)
                            else:
                                context_match = (d == desc)
                            if readable_match and context_match:
                                target_indices.append(idx)

                    for idx in target_indices:
                        row[idx] = f"ne '{info['value']}'"

    # === FASE 2: Calcola colonne derivate (guidato da Regole_Colonne) ===
    calc_offset = len(calc_col_names)

    for i, calc_name in enumerate(calc_col_names):
        value = "No"
        col_rules = _get_rules_for_column(rules, calc_name)

        for col_idx in calc_indices.get(calc_name, []):
            if not row[col_idx]:
                continue

            bundle, desc, readable = col_structure[col_idx]
            cell_value = row[col_idx]
            is_negated = 'ne' in cell_value.lower() and 'true' in cell_value.lower()

            # Cerca la regola che corrisponde a questa colonna sorgente
            matched_rule = None
            for rule in col_rules:
                cerca_in = rule['cerca_in']
                tipo = rule['tipo_ricerca']
                valore_cercato = rule['valore_cercato']

                # Verifica match sorgente
                if tipo == 'bundle' and bundle != cerca_in:
                    continue
                if tipo == 'descrizione' and desc != cerca_in:
                    continue

                # Verifica match valore cercato
                if valore_cercato != '*qualsiasi*':
                    readable_norm = _remove_parentheses_variant(readable)
                    if readable != valore_cercato and readable_norm != valore_cercato:
                        continue

                matched_rule = rule
                break

            if matched_rule is None:
                # Nessuna regola trovata: usa comportamento default (readable senza parentesi)
                value = _remove_parentheses_variant(readable)
                break

            # Applica la regola trovata
            risultato_str = matched_rule['risultato']
            risultato_negato_str = matched_rule['risultato_negato']

            # Determina il valore base (positivo)
            if risultato_str == 'readable':
                # Per colonne "come_trovato" si normalizza rimuovendo le parentesi,
                # cosi' "Ohlins Smart EC (electronic)" diventa "Ohlins Smart EC"
                # e "Standard (APAC)" diventa "Standard", come nel codice originale.
                if matched_rule['risultato_finale'] == 'come_trovato':
                    base_value = _remove_parentheses_variant(readable)
                else:
                    base_value = readable
            elif risultato_str == 'desc':
                base_value = desc
            else:
                # Stringa letterale fissa (es: "Enduro Pack - Black")
                base_value = risultato_str

            # Gestisci negazione
            if is_negated and risultato_negato_str:
                # Applica template negazione con sostituzione placeholder
                value = risultato_negato_str.replace('{readable}', readable).replace('{desc}', desc)
            else:
                value = base_value

            break  # Primo match trovato nella colonna: esci

        # === FASE 3: Applica pattern negativi (guidato da Pattern_Negativi) ===
        if calc_name in negative_patterns:
            value_lower = value.lower()
            for pattern in negative_patterns[calc_name]:
                if pattern in value_lower:
                    value = "Not"
                    break

        # === FASE 4: Applica riduzione finale (Risultato Finale) ===
        if value not in ['No', 'Not']:
            # Prendi il risultato_finale dalla prima regola della colonna
            # (tutte le regole della stessa colonna condividono lo stesso risultato_finale)
            if col_rules and col_rules[0]['risultato_finale'] == 'yes':
                value = 'Yes'

        row[-(calc_offset - i)] = value


# ============================================================================
# FUNZIONE PRINCIPALE: ELABORAZIONE BOM -> MATRICE
# ============================================================================

def process_excel_to_matrix(excel_path: str,
                            mapping_file: str = '.\\Input\\Mappatura_v2.xlsx',
                            data_sheet: str = 'BOM150',
                            output_file: str = '.\\Input\\matrix_output_v2.xlsx'):
    """
    Funzione principale: converte il file BOM150 in matrice espansa.

    PROCESSO:
    1. Carica configurazione (Schema + Regole + Pattern + Sostituzioni)
    2. Legge BOM150, filtra righe valide
    3. Costruisce struttura colonne (fisse + mapping + calcolate)
    4. Espande condizioni OR/IN in righe atomiche
    5. Popola matrice usando motore generico guidato da configurazione
    6. Applica sostituzioni post-elaborazione
    7. Salva in Excel con formato tabella

    Args:
        excel_path: Path al file BOM150 (.xlsm o .xlsx)
        mapping_file: Path al file Mappatura_v2.xlsx
        data_sheet: Nome foglio contenente BOM150
        output_file: Path file output

    Returns:
        Lista di righe della matrice generata
    """
    print(f"\n=== CARICAMENTO ===")
    # -------------------------------------------------------------------------
    # FASE 1: Caricamento dati e configurazione
    # -------------------------------------------------------------------------

    # Carica mapping CHAR/VALUE -> descrizioni (foglio Schema)
    mapping_df = load_mapping_from_file(mapping_file)

    # Carica configurazione colonne calcolate (fogli Regole, Pattern, Sostituzioni)
    rules, negative_patterns, post_substitutions = load_calculated_columns_config(mapping_file)

    # Legge BOM150 (header sulla riga 3, quindi header=2 per 0-indexing)
    df = pd.read_excel(excel_path, sheet_name=data_sheet, header=2)

    # Filtra solo righe con condizioni valide
    df = df[df['Blocco dati per archivio dipendenze'].notnull()]

    start_col = 'Blocco dati per archivio dipendenze'
    if start_col not in df.columns:
        raise ValueError(f"Colonna '{start_col}' non trovata")

    bom_data = df.copy()

    # Inizializza parser
    parser = ConditionParser(mapping_df)

    # -------------------------------------------------------------------------
    # FASE 2: Costruzione struttura colonne
    # -------------------------------------------------------------------------

    # Colonne fisse iniziali
    col_structure = [
        ('', '', 'Numero componenti'),
        ('', '', 'Condizione_Originale'),
        ('', '', 'Condizione_Espansa'),
        (start_col, '', '')
    ]

    # Colonne dal mapping
    for row in mapping_df.itertuples(index=False):
        col_tuple = (row.Bundle, row.Descrizione, row.Readable)
        col_structure.append(col_tuple)

    print(f"[OK] Colonne totali dopo mapping: {len(col_structure)}")

    # -------------------------------------------------------------------------
    # FASE 3: Colonne calcolate (da configurazione esterna)
    # -------------------------------------------------------------------------

    # Estrai nomi colonne calcolate dalle regole (ordine preservato)
    calc_col_names = _get_ordered_calc_col_names(rules)

    # Aggiungi colonne calcolate alla struttura
    for col_name in calc_col_names:
        col_structure.append(('', '', col_name))

    print(f"[OK] Colonne calcolate da configurazione: {len(calc_col_names)}")
    for name in calc_col_names:
        n_rules = len(_get_rules_for_column(rules, name))
        n_patterns = len(negative_patterns.get(name, []))
        print(f"     - {name}: {n_rules} regole, {n_patterns} pattern negativi")

    # -------------------------------------------------------------------------
    # FASE 4: Pre-calcolo indici per performance
    # -------------------------------------------------------------------------

    mapping_cols_start = 4
    mapping_cols_end = len(col_structure) - len(calc_col_names)

    calc_indices = _build_calc_indices_from_rules(
        rules, col_structure, mapping_cols_start, mapping_cols_end
    )

    col_index = _build_column_index(col_structure)

    # -------------------------------------------------------------------------
    # FASE 5: Espansione condizioni e popolamento matrice
    # -------------------------------------------------------------------------

    print(f"\n=== ELABORAZIONE CON MAPPATURE ESPANSE ===")
    print(f"Colonne totali: {len(col_structure)}")
    print(f"Righe BOM150 originali: {len(bom_data)}")

    expansion_mapping = []

    def expand_all_conditions(condition_text, original_condition):
        """Espande ricorsivamente le condizioni contenenti OR e IN."""
        if not condition_text or pd.isna(condition_text):
            return [(condition_text, original_condition)]

        or_parts = parser.split_or_conditions(condition_text)
        expanded_conditions = []

        for or_part in or_parts:
            if not or_part or pd.isna(or_part):
                expanded_conditions.append((or_part, original_condition))
                continue

            in_matches = list(re.finditer(r'\$root\.(\w+)\s+IN\s*\(([^)]+)\)', or_part, re.IGNORECASE))

            if not in_matches:
                expanded_conditions.append((or_part, original_condition))
            else:
                first_match = in_matches[0]
                var, values_str = first_match.groups()
                values = re.findall(r"['\"]([^'\"]+)['\"]", values_str)

                for value in values:
                    new_condition = or_part[:first_match.start()] + f"$root.{var} eq '{value}'" + or_part[first_match.end():]
                    recursive_expansions = expand_all_conditions(new_condition, original_condition)
                    expanded_conditions.extend(recursive_expansions)

        return expanded_conditions

    # Container per righe matrice finale
    matrix_rows = []

    for bom_row in bom_data.itertuples():
        bom_index = bom_row.Index

        condition = getattr(bom_row, start_col.replace(' ', '_').replace('/', '_'), None)
        if condition is None:
            condition = bom_data.loc[bom_index, start_col]

        numero_componenti = getattr(bom_row, 'Numero_componenti', '') if hasattr(bom_row, 'Numero_componenti') else ''
        if not numero_componenti:
            numero_componenti = bom_data.loc[bom_index, 'Numero componenti'] if 'Numero componenti' in bom_data.columns else ''

        if not condition or pd.isna(condition):
            row = ['' for _ in col_structure]
            row[0] = numero_componenti
            row[1] = condition if condition else ''
            row[2] = condition if condition else ''
            row[3] = condition if condition else ''
            matrix_rows.append(row)
            expansion_mapping.append({
                'bom_index': bom_index,
                'original_condition': condition,
                'expanded_condition': condition,
                'numero_componenti': numero_componenti
            })
        else:
            all_expanded_conditions = expand_all_conditions(condition, condition)

            for expanded_cond, original_cond in all_expanded_conditions:
                row = ['' for _ in col_structure]
                row[0] = numero_componenti
                row[1] = original_cond
                row[2] = expanded_cond
                row[3] = expanded_cond

                _populate_row_from_condition(
                    row, expanded_cond, parser, col_structure,
                    calc_indices, calc_col_names, rules,
                    negative_patterns, col_index
                )
                matrix_rows.append(row)

                expansion_mapping.append({
                    'bom_index': bom_index,
                    'original_condition': original_cond,
                    'expanded_condition': expanded_cond,
                    'numero_componenti': numero_componenti
                })

    if DEBUG_MODE:
        mapping_df_debug = pd.DataFrame(expansion_mapping)
        mapping_df_debug.to_excel('expansion_mapping_debug.xlsx', index=False)
        print(f"[OK] Mapping espansioni salvato: expansion_mapping_debug.xlsx")

    print(f"[OK] Righe totali matrice: {len(matrix_rows)}")

    # -------------------------------------------------------------------------
    # FASE 6: Post-processing - Sostituzioni condizionali (da configurazione)
    # -------------------------------------------------------------------------

    if post_substitutions:
        print(f"\n=== SOSTITUZIONI POST-ELABORAZIONE ===")

        # Pre-calcola indici delle colonne readable per lookup veloce
        col_names = [c[2] if c[2] else c[0] for c in col_structure]
        col_name_to_idx = {}
        for idx, name in enumerate(col_names):
            if name and name not in col_name_to_idx:
                col_name_to_idx[name] = idx

        for sost in post_substitutions:
            target_idx = col_name_to_idx.get(sost['colonna_target'])
            cond_idx = col_name_to_idx.get(sost['colonna_condizione'])

            if target_idx is None or cond_idx is None:
                print(f"[WARN] Sostituzione ignorata: colonna '{sost['colonna_target']}' "
                      f"o '{sost['colonna_condizione']}' non trovata")
                continue

            substituted_count = 0
            for row in matrix_rows:
                # Verifica che il valore attuale corrisponda
                if row[target_idx] != sost['valore_attuale']:
                    continue

                # Verifica la condizione sulla colonna di controllo
                condition_met = False
                if sost['condizione'] == 'non_vuoto':
                    condition_met = bool(row[cond_idx])
                elif sost['condizione'].startswith("eq '"):
                    # eq 'valore' -> verifica uguaglianza
                    check_val = sost['condizione'][4:-1]  # estrai valore tra apici
                    condition_met = (row[cond_idx] == check_val)
                elif sost['condizione'].startswith("contiene '"):
                    # contiene 'testo' -> verifica substring
                    check_val = sost['condizione'][10:-1]
                    condition_met = check_val in str(row[cond_idx]).lower()

                if condition_met:
                    row[target_idx] = sost['valore_nuovo']
                    substituted_count += 1

            print(f"[OK] Sostituzione '{sost['colonna_target']}' "
                  f"({sost['valore_attuale']} -> {sost['valore_nuovo']}): "
                  f"{substituted_count} righe modificate")

    # -------------------------------------------------------------------------
    # FASE 7: Salvataggio matrice in Excel
    # -------------------------------------------------------------------------

    print(f"\n=== SALVATAGGIO MATRICE FINALE ===")

    wb = load_workbook(output_file) if pd.io.common.file_exists(output_file) else None
    if wb and 'Matrix' in wb.sheetnames:
        ws = wb['Matrix']
        existing_r1 = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        existing_r2 = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    else:
        existing_r1, existing_r2 = None, None

    with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
        pd.DataFrame(matrix_rows).to_excel(writer, sheet_name='Matrix', index=False, header=False, startrow=3)

        ws = writer.sheets['Matrix']

        for col_idx, (bundle, desc, readable) in enumerate(col_structure, start=1):
            ws.cell(1, col_idx, bundle if bundle else (existing_r1[col_idx-1] if existing_r1 and col_idx <= len(existing_r1) else ''))
            ws.cell(2, col_idx, desc if desc else (existing_r2[col_idx-1] if existing_r2 and col_idx <= len(existing_r2) else ''))
            ws.cell(3, col_idx, readable if readable else col_structure[col_idx-1][0])

        last_col = get_column_letter(len(col_structure))
        tab = Table(displayName="MatrixTable", ref=f"A3:{last_col}{len(matrix_rows)+3}")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tab)

    print(f"\n=== RISULTATO ===")
    print(f"Salvato: {output_file}")
    print(f"Dimensioni: {len(matrix_rows)} righe x {len(col_structure)} colonne")

    return matrix_rows


# ============================================================================
# ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    excel_path = r'.\Input\BOM150_V21E.xlsm'

    process_excel_to_matrix(
        excel_path=excel_path,
        mapping_file='.\\Input\\Mappatura_v2.xlsx',
        data_sheet='BOM150',
        output_file='.\\Input\\matrix_output.xlsx'
    )
