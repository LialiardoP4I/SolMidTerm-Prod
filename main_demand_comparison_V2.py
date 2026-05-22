# -*- coding: utf-8 -*-
"""
================================================================================
DEMAND COMPARISON SCRIPT V2 - ACTUAL vs FORECAST TR
================================================================================

Versione V2 di main_demand_comparison.py.

DIFFERENZE RISPETTO A V1:
  - Importa da mc_simulator_V2 (TR TOTALV21E - Copia.xlsx di default)
  - Importa da sku_matching_V2 (COLUMN_MAPPING auto-rilevato, zero hardcoding)
  - Catalogo SKU: tutte_righe_univoche_V2.xlsx
  - Output:       demand_comparison_actual_vs_forecast_V2.xlsx
  - Multi-model:  righe con Model='MSV4+MSV4S' espanse correttamente
  - _reinit_norm_cache() chiamata prima del matching per prefissi dinamici

SCOPO:
  Calcola la domanda media per SKU con due file TR distinti (actual vs forecast)
  e produce un confronto SKU per SKU con delta domanda e delta economico.

OUTPUT Excel — un unico file multi-foglio:
  - "Con Prezzo":    SKU con prezzo disponibile (include delta_prezzo)
  - "Senza Prezzo":  SKU senza prezzo
  - "Actual":        medie raw TR_actual
  - "Forecast":      medie raw TR_forecast

================================================================================
@author: AntonioLiardo-SMOPS
================================================================================
"""

import time
import gc
import psutil
import os
import numpy as np
import pandas as pd
from typing import Dict, List

# ============================================================================
# IMPORTAZIONE MODULI V2
# ============================================================================
from tr_parser_V2 import parse_tr_file, load_monthly_tr, get_tr_for_month
from mc_simulator_V2 import (
    SimulationConfig,
    load_monthly_forecast,
    calculate_n_months_needed,
    parse_exclusions,
    run_monthly_simulations_optimized,
    verify_demand_conservation,
    build_wide_dataframe,
)
from sku_matching_V2 import (
    NormalizationCache,
    normalize_value_cached,
    load_lead_times,
    prenormalize_simulation_output_fast,
    sku_matches_combination_numpy,
    aggregate_monthly_demands_numpy,
    parse_month_to_sortable,
    get_memory_usage_mb,
    build_column_mapping_auto,
    _reinit_norm_cache,
)


# ============================================================================
# UTILITY
# ============================================================================

def get_memory_mb() -> float:
    process = psutil.Process(os.getpid())
    return process.memory_info().rss / 1024 / 1024


def load_prezzi(filepath: str = ".\\Input\\Prezzi.xlsx") -> pd.DataFrame:
    """
    Carica e normalizza la tabella prezzi da Prezzi.xlsx.

    Logica (identica a sku_matching_V2 per coerenza):
    - Legge Sheet1: Materiale, Prezzo netto, Unità di prezzo, Divisa, Data documento
    - Calcola prezzo unitario: Prezzo netto / Unità di prezzo
    - Converte in EUR con 4 livelli di fallback per il cambio valuta:
        1. File 'Cambio Valuta.xlsx' separato
        2. Foglio con "cambio"/"valuta"/"eur" nel nome dentro Prezzi.xlsx
        3. Stima da Sheet1 (per valute multiple)
        4. Default EUR=1.0
    - Per ogni Materiale prende il prezzo più recente (Data documento più alta)

    Returns:
        DataFrame con colonne [Materiale, Prezzo_Unitario], indicizzato su Materiale.
        Restituisce DataFrame vuoto se il file non esiste.
    """
    if not os.path.exists(filepath):
        print(f"  [WARN] File prezzi non trovato: {filepath} → colonna prezzo omessa")
        return pd.DataFrame(columns=["Materiale", "Prezzo_Unitario"]).set_index("Materiale")

    # Leggi Sheet1 (stessa logica di sku_matching_V2)
    all_data = pd.read_excel(filepath, sheet_name='Sheet1')

    # Identifica colonne con caratteri speciali
    material_col   = "Materiale"
    prezzo_col     = "Prezzo netto"
    data_doc_col   = "Data documento"
    unit_prezzo_col = None
    divisa_col      = None

    for col in all_data.columns:
        col_lower = str(col).lower().replace('\xa0', ' ')
        if 'unit' in col_lower and 'prezzo' in col_lower:
            unit_prezzo_col = col
        if 'divisa' in col_lower or 'currency' in col_lower:
            divisa_col = col

    if unit_prezzo_col is None or divisa_col is None:
        print(f"  [WARN] Colonne non trovate (unit_prezzo={unit_prezzo_col}, divisa={divisa_col}) → prezzo omesso")
        return pd.DataFrame(columns=["Materiale", "Prezzo_Unitario"]).set_index("Materiale")

    df = all_data[[material_col, prezzo_col, unit_prezzo_col, divisa_col, data_doc_col]].drop_duplicates()
    df["Prezzo_Unitario"] = df[prezzo_col] / df[unit_prezzo_col].replace(0, np.nan)

    # ── Cambio valuta con 4 livelli di fallback (identico a sku_matching_V2) ──
    cambio = None

    # Fallback 1: file Cambio Valuta.xlsx separato
    cambio_path = os.path.join(os.path.dirname(filepath), "Cambio Valuta.xlsx")
    try:
        if os.path.exists(cambio_path):
            cambio_temp = pd.read_excel(cambio_path)
            if len(cambio_temp) > 0 and len(cambio_temp.columns) > 0:
                code_col = rate_col = None
                for col in cambio_temp.columns:
                    cl = str(col).lower()
                    if "codice" in cl or "currency" in cl or "divisa" in cl:
                        code_col = col
                    if "eur" in cl or "tasso" in cl or "rate" in cl:
                        rate_col = col
                if code_col and rate_col:
                    cambio = cambio_temp[[code_col, rate_col]].rename(
                        columns={code_col: "Codice", rate_col: "tasso_eur"}
                    )
                    print(f"  [INFO] Tassi di cambio da Cambio Valuta.xlsx")
    except Exception as e:
        print(f"  [WARN] Errore lettura Cambio Valuta.xlsx: {e}")

    # Fallback 2: foglio con "cambio"/"valuta"/"eur" nel nome dentro Prezzi.xlsx
    if cambio is None or len(cambio) == 0:
        try:
            for sheet_name in pd.ExcelFile(filepath).sheet_names:
                if any(k in sheet_name.lower() for k in ("cambio", "valuta", "eur")):
                    tmp = pd.read_excel(filepath, sheet_name=sheet_name)
                    code_col = rate_col = None
                    for col in tmp.columns:
                        cl = str(col).lower()
                        if "codice" in cl or "currency" in cl or "divisa" in cl:
                            code_col = col
                        if "eur" in cl or "tasso" in cl or "rate" in cl:
                            rate_col = col
                    if code_col and rate_col:
                        cambio = tmp[[code_col, rate_col]].rename(
                            columns={code_col: "Codice", rate_col: "tasso_eur"}
                        )
                        print(f"  [INFO] Tassi di cambio dal foglio '{sheet_name}' in Prezzi.xlsx")
                        break
        except Exception:
            pass

    # Fallback 3: stima da Sheet1 (assume tasso=1.0 per tutte le valute)
    if cambio is None or len(cambio) == 0:
        valute = df[divisa_col].dropna().unique()
        cambio = pd.DataFrame({"Codice": valute, "tasso_eur": 1.0})
        print(f"  [INFO] Tassi di cambio non trovati → tasso=1.0 per tutte le valute ({list(valute)})")

    # Applica il cambio
    df = df.merge(cambio, left_on=divisa_col, right_on="Codice", how="left")
    n_missing = df["tasso_eur"].isna().sum()
    if n_missing > 0:
        print(f"  [WARN] {n_missing} righe con valuta non trovata → tasso=1.0 (EUR)")
    df["tasso_eur"] = df["tasso_eur"].fillna(1.0)
    df["Prezzo_Unitario"] = df["Prezzo_Unitario"] * df["tasso_eur"]

    # Prezzo più recente per Materiale
    df = (
        df[["Materiale", data_doc_col, "Prezzo_Unitario"]]
        .dropna(subset=["Prezzo_Unitario"])
        .sort_values(data_doc_col)
        .groupby("Materiale", as_index=True)[["Prezzo_Unitario"]]
        .last()
    )
    print(f"  Prezzi caricati: {len(df)} materiali")
    return df


# ============================================================================
# NORMALIZZAZIONE FILE TR (fix typo "Bunde" + rimozione riga ALTRO)
# ============================================================================

def _normalize_tr_file(tr_path: str) -> str:
    """
    Crea un file TR temporaneo con:
    1. Typo 'Bunde' -> 'Bundle' corretto in tutte le celle
    2. Righe ALTRO eliminate fisicamente dai fogli TR_Model
    3. TR restanti normalizzate a 1 se necessario
    Restituisce il path del file temporaneo.
    """
    import tempfile
    import shutil
    from openpyxl import load_workbook

    RENAME_MAP = {
        "Nome Opzione/Bunde": "Nome Opzione/Bundle",
        "CONCAT Nome Opzione/Bunde - Caratteristica": "CONCAT Nome Opzione/Bundle - Caratteristica",
        "Take Rate Model\xa0": "Take Rate Model",
    }

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    shutil.copy2(tr_path, tmp.name)

    wb = load_workbook(tmp.name)

    for ws in wb.worksheets:
        # Fix typo in tutte le celle
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    stripped = cell.value.strip()
                    if stripped in RENAME_MAP:
                        cell.value = RENAME_MAP[stripped]

        # Nei fogli TR_Model: elimina righe ALTRO e normalizza TR a 1
        if ws.title.lower().startswith('tr_model'):
            rows_data = list(ws.iter_rows(values_only=True))
            hdr_idx = None
            for i, row in enumerate(rows_data):
                if any(str(v).strip() == 'Take Rate Model' for v in row if v):
                    hdr_idx = i
                    break
            if hdr_idx is not None:
                headers = [str(v).strip() if v else '' for v in rows_data[hdr_idx]]
                tr_col_idx = headers.index('Take Rate Model') if 'Take Rate Model' in headers else None
                concat_col_idx = next(
                    (i for i, h in enumerate(headers) if 'CONCAT' in h), None
                )
                if tr_col_idx is not None and concat_col_idx is not None:
                    # Identifica righe ALTRO (1-indexed per openpyxl)
                    altro_excel_rows = []
                    for row_0idx, row in enumerate(rows_data[hdr_idx+1:], start=hdr_idx+2):
                        if not any(v for v in row if v):
                            continue
                        concat_val = str(row[concat_col_idx]).strip()
                        if 'ALTRO' in concat_val.upper():
                            altro_excel_rows.append(row_0idx)

                    # Elimina dal basso verso l'alto per non shiftare indici
                    for excel_row in sorted(altro_excel_rows, reverse=True):
                        ws.delete_rows(excel_row)
                        print(f"  [TR] Eliminata riga ALTRO (row {excel_row}) da '{ws.title}'")

                    # Ri-legge e normalizza se necessario
                    rows_data2 = list(ws.iter_rows(values_only=True))
                    hdr_idx2 = next(
                        (i for i, row in enumerate(rows_data2)
                         if any(str(v).strip() == 'Take Rate Model' for v in row if v)),
                        None
                    )
                    if hdr_idx2 is not None:
                        tr_vals = []
                        for row in rows_data2[hdr_idx2+1:]:
                            if not any(v for v in row if v):
                                continue
                            try:
                                tr_vals.append(float(row[tr_col_idx]))
                            except (TypeError, ValueError):
                                tr_vals.append(0.0)
                        total = sum(tr_vals)
                        if total > 0 and abs(total - 1.0) > 0.001:
                            written = 0
                            for row in ws.iter_rows(min_row=hdr_idx2+2, max_row=ws.max_row):
                                if written >= len(tr_vals):
                                    break
                                if any(c.value for c in row):
                                    row[tr_col_idx].value = round(tr_vals[written] / total, 6)
                                    written += 1

    wb.save(tmp.name)
    print(f"  [TR] Normalizzato '{os.path.basename(tr_path)}' -> temp file")
    return tmp.name


# ============================================================================
# SIMULAZIONE CON TR PARAMETRIZZABILE (V2)
# ============================================================================

def run_simulation_with_tr(
    tr_path: str,
    n_runs: int,
    start_month_offset: int = 0,
    label: str = ""
) -> pd.DataFrame:
    """
    Esegue la simulazione Monte Carlo usando il file TR specificato.

    Versione V2: importa da mc_simulator_V2 (stessa logica, TR default aggiornato).

    Args:
        tr_path:             Percorso al file TR Excel (actual o forecast)
        n_runs:              Numero run Monte Carlo
        start_month_offset:  Quanti mesi saltare nel forecast (0 = primo mese)
        label:               Etichetta per il logging (es. "ACTUAL", "FORECAST")

    Returns:
        DataFrame wide con colonne: config_columns + "<mese>_Run_<i>"
    """
    sep = "=" * 60
    print(f"\n{sep}")
    print(f"SIMULAZIONE MONTE CARLO V2 - TR {label}")
    print(f"  File TR: {tr_path}")
    print(f"{sep}")

    config = SimulationConfig(
        n_runs=n_runs,
        random_seed=42,
        start_month_offset=start_month_offset,
        use_int16=True,
    )

    # 1. Forecast domanda
    print("\n[1] Caricamento forecast domanda...")
    forecast_values, month_names = load_monthly_forecast(
        ".\\Input\\Total_demand.xlsx",
        start_offset=config.start_month_offset,
    )

    # 2. Calcola mesi necessari
    print("\n[2] Calcolo mesi necessari...")
    n_months_needed = calculate_n_months_needed(".\\Input\\quantity e residual.xlsx")
    if n_months_needed > len(forecast_values):
        n_months_needed = len(forecast_values)

    # 3. Carica TR (con normalizzazione typo + rimozione ALTRO)
    print(f"\n[3] Caricamento TR da: {tr_path}")
    _tmp_tr_path = _normalize_tr_file(tr_path)
    try:
        model_mix, characteristics = parse_tr_file(_tmp_tr_path)
    finally:
        try:
            os.remove(_tmp_tr_path)
        except Exception:
            pass
    print(f"  Modelli: {len(model_mix.models)}")
    print(f"  Caratteristiche: {len(characteristics)}")

    monthly_tr = None
    try:
        monthly_tr = load_monthly_tr(tr_path)
        print(f"  TR mensili: {len(monthly_tr)} mesi")
    except (FileNotFoundError, ValueError):
        print("  Nessun TR mensile trovato → uso TR globali")
    except Exception as e:
        print(f"  WARN TR mensili: {e} → uso TR globali")

    # 4. Esclusioni
    print("\n[4] Caricamento esclusioni...")
    exclusions = parse_exclusions(".\\Input\\esclusioni.xlsx")

    # 5. Simulazione
    print(f"\n[5] Simulazione ({n_runs} run, {n_months_needed} mesi)...")
    results_by_month = run_monthly_simulations_optimized(
        forecast_values=forecast_values,
        month_names=month_names,
        n_months=n_months_needed,
        model_mix=model_mix,
        characteristics=characteristics,
        exclusions=exclusions,
        config=config,
        monthly_tr=monthly_tr,
    )

    # 6. Verifica
    verify_demand_conservation(
        results_by_month=results_by_month,
        n_runs=config.n_runs,
        forecast_values=forecast_values[:n_months_needed],
    )

    # 7. DataFrame wide
    print("\n[7] Costruzione DataFrame output...")
    df_wide = build_wide_dataframe(
        results_by_month=results_by_month,
        month_names=month_names,
        n_runs=config.n_runs,
        save_to_file=False,
    )

    gc.collect()
    print(f"\n[OK] Simulazione V2 TR {label} completata — {len(df_wide)} combinazioni")
    return df_wide


# ============================================================================
# CALCOLO DOMANDA MEDIA PER SKU (V2)
# ============================================================================

def compute_mean_demands(
    simulation_output: pd.DataFrame,
    sku_catalog: pd.DataFrame,
    lead_times: pd.DataFrame,
    column_mapping: Dict[str, str],
    n_runs: int,
) -> pd.DataFrame:
    """
    Calcola la domanda media per ogni SKU a partire dal DataFrame di simulazione.

    Versione V2:
    - Chiama _reinit_norm_cache(column_mapping) per prefissi dinamici
    - Espande righe multi-modello (Model='MSV4+MSV4S' → due entry separate)
    - Nessun COLUMN_MAPPING hardcoded

    Args:
        simulation_output: DataFrame wide dal Monte Carlo
        sku_catalog:       DataFrame catalogo SKU (tutte_righe_univoche_V2.xlsx)
        lead_times:        DataFrame lead time (load_lead_times())
        column_mapping:    Mapping colonne SKU → simulazione (auto-rilevato)
        n_runs:            Numero run Monte Carlo

    Returns:
        DataFrame con colonne: SKU, Description, Quantity_Per_Bike,
                               Lead_Time_Months_Ceil, N_Matching_Combinations,
                               mean_demand
    """
    sep = "-" * 60
    print(f"\n{sep}")
    print("CALCOLO DOMANDA MEDIA PER SKU (V2)")
    print(f"{sep}")
    print(f"Memoria iniziale: {get_memory_usage_mb():.1f} MB")

    # ------------------------------------------------------------------
    # [V2] Inizializza la cache di normalizzazione con prefissi dinamici
    # ------------------------------------------------------------------
    _reinit_norm_cache(column_mapping)

    # ------------------------------------------------------------------
    # Pre-normalizzazione
    # ------------------------------------------------------------------
    normalized_df, norm_arrays_raw = prenormalize_simulation_output_fast(
        simulation_output, column_mapping
    )

    # Mappe categoria → codice per matching numpy
    from sku_matching_V2 import _norm_cache
    category_maps = {}
    norm_arrays = {}
    for col_name in norm_arrays_raw.keys():
        if col_name in normalized_df.columns:
            cat_col = normalized_df[col_name]
            if hasattr(cat_col, "cat"):
                category_maps[col_name] = {
                    cat: code for code, cat in enumerate(cat_col.cat.categories)
                }
                norm_arrays[col_name] = cat_col.cat.codes.values

    n_rows = len(normalized_df)

    # ------------------------------------------------------------------
    # Matrice numpy delle colonne run
    # ------------------------------------------------------------------
    month_run_columns = [col for col in simulation_output.columns if "_Run_" in col]

    month_prefixes = []
    seen = set()
    for col in month_run_columns:
        prefix = col.split("_Run_")[0]
        if prefix not in seen:
            month_prefixes.append(prefix)
            seen.add(prefix)
    month_prefixes.sort(key=parse_month_to_sortable)

    run_columns = []
    for prefix in month_prefixes:
        for i in range(n_runs):
            col_name = f"{prefix}_Run_{i}"
            if col_name in simulation_output.columns:
                run_columns.append(col_name)

    data_matrix = simulation_output[run_columns].values.astype(np.int16)

    # Pre-calcola indici colonne per mese
    run_col_to_idx = {col: idx for idx, col in enumerate(run_columns)}
    month_col_indices = {}
    for month_idx, prefix in enumerate(month_prefixes):
        indices = []
        for i in range(n_runs):
            col_name = f"{prefix}_Run_{i}"
            if col_name in run_col_to_idx:
                indices.append(run_col_to_idx[col_name])
        month_col_indices[month_idx] = indices

    # ------------------------------------------------------------------
    # Prepara SKU con lead time
    # ------------------------------------------------------------------
    sku_with_lt = sku_catalog.merge(
        lead_times[["SKU", "Lead_Time_Months", "Lead_Time_Months_Ceil",
                    "Quantity_Per_Bike", "Description"]],
        left_on="Numero componenti",
        right_on="SKU",
        how="left",
    )
    sku_with_lt = sku_with_lt[sku_with_lt["Lead_Time_Months_Ceil"].notna()]
    sku_with_lt = sku_with_lt[sku_with_lt["Lead_Time_Months_Ceil"] >= 1]
    sku_with_lt["Lead_Time_Months_Ceil"] = sku_with_lt["Lead_Time_Months_Ceil"].astype(int)
    sku_with_lt["Quantity_Per_Bike"] = sku_with_lt["Quantity_Per_Bike"].fillna(1.0)

    sku_groups = sku_with_lt.groupby("SKU")
    n_sku_groups = len(sku_groups)
    print(f"SKU da processare: {n_sku_groups}")

    # [V2] Trova la colonna Model nel catalogo SKU (per multi-model expansion)
    model_sku_col = next(
        (c for c in column_mapping.keys() if c.lower().strip() == "model"),
        None,
    )

    # Pre-estrai dati per ogni SKU con espansione multi-modello [V2]
    sku_ids = []
    sku_data_list = []
    for sku_id, sku_rows_group in sku_groups:
        sku_ids.append(sku_id)
        rows_values = []
        for _, row in sku_rows_group.iterrows():
            row_values = {
                sku_col: row[sku_col]
                for sku_col in column_mapping.keys()
                if sku_col in row.index
            }
            # [V2] Multi-model expansion: "MSV4+MSV4S" → due entry separate
            if model_sku_col and model_sku_col in row_values:
                model_val = str(row_values[model_sku_col]).strip()
                if "+" in model_val:
                    for single_model in model_val.split("+"):
                        expanded = dict(row_values)
                        expanded[model_sku_col] = single_model.strip()
                        rows_values.append(expanded)
                else:
                    rows_values.append(row_values)
            else:
                rows_values.append(row_values)

        sku_data_list.append({
            "values": rows_values,
            "lead_time": int(sku_rows_group["Lead_Time_Months_Ceil"].max()),
            "quantity": float(sku_rows_group["Quantity_Per_Bike"].mean()),
            "description": sku_rows_group["Description"].iloc[0]
                           if "Description" in sku_rows_group.columns else "",
            "lead_time_raw": sku_rows_group["Lead_Time_Months"].iloc[0]
                             if "Lead_Time_Months" in sku_rows_group.columns else np.nan,
        })

    # ------------------------------------------------------------------
    # Loop principale: matching + media
    # ------------------------------------------------------------------
    print(f"Matching SKU... (ogni '.' = 50 SKU)")
    start_time = time.time()
    results = []

    for sku_idx, (sku_id, sku_data) in enumerate(zip(sku_ids, sku_data_list)):
        if sku_idx % 50 == 0:
            print(".", end="", flush=True)
        if sku_idx % 500 == 0 and sku_idx > 0:
            elapsed = time.time() - start_time
            rate = sku_idx / elapsed
            remaining = (n_sku_groups - sku_idx) / rate if rate > 0 else 0
            print(f" [{sku_idx}/{n_sku_groups}] {rate:.1f} SKU/s, ETA: {remaining/60:.1f}min")

        # Matching: OR tra le righe dello SKU (incluse espansioni multi-modello)
        combined_mask = np.zeros(n_rows, dtype=bool)
        for row_values in sku_data["values"]:
            row_mask = sku_matches_combination_numpy(
                row_values, norm_arrays, category_maps, column_mapping, n_rows
            )
            combined_mask |= row_mask

        matching_indices = np.where(combined_mask)[0]
        if len(matching_indices) == 0:
            continue

        # Aggregazione domanda per i mesi del lead time
        run_demands = aggregate_monthly_demands_numpy(
            matching_indices, data_matrix, month_col_indices,
            sku_data["lead_time"], n_runs,
        )

        # Media tra i run
        mean_demand = float(np.mean(run_demands))

        results.append({
            "SKU": sku_id,
            "Description": sku_data["description"],
            "Quantity_Per_Bike": round(sku_data["quantity"], 2),
            "Lead_Time_Months": sku_data["lead_time_raw"],
            "Lead_Time_Months_Ceil": sku_data["lead_time"],
            "N_Matching_Combinations": int(len(matching_indices)),
            "mean_demand": round(mean_demand, 2),
        })

        if sku_idx % 1000 == 0:
            gc.collect()

    elapsed = time.time() - start_time
    print(f"\n\n[OK] Completato in {elapsed:.1f}s — {len(results)} SKU con match")
    return pd.DataFrame(results)


# ============================================================================
# PIPELINE PRINCIPALE (V2)
# ============================================================================

def main(
    tr_actual_path: str   = ".\\Input\\TR_actual.xlsx",
    tr_forecast_path: str = ".\\Input\\TR_forecast.xlsx",
    n_runs: int = 200,
    start_month_offset: int = 0,
):
    """
    Pipeline V2: due simulazioni (actual vs forecast) + confronto medie SKU.

    Versione V2:
    - Usa mc_simulator_V2 per la simulazione
    - COLUMN_MAPPING auto-rilevato (build_column_mapping_auto)
    - Catalogo SKU da tutte_righe_univoche_V2.xlsx
    - Output: demand_comparison_actual_vs_forecast_V2.xlsx

    Args:
        tr_actual_path:    File TR attuale
        tr_forecast_path:  File TR previsionale
        n_runs:            Numero run Monte Carlo per ogni simulazione
        start_month_offset: Offset mese nel file forecast
    """
    total_start = time.time()
    print(f"\nMemoria iniziale: {get_memory_mb():.1f} MB")

    # =========================================================================
    # STEP 1: Simulazione con TR_actual
    # =========================================================================
    print("\n" + "=" * 70)
    print("STEP 1: SIMULAZIONE V2 CON TR_ACTUAL")
    print("=" * 70)

    t1 = time.time()
    df_actual = run_simulation_with_tr(
        tr_path=tr_actual_path,
        n_runs=n_runs,
        start_month_offset=start_month_offset,
        label="ACTUAL",
    )
    print(f"Step 1 completato in {time.time() - t1:.1f}s")
    print(f"Memoria: {get_memory_mb():.1f} MB")
    gc.collect()

    # =========================================================================
    # STEP 2: Simulazione con TR_forecast
    # =========================================================================
    print("\n" + "=" * 70)
    print("STEP 2: SIMULAZIONE V2 CON TR_FORECAST")
    print("=" * 70)

    t2 = time.time()
    df_forecast = run_simulation_with_tr(
        tr_path=tr_forecast_path,
        n_runs=n_runs,
        start_month_offset=start_month_offset,
        label="FORECAST",
    )
    print(f"Step 2 completato in {time.time() - t2:.1f}s")
    print(f"Memoria: {get_memory_mb():.1f} MB")
    gc.collect()

    # =========================================================================
    # STEP 3: Caricamento dati comuni (catalogo SKU V2 e lead time)
    # =========================================================================
    print("\n" + "=" * 70)
    print("STEP 3: CARICAMENTO DATI SKU (V2)")
    print("=" * 70)

    # [V2] Catalogo da tutte_righe_univoche_V2.xlsx
    sku_catalog_raw = pd.read_excel(".\\Input\\tutte_righe_univoche_V2.xlsx")
    if "ID" in sku_catalog_raw.columns:
        sku_catalog_raw = sku_catalog_raw.drop(columns=["ID"])
    sku_catalog = sku_catalog_raw.drop_duplicates(keep="first")
    print(f"  SKU catalogo V2: {len(sku_catalog)}")

    lead_times = load_lead_times(".\\Input\\quantity e residual.xlsx")
    print(f"  Lead time SKU: {len(lead_times)}")

    # [V2] Auto-rileva COLUMN_MAPPING da colonne catalogo e simulazione
    sim_cols = [c for c in df_actual.columns if "_Run_" not in c]
    column_mapping = build_column_mapping_auto(
        sku_cols=list(sku_catalog.columns),
        sim_cols=sim_cols,
    )
    print(f"  COLUMN_MAPPING auto-rilevato: {len(column_mapping)} colonne")
    if column_mapping:
        print(f"  Esempio mapping: {list(column_mapping.items())[:3]}")

    # Identifica n_runs dal DataFrame (robusto)
    month_run_cols = [c for c in df_actual.columns if "_Run_" in c]
    if month_run_cols:
        first_prefix = month_run_cols[0].split("_Run_")[0]
        n_runs_actual = len([c for c in month_run_cols if c.startswith(first_prefix)])
    else:
        n_runs_actual = n_runs

    month_run_cols_f = [c for c in df_forecast.columns if "_Run_" in c]
    if month_run_cols_f:
        first_prefix_f = month_run_cols_f[0].split("_Run_")[0]
        n_runs_forecast = len([c for c in month_run_cols_f if c.startswith(first_prefix_f)])
    else:
        n_runs_forecast = n_runs

    # =========================================================================
    # STEP 4: Calcolo domanda media — ACTUAL
    # =========================================================================
    print("\n" + "=" * 70)
    print("STEP 4: DOMANDA MEDIA SKU V2 — TR_ACTUAL")
    print("=" * 70)

    t4 = time.time()
    df_mean_actual = compute_mean_demands(
        simulation_output=df_actual,
        sku_catalog=sku_catalog,
        lead_times=lead_times,
        column_mapping=column_mapping,
        n_runs=n_runs_actual,
    )
    print(f"Step 4 completato in {time.time() - t4:.1f}s")
    del df_actual
    gc.collect()

    # =========================================================================
    # STEP 5: Calcolo domanda media — FORECAST
    # =========================================================================
    print("\n" + "=" * 70)
    print("STEP 5: DOMANDA MEDIA SKU V2 — TR_FORECAST")
    print("=" * 70)

    t5 = time.time()
    df_mean_forecast = compute_mean_demands(
        simulation_output=df_forecast,
        sku_catalog=sku_catalog,
        lead_times=lead_times,
        column_mapping=column_mapping,
        n_runs=n_runs_forecast,
    )
    print(f"Step 5 completato in {time.time() - t5:.1f}s")
    del df_forecast
    gc.collect()

    # =========================================================================
    # STEP 6: Merge, prezzi e calcolo delta
    # =========================================================================
    print("\n" + "=" * 70)
    print("STEP 6: CONFRONTO ACTUAL vs FORECAST + PREZZI")
    print("=" * 70)

    id_cols = ["SKU", "Description", "Quantity_Per_Bike",
               "Lead_Time_Months", "Lead_Time_Months_Ceil", "N_Matching_Combinations"]

    df_merged = df_mean_actual[id_cols + ["mean_demand"]].merge(
        df_mean_forecast[["SKU", "mean_demand"]],
        on="SKU",
        how="outer",
        suffixes=("_actual", "_forecast"),
    )

    df_merged["mean_demand_actual"]   = df_merged["mean_demand_actual"].fillna(0.0).round(2)
    df_merged["mean_demand_forecast"] = df_merged["mean_demand_forecast"].fillna(0.0).round(2)

    df_merged["delta_demand"] = (
        df_merged["mean_demand_forecast"] - df_merged["mean_demand_actual"]
    ).round(2)

    prezzi = load_prezzi(".\\Input\\Prezzi.xlsx")

    df_merged = df_merged.merge(
        prezzi,
        left_on="SKU",
        right_index=True,
        how="left",
    )

    # Costo assoluto per scenario (analogia con prezzo_safety nel Safety Stock)
    df_merged["costo_actual"]   = (
        df_merged["mean_demand_actual"]   * df_merged["Prezzo_Unitario"]
    ).round(2)
    df_merged["costo_forecast"] = (
        df_merged["mean_demand_forecast"] * df_merged["Prezzo_Unitario"]
    ).round(2)
    df_merged["delta_prezzo"] = (
        df_merged["delta_demand"] * df_merged["Prezzo_Unitario"]
    ).round(2)

    df_merged = (
        df_merged
        .sort_values("delta_demand", key=abs, ascending=False)
        .reset_index(drop=True)
    )

    mask_con_prezzo = df_merged["Prezzo_Unitario"].notna()
    df_con_prezzo   = df_merged[mask_con_prezzo].reset_index(drop=True)
    df_senza_prezzo = df_merged[~mask_con_prezzo].drop(
        columns=["Prezzo_Unitario", "costo_actual", "costo_forecast", "delta_prezzo"]
    ).reset_index(drop=True)

    n_con   = len(df_con_prezzo)
    n_senza = len(df_senza_prezzo)

    print(f"  SKU totali nel confronto:              {len(df_merged)}")
    print(f"  SKU con prezzo disponibile:            {n_con}")
    print(f"  SKU senza prezzo:                      {n_senza}")
    print(f"  SKU con delta > 0 (forecast > actual): {(df_merged['delta_demand'] > 0).sum()}")
    print(f"  SKU con delta < 0 (forecast < actual): {(df_merged['delta_demand'] < 0).sum()}")
    print(f"  SKU invariati    (delta = 0):          {(df_merged['delta_demand'] == 0).sum()}")

    # =========================================================================
    # STEP 7: Salvataggio output V2
    # =========================================================================
    print("\n" + "=" * 70)
    print("STEP 7: SALVATAGGIO OUTPUT V2")
    print("=" * 70)

    os.makedirs(".\\Output", exist_ok=True)
    # [V2] Nome file output aggiornato
    out_path = ".\\Output\\demand_comparison_actual_vs_forecast_V2.xlsx"

    cols_con_prezzo = [
        "SKU", "Description", "Quantity_Per_Bike",
        "Lead_Time_Months", "Lead_Time_Months_Ceil", "N_Matching_Combinations",
        "mean_demand_actual", "mean_demand_forecast", "delta_demand",
        "Prezzo_Unitario", "costo_actual", "costo_forecast", "delta_prezzo",
    ]
    cols_senza_prezzo = [
        "SKU", "Description", "Quantity_Per_Bike",
        "Lead_Time_Months", "Lead_Time_Months_Ceil", "N_Matching_Combinations",
        "mean_demand_actual", "mean_demand_forecast", "delta_demand",
    ]

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df_con_prezzo[cols_con_prezzo].to_excel(
            writer, sheet_name="Con Prezzo", index=False
        )
        df_senza_prezzo[cols_senza_prezzo].to_excel(
            writer, sheet_name="Senza Prezzo", index=False
        )
        df_mean_actual.sort_values("mean_demand", ascending=False).to_excel(
            writer, sheet_name="Actual", index=False
        )
        df_mean_forecast.sort_values("mean_demand", ascending=False).to_excel(
            writer, sheet_name="Forecast", index=False
        )

    print(f"  [OK] {out_path}")
    print(f"       Foglio 'Con Prezzo'   → {n_con} SKU")
    print(f"       Foglio 'Senza Prezzo' → {n_senza} SKU")
    print(f"       Foglio 'Actual'       → {len(df_mean_actual)} SKU")
    print(f"       Foglio 'Forecast'     → {len(df_mean_forecast)} SKU")
    if n_con > 0:
        tot_actual   = df_con_prezzo["costo_actual"].sum()
        tot_forecast = df_con_prezzo["costo_forecast"].sum()
        tot_delta    = df_con_prezzo["delta_prezzo"].sum()
        print(f"\n  Costo totale Actual:   € {tot_actual:,.2f}")
        print(f"  Costo totale Forecast: € {tot_forecast:,.2f}")
        print(f"  Delta costo (Δ€):      € {tot_delta:+,.2f}")

    # =========================================================================
    # RIEPILOGO FINALE
    # =========================================================================
    total_elapsed = time.time() - total_start

    print("\n" + "=" * 70)
    print("PIPELINE V2 COMPLETATA!")
    print("=" * 70)
    print(f"\nTempo totale: {total_elapsed:.1f}s ({total_elapsed/60:.1f} min)")
    print(f"Memoria finale: {get_memory_mb():.1f} MB")
    print(f"\nFile generato:")
    print(f"  {out_path}")
    print(f"\nFogli Excel:")
    print(f"  'Con Prezzo'   — SKU, medie, delta_demand, Prezzo_Unitario, delta_prezzo")
    print(f"  'Senza Prezzo' — SKU, medie, delta_demand")
    print(f"  'Actual'       — domanda media grezza per TR_actual")
    print(f"  'Forecast'     — domanda media grezza per TR_forecast")

    return df_merged


# ============================================================================
# ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    # -------------------------------------------------------------------------
    # CONFIGURAZIONE — modifica questi path per puntare ai tuoi file TR
    # -------------------------------------------------------------------------
    TR_ACTUAL_PATH   = ".\\Input\\TR_actual.xlsx"
    TR_FORECAST_PATH = ".\\Input\\TR_forecast.xlsx"
    N_RUNS           = 200   # Aumenta per maggiore precisione statistica
    START_MONTH_OFFSET = 0   # 0 = primo mese del forecast

    results = main(
        tr_actual_path=TR_ACTUAL_PATH,
        tr_forecast_path=TR_FORECAST_PATH,
        n_runs=N_RUNS,
        start_month_offset=START_MONTH_OFFSET,
    )
