# -*- coding: utf-8 -*-
"""
================================================================================
GESTIONE DIPENDENZE - VERSIONE 2 (SCALABILE, ZERO HARDCODING)
================================================================================

Questa versione elimina tutti gli hardcoding presenti nella V1:

HARDCODING ELIMINATI:
---------------------
1. calculated_cols (19 tuple fisse) -> derivate dinamicamente da TR colonna B (CONCAT)
2. rider_seat_order (6 entry fisse) -> derivato dall'ordine di apparizione nel TR
3. Pattern negativi (5 if/elif su stringhe) -> gestiti tramite Bundle "Not XYZ" in Schema
4. Logica speciale Enduro/Smoked Windshield -> unificata con logica generica bundle
5. Keyword "pack","auxiliary" per normalizzare "Yes" -> usa definizione bundle dinamica
6. Nomi modelli RIMS (MSV4PP/MSV4S) -> letti dal foglio RIMS_Substitutions in Mappatura

FONTE DI VERITÀ:
----------------
- TR TOTALV21E.xlsx colonna B (CONCAT): definisce le colonne calcolate
  * Parte sinistra di " - " -> colonne caratteristiche (FAIRING COLOR, RIDER SEAT, ...)
  * "BUNDLE - NomePack" (senza NOT) -> colonne bundle (Radar Pack, Tech Pack, ...)
- Mappatura.xlsx foglio Schema: definisce il mapping CHAR/VALUE con Bundle
  * Bundle "Not XYZ" -> indica l'assenza del pack (sostituisce pattern negativi)
- Mappatura.xlsx foglio RIMS_Substitutions (opzionale): configurazione post-processing RIMS

SCALABILITÀ:
------------
- Aggiungere un nuovo pack: aggiungere "BUNDLE - NomePack" nel TR
- Aggiungere un bundle negativo: aggiungere righe con Bundle="Not NomePack" in Schema
- Modificare sostituzione RIMS: modificare foglio RIMS_Substitutions in Mappatura
- Aggiungere varianti sedile: aggiungere righe RIDER SEAT nel TR (l'ordine è la priorità)

@author: AntonioLiardo-SMOPS
"""

import os
import pandas as pd
import numpy as np
import re
from pathlib import Path
from typing import Dict, Any, List, Tuple, Optional, Set
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# ============================================================================
# CONFIGURAZIONE
# ============================================================================

DEBUG_MODE = False


# ============================================================================
# FUNZIONI DI CARICAMENTO DINAMICO DAL TR
# ============================================================================

def load_calculated_cols_from_tr(tr_file: str) -> List[Tuple[str, str, str, Optional[str]]]:
    """
    Deriva dinamicamente le colonne calcolate dal file TR (colonna B - CONCAT).

    LOGICA:
    - Legge il foglio TR_Optional più vecchio disponibile (fonte di riferimento globale)
    - Parte sinistra di " - " -> gruppi caratteristiche (es: FAIRING COLOR, RIDER SEAT)
    - "BUNDLE - NomePack" (senza NOT) -> bundle positivi (es: Radar Pack, Tech Pack)

    Il quarto elemento della tupla:
    - Per caratteristiche: uguale al nome gruppo (chiave di ricerca su DESC)
    - Per bundle: None (ricerca per Bundle name)

    Args:
        tr_file: Path al file TR Excel

    Returns:
        Lista di tuple (bundle, desc, nome_colonna, chiave_ricerca)
    """
    # Individua il foglio TR_Optional più vecchio disponibile
    xl = pd.ExcelFile(tr_file)
    sheets = xl.sheet_names

    from tr_parser_V2 import _parse_month_str as _pms
    opt_sheets = sorted(
        [s for s in sheets if s.lower().startswith('tr_optional_')],
        key=lambda s: _pms(s[len('tr_optional_'):])  # ordine CRONOLOGICO (coerente con tr_parser_V2)
    )

    if not opt_sheets:
        raise ValueError(f"Nessun foglio TR_Optional_* trovato in {tr_file}")

    first_opt = opt_sheets[0]
    print(f"[V2] TR fonte di verità: '{first_opt}'")

    tr_df = pd.read_excel(tr_file, sheet_name=first_opt, header=0)

    # Individua colonna CONCAT (seconda colonna dati)
    concat_col = None
    for col in tr_df.columns:
        if col and 'concat' in str(col).lower():
            concat_col = col
            break
    if concat_col is None:
        # fallback: prendi seconda colonna
        concat_col = tr_df.columns[1]

    calculated = []
    seen_chars = set()
    seen_bundles = set()

    for raw_val in tr_df[concat_col]:
        if not raw_val or pd.isna(raw_val):
            continue
        val_str = str(raw_val).strip()

        # Salta la riga di header interna
        if 'concat' in val_str.lower() or 'nome opzione' in val_str.lower():
            continue

        parts = val_str.split(' - ', 1)
        if len(parts) < 2:
            continue

        group = parts[0].strip()
        rest = parts[1].strip()

        if group == 'BUNDLE':
            # Bundle: scarta quelli con NOT
            if rest.upper().startswith('NOT '):
                continue
            bundle_name = rest  # già il nome completo del bundle
            if bundle_name not in seen_bundles:
                # (bundle_intestazione, desc_intestazione, nome_colonna, chiave_ricerca)
                calculated.append(('', '', bundle_name, None))
                seen_bundles.add(bundle_name)
        else:
            # Caratteristica: usa il gruppo come nome colonna e chiave ricerca
            if group not in seen_chars:
                calculated.append(('', '', group, group))
                seen_chars.add(group)

    print(f"[V2] Colonne calcolate derivate dal TR: {len(calculated)}")
    for c in calculated:
        print(f"     - '{c[2]}' (chiave: {c[3]})")

    return calculated


def load_priority_order_from_tr(tr_file: str, group_prefix: str) -> List[str]:
    """
    Estrae l'ordine di priorità per un gruppo dal TR.

    L'ordine di apparizione nel TR definisce la priorità:
    la prima opzione trovata per un componente viene usata.

    Args:
        tr_file: Path al file TR Excel
        group_prefix: Prefisso del gruppo (es: 'RIDER SEAT', 'PILLION SEAT')

    Returns:
        Lista di nomi opzioni in ordine di priorità
    """
    xl = pd.ExcelFile(tr_file)
    sheets = xl.sheet_names
    from tr_parser_V2 import _parse_month_str as _pms
    opt_sheets = sorted(
        [s for s in sheets if s.lower().startswith('tr_optional_')],
        key=lambda s: _pms(s[len('tr_optional_'):])  # ordine CRONOLOGICO (coerente con tr_parser_V2)
    )

    if not opt_sheets:
        return []

    tr_df = pd.read_excel(tr_file, sheet_name=opt_sheets[0], header=0)

    concat_col = None
    for col in tr_df.columns:
        if col and 'concat' in str(col).lower():
            concat_col = col
            break
    if concat_col is None:
        concat_col = tr_df.columns[1]

    prefix_search = group_prefix + ' - '
    order = []
    seen = set()

    for raw_val in tr_df[concat_col]:
        if not raw_val or pd.isna(raw_val):
            continue
        val_str = str(raw_val).strip()
        if val_str.startswith(prefix_search):
            option_name = val_str[len(prefix_search):]
            if option_name not in seen:
                order.append(option_name)
                seen.add(option_name)

    print(f"[V2] Ordine priorità '{group_prefix}': {order}")
    return order


def load_rims_substitutions(mapping_file: str) -> Dict[str, List[Tuple[str, str]]]:
    """
    Carica le sostituzioni RIMS dal foglio RIMS_Substitutions in Mappatura.

    Il foglio deve avere queste colonne:
    | Base_Value | Model_Code | Substitution |
    |------------|-----------|--------------|
    | Forged     | MSV4PP    | Forged(17"/17") - Pikes Peak |
    | Forged     | MSV4S     | Forged(19"/17") - Black |

    Se il foglio non esiste, ritorna dizionario vuoto (nessuna sostituzione).

    Args:
        mapping_file: Path al file Mappatura Excel

    Returns:
        Dict: {base_value: [(model_code, substitution), ...]}
    """
    try:
        xl = pd.ExcelFile(mapping_file)
        if 'RIMS_Substitutions' not in xl.sheet_names:
            print(f"[V2] Foglio RIMS_Substitutions non trovato — sostituzione RIMS disabilitata")
            return {}

        config_df = pd.read_excel(mapping_file, sheet_name='RIMS_Substitutions', header=0)
        substitutions = {}

        for row in config_df.itertuples(index=False):
            base = str(row[0]).strip()
            model = str(row[1]).strip()
            sub = str(row[2]).strip()

            if base and model and sub:
                if base not in substitutions:
                    substitutions[base] = []
                substitutions[base].append((model, sub))

        print(f"[V2] Sostituzioni RIMS caricate: {sum(len(v) for v in substitutions.values())} regole")
        return substitutions

    except Exception as e:
        print(f"[V2] RIMS_Substitutions non disponibile: {e}")
        return {}


# ============================================================================
# FUNZIONI DI UTILITÀ
# ============================================================================

def _remove_parentheses_variant(text: str) -> str:
    """Rimuove la parte tra parentesi da un testo per normalizzazione."""
    if not text:
        return text
    return re.sub(r'\s*\([^)]*\)', '', text).strip()


def _build_column_index(col_structure: List[Tuple]) -> Dict[Tuple, List[int]]:
    """
    Pre-calcola indice per ricerca veloce delle colonne nella matrice.
    Struttura: (bundle, desc, readable) -> [indici]
    """
    index = {}
    for idx, col_data in enumerate(col_structure):
        if len(col_data) >= 3:
            b, d, r = col_data[0], col_data[1], col_data[2]
            key = (b, d, r)
            if key not in index:
                index[key] = []
            index[key].append(idx)

            r_norm = _remove_parentheses_variant(r)
            if r_norm != r:
                key_norm = (b, d, r_norm)
                if key_norm not in index:
                    index[key_norm] = []
                index[key_norm].append(idx)

    return index


def _is_bundle_col(calc_name: str, calculated_cols: List[Tuple]) -> bool:
    """
    Verifica se una colonna calcolata è un bundle (chiave_ricerca = None).

    Le colonne caratteristiche (Fairing colors, Rider Seat, ...) hanno
    una chiave di ricerca sulla DESC (es: 'FAIRING COLOR').
    Le colonne bundle (Radar Pack, Tech Pack, ...) hanno chiave None.

    Args:
        calc_name: Nome della colonna calcolata
        calculated_cols: Lista completa delle colonne calcolate

    Returns:
        True se è un bundle, False se è una caratteristica
    """
    for col_tuple in calculated_cols:
        if col_tuple[2] == calc_name:
            return col_tuple[3] is None
    return False


def _get_not_bundle_name(bundle_name: str) -> str:
    """
    Ritorna il nome del bundle negativo corrispondente.
    Es: 'Tech Pack' -> 'Not Tech Pack'
    """
    return f"Not {bundle_name}"


# ============================================================================
# CLASSE PRINCIPALE: ConditionParser
# ============================================================================

class ConditionParser:
    """
    Parser per condizioni logiche presenti nel BOM150.
    Identica alla V1 — nessuna modifica necessaria.
    """

    def __init__(self, mapping_df=None, active_cv_sets: Dict[str, Set[str]] = None,
                 filter_untraced: bool = False,
                 unknown_ct: Set[str] = None,
                 unknown_cv: Set[tuple] = None):
        self.mapping_dict = self._build_mapping_dict(mapping_df) if mapping_df is not None else {}
        # active_cv_sets: {CHAR_code: {VALUE_code, ...}} — solo CV con TR > 0.
        # Dict vuoto = nessun filtro (tutti i CV vengono restituiti).
        self.active_cv_sets: Dict[str, Set[str]] = active_cv_sets or {}
        # Filtro CT/CV non tracciati
        self.filter_untraced: bool = filter_untraced
        self.unknown_ct: Set[str]    = unknown_ct or set()
        self.unknown_cv: Set[tuple]  = unknown_cv or set()  # {(ct, cv), ...}

    def _build_mapping_dict(self, mapping_df: pd.DataFrame) -> Dict[str, Dict[str, List[Tuple[str, str, str]]]]:
        mapping = {}
        for row in mapping_df.itertuples(index=False):
            char = str(getattr(row, 'CHAR', '')).strip()
            value = str(getattr(row, 'VALUE', '')).strip()
            bundle = str(getattr(row, 'Bundle', '')).strip()
            descrizione = str(getattr(row, 'Descrizione', '')).strip()
            readable = str(getattr(row, 'Readable', '')).strip()

            if not char or not value:
                continue
            if char not in mapping:
                mapping[char] = {}
            if value not in mapping[char]:
                mapping[char][value] = []
            mapping[char][value].append((bundle, descrizione, readable))

        print(f"\n[OK] Mapping: {len(mapping)} CHAR codes, {sum(len(v) for v in mapping.values())} totali")

        if DEBUG_MODE:
            debug_data = []
            for c, vals in mapping.items():
                for v, entries in vals.items():
                    for b, d, r in entries:
                        debug_data.append({'CHAR': c, 'VALUE': v, 'Bundle': b, 'Descrizione': d, 'Readable': r})
            pd.DataFrame(debug_data).to_excel('mapping_debug_v2.xlsx', index=False)

        return mapping

    def get_column_info(self, char_code: str, value_code: str) -> List[Tuple[str, str, str]]:
        return self.mapping_dict.get(char_code, {}).get(value_code, [])

    def get_all_values_for_char(self, char_code: str) -> List[str]:
        """
        Restituisce i VALUE definiti in mappatura per un dato CHAR.

        Filtri applicati (in ordine):
        1. Esclude marker interni 'NE ...' (es: 'NE MSV4PP')
        2. Se active_cv_sets è disponibile per questo CHAR, filtra solo i CV
           con TR > 0 nel TR file. Evita di espandere NOT(IN) con CV "morti"
           che la simulazione non può mai generare (TR=0 -> match impossibile).

        Usato per espandere:
        - ZPRODGRP ne 'VAL'          -> una riga eq per ogni modello eccetto VAL
        - NOT ($root.CT IN (lista))  -> una riga eq per ogni CV non nella lista
        """
        all_values = [
            v for v in self.mapping_dict.get(char_code, {}).keys()
            if not str(v).upper().startswith('NE ')
        ]
        if char_code in self.active_cv_sets:
            active = self.active_cv_sets[char_code]
            filtered = [v for v in all_values if v in active]
            if len(filtered) < len(all_values):
                print(f"  [V2] Filtro TR attivo: {char_code} "
                      f"{len(all_values)} -> {len(filtered)} CV (scartati {len(all_values)-len(filtered)} con TR=0)")
            return filtered
        return all_values

    def split_or_conditions(self, condition: str) -> List[str]:
        if not condition or pd.isna(condition):
            return [condition]
        return [p.strip() for p in re.split(r'\s+OR\s+', condition, flags=re.IGNORECASE)]

    def parse_condition(self, condition: str) -> Dict[str, Any]:
        if not condition or pd.isna(condition):
            return {}
        result = {}

        # ── STEP 1: NOT ($root.VAR eq 'VALUE') -> tratta come 'ne' ──────────
        # Registra anche gli span di questi blocchi per escluderli dall'eq scan.
        # Es: NOT ($root.ZCOUNTRY eq 'CDN') -> ZCOUNTRY: {op='ne', value='CDN'}
        not_spans = []
        for m in re.finditer(
            r'NOT\s*\(\s*\$root\.(\w+)\s+eq\s+["\']([^"\']+)["\']\s*\)',
            condition, re.IGNORECASE
        ):
            var, value = m.groups()
            result[var] = {'op': 'ne', 'value': value}
            not_spans.append((m.start(), m.end()))

        def _in_not_span(pos: int) -> bool:
            return any(s <= pos < e for s, e in not_spans)

        # ── STEP 2: eq normali (escludi quelli coperti da NOT) ──────────────
        for match in re.finditer(r'\$root\.(\w+)\s+eq\s+["\']([^"\']+)["\']', condition):
            if not _in_not_span(match.start()):
                var, value = match.groups()
                result[var] = {'op': 'eq', 'value': value}

        # ── STEP 3: ne espliciti ─────────────────────────────────────────────
        for match in re.finditer(r'\$root\.(\w+)\s+ne\s+["\']([^"\']+)["\']', condition, re.IGNORECASE):
            var, value = match.groups()
            result[var] = {'op': 'ne', 'value': value}

        # ── STEP 4: IN normali (escludi quelli dentro NOT) ──────────────────
        # Raccogli gli span dei blocchi NOT (IN) per non riprocessarli come IN.
        # Se expand_all_conditions funziona correttamente questa condizione non
        # dovrebbe mai arrivare qui, ma la protezione evita bug silenziosi.
        not_in_spans = [
            (m.start(), m.end())
            for m in re.finditer(
                r'NOT\s*\(\s*\$root\.\w+\s+IN\s*\([^)]+\)\s*\)',
                condition, re.IGNORECASE
            )
        ]

        def _in_not_in_span(pos: int) -> bool:
            return any(s <= pos < e for s, e in not_in_spans)

        for match in re.finditer(r'\$root\.(\w+)\s+IN\s*\(([^)]+)\)', condition, re.IGNORECASE):
            if not _in_not_in_span(match.start()):
                var, values_str = match.groups()
                values = re.findall(r"['\"]([^'\"]+)['\"]", values_str)
                result[var] = {'op': 'IN', 'values': values}

        # ── Filtro CT/CV non tracciati ────────────────────────────────────────
        # Comportamento (filter_untraced=True):
        # - CT sconosciuto intero   : scarta il predicato, valuta il resto (AND)
        # - eq/ne con CV sconosciuto: scarta il predicato, valuta il resto (AND)
        # - IN con CV sconosciuto   : rimuovi quel CV dalla lista; se lista vuota scarta predicato
        # - Tutti i predicati scartati: elimina intera dipendenza (return None)
        if self.filter_untraced and result:
            filtered = {}
            for ct, info in result.items():
                if ct in self.unknown_ct:
                    continue                          # CT intero non tracciato -> scarta predicato
                op = info['op']
                if op in ('eq', 'ne'):
                    if (ct, info['value']) in self.unknown_cv:
                        continue                      # CV sconosciuto -> scarta solo questo predicato
                    filtered[ct] = info
                elif op == 'IN':
                    known_values = [v for v in info['values']
                                    if (ct, v) not in self.unknown_cv]
                    if not known_values:
                        continue                      # lista svuotata -> scarta predicato
                    filtered[ct] = {**info, 'values': known_values}
                else:
                    filtered[ct] = info
            # Se tutti i predicati erano sconosciuti -> escludi intera dipendenza
            if not filtered:
                return None
            result = filtered

        return result


# ============================================================================
# CARICAMENTO MAPPING
# ============================================================================

def build_shared_cv_map(mapping_df: pd.DataFrame) -> Dict[Tuple[str, str], List[str]]:
    """
    Pre-calcola da Schema quali (CHAR, VALUE) sono condivisi tra più bundle.

    Restituisce un dizionario:
        (CHAR, VALUE) -> [bundle1, bundle2, ...]

    Se la lista ha 1 solo elemento -> CT/CV esclusivo di quel bundle.
    Se la lista ha N elementi -> CT/CV condiviso tra N bundle.

    Usato per la logica di split/discriminante:
    - Condiviso + nessun discriminante esclusivo -> split in N righe
    - Condiviso + almeno un discriminante esclusivo -> attiva solo il bundle esclusivo

    NOTA: considera solo i bundle non vuoti e non negativi ("Not ..."),
    perché i bundle negativi non discriminano la variante positiva.
    """
    cv_to_bundles: Dict[Tuple[str, str], List[str]] = {}

    for row in mapping_df.itertuples(index=False):
        char  = str(row.CHAR).strip()
        value = str(row.VALUE).strip()
        bundle = str(row.Bundle).strip()

        if not char or not value or not bundle:
            continue
        if bundle.lower().startswith('not '):
            continue  # i bundle negativi non contano come discriminanti

        key = (char, value)
        if key not in cv_to_bundles:
            cv_to_bundles[key] = []
        if bundle not in cv_to_bundles[key]:
            cv_to_bundles[key].append(bundle)

    shared = {k: v for k, v in cv_to_bundles.items() if len(v) > 1}
    exclusive = {k: v for k, v in cv_to_bundles.items() if len(v) == 1}

    print(f"[V2] CT/CV condivisi tra più bundle: {len(shared)}")
    print(f"[V2] CT/CV esclusivi di un solo bundle: {len(exclusive)}")
    if shared:
        for (c, v), bundles in list(shared.items())[:5]:
            print(f"     ({c}, {v}) -> {bundles}")

    # Restituisce il mapping completo (sia shared che exclusive)
    return cv_to_bundles


def load_mapping_from_file(mapping_file: str = '.\\Input\\Mappatura.xlsx') -> pd.DataFrame:
    """
    Carica il file di mappatura e crea DataFrame normalizzato.
    Tenta prima il formato long (sheet 'Schema_Long', colonne Bundle/CT/CT_Desc/CV/CV_Desc),
    altrimenti cade in fallback sul formato wide (sheet 'Schema', colonne Opz1/Desc1...).
    """
    xl = pd.ExcelFile(mapping_file)

    # ── Formato long (Schema_Long) ────────────────────────────────────────────
    if 'Schema_Long' in xl.sheet_names:
        df = xl.parse('Schema_Long', header=0)
        required = {'Bundle', 'CT', 'CT_Desc', 'CV', 'CV_Desc'}
        if required.issubset(set(df.columns)):
            mapping_rows = []
            for _, row in df.iterrows():
                bundle    = str(row.get('Bundle', '')).strip()
                char_code = str(row.get('CT', '')).strip()
                char_desc = str(row.get('CT_Desc', '')).strip()
                value     = str(row.get('CV', '')).strip()
                readable  = str(row.get('CV_Desc', '')).strip()

                if not char_code or char_code.lower() in ('nan', ''):
                    continue
                if not value or value.lower() in ('nan', ''):
                    continue

                if value.upper() in ('TRUE', 'FALSE', 'NE TRUE'):
                    if not readable or readable.upper() in ('TRUE', 'FALSE', 'NE TRUE', 'NAN'):
                        readable = value

                mapping_rows.append({
                    'Bundle':      bundle if bundle and bundle.lower() not in ('nan', 'no') else '',
                    'Descrizione': char_desc,
                    'CHAR':        char_code,
                    'VALUE':       value,
                    'Readable':    readable if readable and readable.lower() not in ('nan', '') else value,
                })

            result_df = pd.DataFrame(mapping_rows)
            print(f"[OK] Mappatura Long: {len(result_df)} entries da {len(df)} righe Schema_Long")
            return result_df

    # ── Fallback formato wide (Schema) ────────────────────────────────────────
    df = xl.parse('Schema', header=0)
    mapping_rows = []

    for row in df.itertuples(index=False):
        bundle = str(getattr(row, 'Bundle', '')).strip() if hasattr(row, 'Bundle') else ''
        char_code = str(row[1]).strip() if len(row) > 1 else ''
        char_desc = str(row[2]).strip() if len(row) > 2 else ''

        if not char_code or pd.isna(row[1]):
            continue

        col_idx = 3
        row_list = list(row)
        while col_idx < len(row_list):
            opz_col_val  = row_list[col_idx]     if col_idx     < len(row_list) else None
            desc_col_val = row_list[col_idx + 1] if col_idx + 1 < len(row_list) else None

            value_code = str(opz_col_val).strip()  if opz_col_val  is not None else ''
            readable   = str(desc_col_val).strip() if desc_col_val is not None else ''

            if value_code and value_code.lower() not in ('', 'nan', 'none'):
                if value_code.upper() in ('TRUE', 'FALSE', 'NE TRUE'):
                    if not readable or readable.upper() in ('TRUE', 'FALSE', 'NE TRUE', 'NAN'):
                        readable = value_code
                else:
                    readable = readable if readable and readable.lower() not in ('', 'nan') else value_code

                mapping_rows.append({
                    'Bundle':      bundle if bundle and bundle.lower() not in ('nan', 'no') else '',
                    'Descrizione': char_desc,
                    'CHAR':        char_code,
                    'VALUE':       value_code,
                    'Readable':    readable,
                })

            col_idx += 2

    result_df = pd.DataFrame(mapping_rows)
    print(f"[OK] Mappatura Wide: {len(result_df)} entries da {len(df)} righe Schema")
    return result_df


def load_mapping_from_schema_df(schema_df: pd.DataFrame) -> pd.DataFrame:
    """
    Variante di load_mapping_from_file che accetta direttamente il DataFrame
    Schema (già in memoria) invece di leggerlo da file.
    Utile quando la mappatura unificata viene costruita a runtime in
    process_all_bom_combinations.py senza passare per un file intermedio.
    """
    return _load_mapping_from_df(schema_df)


def _load_mapping_from_df(df: pd.DataFrame) -> pd.DataFrame:
    """Elabora un DataFrame Schema e restituisce il mapping normalizzato."""
    mapping_rows = []
    for row in df.itertuples(index=False):
        bundle = str(getattr(row, 'Bundle', '')).strip() if hasattr(row, 'Bundle') else ''
        char_code = str(row[1]).strip() if len(row) > 1 else ''
        char_desc = str(row[2]).strip() if len(row) > 2 else ''
        if not char_code or pd.isna(row[1]):
            continue
        col_idx = 3
        row_list = list(row)
        while col_idx < len(row_list):
            opz_col_val  = row_list[col_idx]     if col_idx     < len(row_list) else None
            desc_col_val = row_list[col_idx + 1] if col_idx + 1 < len(row_list) else None
            value_code = str(opz_col_val).strip()  if opz_col_val  is not None else ''
            readable   = str(desc_col_val).strip() if desc_col_val is not None else ''
            if value_code and value_code.lower() not in ['', 'nan', 'none']:
                if value_code.upper() in ['TRUE', 'FALSE', 'NE TRUE']:
                    final_readable = value_code if not readable or readable.upper() in ['TRUE', 'FALSE', 'NE TRUE', 'NAN'] else readable
                else:
                    final_readable = readable if readable and readable.lower() not in ['', 'nan'] else value_code
                mapping_rows.append({
                    'Bundle':     bundle if bundle and bundle.lower() not in ['nan', 'no'] else '',
                    'Descrizione': char_desc,
                    'CHAR':       char_code,
                    'VALUE':      value_code,
                    'Readable':   final_readable,
                })
            col_idx += 2
    result_df = pd.DataFrame(mapping_rows)
    print(f"[OK] Mappatura V2 (da DataFrame): {len(result_df)} entries da {len(df)} righe Schema")
    return result_df


# ============================================================================
# FUNZIONE DI POPOLAMENTO RIGA — VERSIONE 2 (SENZA HARDCODING)
# ============================================================================

def _populate_row_from_condition_v2(row, condition, parser, col_structure,
                                    calc_indices, calculated_cols, col_index,
                                    bundle_set: set):
    """
    Popola una riga della matrice in base alla condizione parsata.

    DIFFERENZE RISPETTO A V1:
    -------------------------
    1. Nessun if/elif hardcoded per Smoked Windshield, Enduro Pack, ecc.
    2. Nessun pattern negativo hardcoded (gestito dal Bundle "Not XYZ" in Schema)
    3. Logica "Yes" basata su bundle_set (non su keyword)
    4. Logica unificata: per bundle -> valore = nome bundle (o "Yes"); per caratteristiche -> readable

    Args:
        row: Lista rappresentante la riga da popolare
        condition: Stringa condizione da parsare
        parser: Istanza ConditionParser
        col_structure: Lista struttura colonne
        calc_indices: Dizionario pre-calcolato per colonne calcolate
        calculated_cols: Lista definizione colonne calcolate
        col_index: Indice pre-calcolato per lookup veloce
        bundle_set: Set dei nomi bundle positivi (da TR)
    """
    if condition and not pd.isna(condition):
        parsed = parser.parse_condition(condition)

        # parse_condition restituisce None quando filter_untraced=True
        # e tutti i predicati erano non tracciati -> componente da escludere
        if parsed is None:
            return False

        for var, info in parsed.items():
            if info['op'] == 'eq':
                all_matches = parser.get_column_info(var, info['value'])
                for bundle, desc, readable in all_matches:
                    target_indices = []
                    key = (bundle if bundle else '', desc, readable)
                    if key in col_index:
                        target_indices.extend(col_index[key])
                    readable_norm = _remove_parentheses_variant(readable)
                    if readable_norm != readable:
                        key_norm = (bundle if bundle else '', desc, readable_norm)
                        if key_norm in col_index:
                            target_indices.extend(col_index[key_norm])
                    if not target_indices:
                        for idx, col_data in enumerate(col_structure):
                            if len(col_data) == 3:
                                b, d, r = col_data
                                readable_match = (r == readable or _remove_parentheses_variant(r) == readable)
                                context_match = (b == bundle) if bundle else (d == desc)
                                if readable_match and context_match:
                                    target_indices.append(idx)
                    for idx in target_indices:
                        row[idx] = f"eq '{info['value']}'"

            elif info['op'] == 'IN':
                for val in info['values']:
                    all_matches = parser.get_column_info(var, val)
                    for bundle, desc, readable in all_matches:
                        target_indices = []
                        for idx, col_data in enumerate(col_structure):
                            if len(col_data) == 3:
                                b, d, r = col_data
                                readable_match = (r == readable or _remove_parentheses_variant(r) == readable)
                                context_match = (b == bundle) if bundle else (d == desc)
                                if readable_match and context_match:
                                    target_indices.append(idx)
                        for idx in target_indices:
                            row[idx] = f"eq '{val}'"

            elif info['op'] == 'ne':
                # Le variabili ZPRODGRP con ne sono già state espanse in condizioni
                # eq separate da expand_all_conditions -> skip per evitare doppio processing.
                if 'ZPRODGRP' in var.upper():
                    continue
                ne_value = f"NE {info['value']}"
                all_matches = parser.get_column_info(var, ne_value)
                if not all_matches:
                    all_matches = parser.get_column_info(var, info['value'])
                for bundle, desc, readable in all_matches:
                    target_indices = []
                    for idx, col_data in enumerate(col_structure):
                        if len(col_data) == 3:
                            b, d, r = col_data
                            readable_match = (r == readable or _remove_parentheses_variant(r) == readable)
                            context_match = (b == bundle) if bundle else (d == desc)
                            if readable_match and context_match:
                                target_indices.append(idx)
                    for idx in target_indices:
                        row[idx] = f"ne '{info['value']}'"

    # =========================================================================
    # CALCOLA COLONNE CALCOLATE — LOGICA UNIFICATA V2 (ZERO HARDCODING)
    # =========================================================================
    calc_offset = len(calculated_cols)

    for i, (calc_name, search_key) in enumerate([(c[2], c[3]) for c in calculated_cols]):
        value = "No"
        is_bundle = (search_key is None)  # True = bundle, False = caratteristica

        for col_idx in calc_indices.get(calc_name, []):
            if row[col_idx]:
                bundle_col, desc_col, readable_col = col_structure[col_idx]

                if is_bundle:
                    # -----------------------------------------------------------
                    # LOGICA BUNDLE (V2):
                    # - Se il Bundle inizia con "Not " -> "Not" (provvisorio, non break)
                    # - Altrimenti -> nome del bundle (definitivo, break)
                    #
                    # REGOLA: il match positivo vince sul match negativo.
                    # Un bundle può avere più indicatori (es: Radar Pack = Cruise
                    # Control OR Blind Spot Detection). Se la condizione attiva uno
                    # positivamente e un altro negativamente, la presenza del
                    # positivo deve vincere.
                    # Esempio: $root.CT001979 eq 'TRUE' AND $root.CT001007 ne 'CV001013'
                    # -> Blind Spot Detection presente -> Radar Pack = 'Yes' (non 'Not')
                    # -----------------------------------------------------------
                    if bundle_col.lower().startswith('not '):
                        value = "Not"
                        # NON fare break: continua a cercare un match positivo
                    else:
                        # Match positivo: definitivo.
                        # Usa bundle_col (nome dal Schema) invece di calc_name (nome dal TR)
                        # per preservare la capitalizzazione originale del Schema.
                        # Es: TR "Enduro pack - Black" (lowercase 'p') vs
                        #     Schema "Enduro Pack - Black" (uppercase 'P') -> usa Schema.
                        # Per bundle senza varianti (es: "Touring Pack") il risultato
                        # è lo stesso perché verrà normalizzato a "Yes" dalla logica sotto.
                        value = bundle_col
                        break
                else:
                    # -----------------------------------------------------------
                    # LOGICA CARATTERISTICA (V2):
                    # - Usa il readable INTEGRALE (con eventuali parentesi).
                    #
                    # NOTA: NON applicare _remove_parentheses_variant qui.
                    # Il simulatore MC usa i nomi opzione COMPLETI dal TR
                    # (es: "SUSPENSION - Ohlins Smart EC (electronic)").
                    # Dopo normalizzazione nel matching, il valore diventa
                    # "ohlins smart ec(electronic)". Se togliessimo le
                    # parentesi, il valore in unique_rows sarebbe
                    # "ohlins smart ec" -> MISMATCH -> 0 risultati per SKU
                    # con opzioni del tipo "Ohlins Smart EC (electronic)".
                    # -----------------------------------------------------------
                    value = readable_col
                    break  # primo match valido vince

        # ---------------------------------------------------------------------
        # NORMALIZZAZIONE FINALE (V2):
        # Per i bundle, se il valore non è "No" o "Not" -> normalizza a "Yes"
        # Nessuna keyword hardcoded: usa is_bundle
        #
        # REGOLA: TUTTI i bundle diventano "Yes" se attivi, indipendentemente
        # dal nome (anche quelli con varianti come "Enduro pack - Black").
        # Le caratteristiche categoriali (Fairing Color, Subframe, ecc.) hanno
        # is_bundle=False e NON vengono normalizzate -> mantengono i valori specifici.
        # ---------------------------------------------------------------------
        if is_bundle and value not in ['No', 'Not']:
            value = 'Yes'

        row[-(calc_offset - i)] = value

    return True  # riga inclusa nella matrice


# ============================================================================
# ALLINEAMENTO VALORI CARATTERISTICHE CON NOMI OPZIONE TR
# ============================================================================

def _build_tr_char_options(tr_file: str) -> Dict[str, List[str]]:
    """
    Legge dal TR i nomi opzione per ogni gruppo caratteristica.

    Ritorna: {group_name: [option_name, ...]}
    Es: {"FAIRING COLOR": ["Viola Blast gloss", "Rosso Ducati", ...]}

    Usato per allineare i valori scritti nella matrice con i nomi
    esatti del TR (che coincidono con quelli generati dalla simulazione).
    """
    try:
        xl = pd.ExcelFile(tr_file)
        sheets = xl.sheet_names

        from tr_parser_V2 import _parse_month_str as _pms
        opt_sheets = sorted(
            [s for s in sheets if s.lower().startswith('tr_optional_')],
            key=lambda s: _pms(s[len('tr_optional_'):])
        )
        if not opt_sheets:
            return {}

        tr_df = pd.read_excel(tr_file, sheet_name=opt_sheets[0], header=0)

        concat_col = None
        for col in tr_df.columns:
            if col and 'concat' in str(col).lower():
                concat_col = col
                break
        if concat_col is None:
            concat_col = tr_df.columns[1]

        options: Dict[str, List[str]] = {}
        for val in tr_df[concat_col]:
            if not val or pd.isna(val):
                continue
            val_str = str(val).strip()
            parts = val_str.split(' - ', 1)
            if len(parts) < 2:
                continue
            group = parts[0].strip()
            if group.upper() == 'BUNDLE':
                continue
            option = parts[1].strip()          # es. "Viola Blast gloss"
            if group not in options:
                options[group] = []
            if option not in options[group]:
                options[group].append(option)

        total = sum(len(v) for v in options.values())
        print(f"[V2] TR char options: {len(options)} gruppi, {total} opzioni totali")
        return options

    except Exception as e:
        print(f"[WARN] _build_tr_char_options: {e} -> allineamento TR disabilitato")
        return {}


def _align_to_tr_option(value: str, group: str,
                         tr_char_options: Dict[str, List[str]]) -> str:
    """
    Cerca il nome opzione TR che meglio corrisponde al readable della Mappatura.

    LOGICA (in ordine):
    1. Match esatto case-insensitive  -> ritorna il nome TR con la sua capitalizzazione
    2. Prefix match (1 candidato)     -> es. "Viola Blast" -> "Viola Blast gloss"
       Sicuro SOLO se esiste UN SOLO candidato: evita falsi positivi quando
       esistono sia "Blue gloss" che "Blue matte" (in quel caso si usa il valore originale).
    3. Nessun match / ambiguo         -> ritorna il valore originale invariato
    """
    options = tr_char_options.get(group, [])
    if not options:
        return value

    value_norm = value.lower().strip()

    # 1. Exact match
    for opt in options:
        if opt.lower().strip() == value_norm:
            return opt

    # 2. Prefix match (unambiguous)
    prefix = value_norm + ' '
    matches = [opt for opt in options if opt.lower().strip().startswith(prefix)]
    if len(matches) == 1:
        return matches[0]

    # 3. Fallback: valore originale
    return value


# ============================================================================
# FILTRO CV ATTIVI (TR > 0)
# ============================================================================

def _build_active_cv_sets(tr_file: str, mapping_df: pd.DataFrame) -> Dict[str, Set[str]]:
    """
    Costruisce {CHAR_code: {VALUE_code, ...}} con solo i CV la cui opzione TR
    ha TR > 0 in almeno un modello.

    Usa parse_tr_file() da tr_parser_V2, che:
    - Salta già interi gruppi con _all_tr_zero (TR=0 su tutti i modelli/opzioni)
    - Mantiene però le singole opzioni TR=0 dentro un gruppo altrimenti attivo

    Questo filtro agisce al livello di singola opzione all'interno del gruppo:
    se un'opzione ha TR=0 su TUTTI i modelli -> esclusa dalla espansione.

    Caratteristiche non presenti nel TR -> tutti i CV inclusi (conservativo).
    In caso di errore ritorna {} -> filtro disabilitato, comportamento invariato.
    """
    try:
        from tr_parser_V2 import parse_tr_file

        _, characteristics = parse_tr_file(tr_file)

        # Per ogni CharacteristicGroup già parsato, raccoglie i nomi delle opzioni
        # con TR > 0 in almeno un modello.
        # active_options: {group_name_lower: {option_name_lower, ...}}
        active_options: Dict[str, Set[str]] = {}
        n_zero = 0

        for char_name, char_group in characteristics.items():
            g_lower = char_name.lower()
            active_options[g_lower] = set()

            for i, concat_val in enumerate(char_group.values):
                # concat_val = "FAIRING COLOR - Viola Blast gloss"
                parts = str(concat_val).split(' - ', 1)
                if len(parts) < 2:
                    continue
                option_name = parts[1].strip()

                # TR > 0 in almeno un modello?
                has_tr = any(
                    float(tr_arr[i]) > 0
                    for tr_arr in char_group.tr.values()
                    if i < len(tr_arr)
                )
                if has_tr:
                    active_options[g_lower].add(option_name.lower())
                else:
                    n_zero += 1

        n_active_opts = sum(len(v) for v in active_options.values())
        print(f"[V2] Opzioni TR attive: {n_active_opts} | scartate TR=0: {n_zero}")

        # Individua CHAR code che hanno almeno una Descrizione nel TR
        chars_in_tr: Set[str] = set()
        for _, mrow in mapping_df.iterrows():
            if str(mrow['Descrizione']).strip().lower() in active_options:
                chars_in_tr.add(str(mrow['CHAR']).strip())

        # Cross-reference Mappatura -> {CHAR_code: {VALUE_code attivi}}
        active_cv_sets: Dict[str, Set[str]] = {}

        for _, mrow in mapping_df.iterrows():
            char_code  = str(mrow['CHAR']).strip()
            value_code = str(mrow['VALUE']).strip()
            char_desc  = str(mrow['Descrizione']).strip().lower()
            readable   = str(mrow['Readable']).strip().lower()

            if char_code not in active_cv_sets:
                active_cv_sets[char_code] = set()

            if char_code not in chars_in_tr:
                # Caratteristica non nel TR -> includi per sicurezza (conservativo)
                active_cv_sets[char_code].add(value_code)
                continue

            if readable in active_options.get(char_desc, set()):
                active_cv_sets[char_code].add(value_code)
            # else: opzione con TR=0 su tutti i modelli -> esclusa

        total_active = sum(len(v) for v in active_cv_sets.values())
        print(f"[V2] CV attivi (TR > 0): {total_active}/{len(mapping_df)}")
        return active_cv_sets

    except Exception as e:
        print(f"[WARN] _build_active_cv_sets: {e} -> filtro TR disabilitato")
        return {}


# ============================================================================
# FILTRO STATI (market-specific)
# ============================================================================

def _load_states_set(states_file: str) -> set:
    """
    Carica l'elenco degli stati/mercati da States.xlsx.
    Ritorna un set di codici (es: {'ARG', 'AUS', 'BRA', 'CAL', 'CDN', ...}).
    Se il file non esiste o non è leggibile, ritorna un set vuoto (filtro disabilitato).
    """
    if not states_file:
        print("[V2] States file non specificato -> filtro stati disabilitato")
        return set()
    try:
        df = pd.read_excel(states_file, header=0)
        col = df.columns[0]
        states = {str(v).strip().upper() for v in df[col] if v and not pd.isna(v)}
        print(f"[V2] States caricati: {sorted(states)}")
        return states
    except Exception as e:
        print(f"[WARN] _load_states_set: {e} -> filtro stati disabilitato")
        return set()


def _is_state_specific(condition_text: str, states_set: set) -> bool:
    """
    Ritorna True se la condizione espansa contiene $root.ZCOUNTRY eq 'STATE'
    per almeno uno STATE presente in states_set.
    Queste righe sono market-specific e vanno escluse dalla matrice globale.

    Esempi che ritornano True (riga scartata):
        $root.ZCOUNTRY eq 'CDN'
        $root.ZPRODGRP1 eq 'MSV4' AND $root.ZCOUNTRY eq 'ARG'

    Esempi che ritornano False (riga mantenuta):
        NOT ($root.ZCOUNTRY eq 'CDN')   ← vale per tutti tranne CDN
        $root.ZCOUNTRY ne 'CDN'         ← idem
        $root.ZPRODGRP1 eq 'MSV4'       ← nessuna clausola paese
    """
    if not condition_text or not states_set:
        return False
    for match in re.finditer(
        r'\$root\.ZCOUNTRY\s+eq\s+["\']([^"\']+)["\']',
        condition_text, re.IGNORECASE
    ):
        if match.group(1).strip().upper() in states_set:
            return True
    return False


def _check_residual_quality(res_df: pd.DataFrame,
                             residual_col: str,
                             component_col: str) -> List[Dict]:
    """
    Diagnostica la qualità del file 'quantity e residual'.
    Stampa un report a console e restituisce una lista di dict con i problemi
    rilevati (usata da _write_quality_report per generare l'Excel).

    Severità:
        ERRORE — residual/quantity null, negativi
        INFO   — residual = 0 (scartato dal filtro, atteso)
        WARN   — quantity null o zero
    """
    print("\n  --- QUALITA' INPUT: file residual/quantity ---")
    print(f"  Totale componenti nel file: {len(res_df)}")

    # ── Colonne ───────────────────────────────────────────────────────────────
    res_series = res_df[residual_col]
    mask_res_null = res_series.isna()
    mask_res_zero = (~mask_res_null) & (res_series == 0)
    mask_res_neg  = (~mask_res_null) & (res_series < 0)

    qty_col = next((c for c in res_df.columns
                    if c != component_col and c != residual_col and (
                    'quant' in c.lower() or 'qty' in c.lower()
                    or 'qt' in c.lower())), None)
    desc_col = next((c for c in res_df.columns
                     if 'testo' in c.lower() or 'descr' in c.lower()
                     or 'text' in c.lower()), None)

    if qty_col:
        qty_series = pd.to_numeric(res_df[qty_col], errors='coerce')
        mask_qty_null = qty_series.isna()
        mask_qty_zero = (~mask_qty_null) & (qty_series == 0)
        mask_qty_neg  = (~mask_qty_null) & (qty_series < 0)
    else:
        mask_qty_null = pd.Series([False] * len(res_df), index=res_df.index)
        mask_qty_zero = mask_qty_null.copy()
        mask_qty_neg  = mask_qty_null.copy()
        qty_series    = pd.Series([None] * len(res_df), index=res_df.index)

    # ── Raccolta dati per Excel ────────────────────────────────────────────────
    checks = [
        (mask_res_null, "Residual NULL/non numerico", "ERRORE"),
        (mask_res_neg,  "Residual < 0 (anomalia)",   "ERRORE"),
        (mask_qty_null, "Quantity NULL/non numerico", "WARN"),
        (mask_qty_zero, "Quantity = 0",               "WARN"),
        (mask_qty_neg,  "Quantity < 0 (anomalia)",    "ERRORE"),
        (mask_res_zero, "Residual = 0 (scartato)",    "INFO"),
    ]

    issues = []
    has_problems = False

    for mask, label, severity in checks:
        n = int(mask.sum())
        if n == 0:
            continue
        if severity in ('ERRORE', 'WARN'):
            has_problems = True

        codes = res_df.loc[mask, component_col].dropna().astype(str).tolist()
        preview = ', '.join(codes[:8]) + (f' ... (+{len(codes)-8} altri)' if len(codes) > 8 else '')
        print(f"  [{severity}] {label}: {n} componenti")
        print(f"           Codici: {preview}")

        for idx in res_df.index[mask]:
            row = res_df.loc[idx]
            issues.append({
                'Severita':    severity,
                'Problema':    label,
                'Codice':      str(row.get(component_col, '')),
                'Descrizione': str(row[desc_col]) if desc_col else '',
                'Residual':    row.get(residual_col, None),
                'Quantity':    qty_series.loc[idx] if qty_col else None,
            })

    if not has_problems:
        print("  [OK] Nessun problema ERRORE/WARN nel file residual/quantity.")
    print("  " + "-" * 55)
    return issues


def _check_price_quality(res_df: pd.DataFrame,
                          residual_col: str,
                          component_col: str,
                          prezzi_file: str,
                          prezzi_sheet: str = 'Purch Docs 26-2',
                          price_col_hint: str = 'Net price for unit') -> List[Dict]:
    """
    Verifica i prezzi dei componenti attivi (residual > 0) nel file PREZZI.xlsx.

    Per ogni codice componente attivo, controlla che esista un prezzo valido
    nel listino acquisti.

    Severità:
        ERRORE — prezzo NULL/mancante, prezzo = 0, o codice non trovato in PREZZI.xlsx
    """
    import os as _os
    print("\n  --- QUALITA' INPUT: verifica prezzi componenti ---")

    if not _os.path.isfile(prezzi_file):
        print(f"  [SKIP] File prezzi non trovato: {prezzi_file}")
        return []

    # Codici attivi: residual > 0
    _res = res_df.copy()
    _res[residual_col] = pd.to_numeric(_res[residual_col], errors='coerce')
    active_codes = set(
        _res.loc[_res[residual_col] > 0, component_col].dropna().astype(str)
    )
    print(f"  Componenti attivi (residual > 0): {len(active_codes)}")

    if not active_codes:
        print("  [SKIP] Nessun componente attivo — skip verifica prezzi.")
        return []

    # Carica PREZZI.xlsx
    try:
        pz_df = pd.read_excel(prezzi_file, sheet_name=prezzi_sheet, header=2)
    except Exception as e:
        print(f"  [WARN] Impossibile caricare {prezzi_file}: {e}")
        return []

    # Individua colonna codice Materiale
    mat_col = next((c for c in pz_df.columns
                    if 'materiale' in str(c).lower()), None)
    if mat_col is None:
        print(f"  [WARN] Colonna 'Materiale' non trovata in PREZZI.xlsx")
        return []

    # Individua colonna prezzo
    price_col = next((c for c in pz_df.columns
                      if price_col_hint.lower() in str(c).lower()), None)
    if price_col is None:
        print(f"  [WARN] Colonna prezzo ('{price_col_hint}...') non trovata in PREZZI.xlsx")
        return []

    # Colonna descrizione (opzionale)
    desc_col = next((c for c in pz_df.columns
                     if any(k in str(c).lower()
                            for k in ('testo', 'descr', 'text', 'breve'))), None)

    pz_df[mat_col]   = pz_df[mat_col].astype(str).str.strip()
    pz_df[price_col] = pd.to_numeric(pz_df[price_col], errors='coerce')

    # Filtra solo righe relative ai codici attivi
    pz_active = pz_df[pz_df[mat_col].isin(active_codes)].copy()

    issues = []
    has_problems = False

    # ── Caso 1: Prezzo NULL ───────────────────────────────────────────────────
    mask_null = pz_active[price_col].isna()
    n_null    = int(mask_null.sum())
    if n_null > 0:
        has_problems = True
        codes_null = pz_active.loc[mask_null, mat_col].tolist()
        preview = (', '.join(codes_null[:8])
                   + (f' ... (+{len(codes_null)-8} altri)' if len(codes_null) > 8 else ''))
        print(f"  [ERRORE] Prezzo NULL/assente: {n_null} componenti")
        print(f"           Codici: {preview}")
        for idx in pz_active.index[mask_null]:
            row = pz_active.loc[idx]
            issues.append({
                'Severita':    'ERRORE',
                'Problema':    'Prezzo assente (NULL)',
                'Codice':      str(row[mat_col]),
                'Descrizione': str(row[desc_col]) if desc_col else '',
                'Prezzo (€)':  None,
            })

    # ── Caso 2: Prezzo = 0 ───────────────────────────────────────────────────
    mask_zero = (~mask_null) & (pz_active[price_col] == 0)
    n_zero    = int(mask_zero.sum())
    if n_zero > 0:
        has_problems = True
        codes_zero = pz_active.loc[mask_zero, mat_col].tolist()
        preview = (', '.join(codes_zero[:8])
                   + (f' ... (+{len(codes_zero)-8} altri)' if len(codes_zero) > 8 else ''))
        print(f"  [ERRORE] Prezzo = 0: {n_zero} componenti")
        print(f"           Codici: {preview}")
        for idx in pz_active.index[mask_zero]:
            row = pz_active.loc[idx]
            issues.append({
                'Severita':    'ERRORE',
                'Problema':    'Prezzo = 0',
                'Codice':      str(row[mat_col]),
                'Descrizione': str(row[desc_col]) if desc_col else '',
                'Prezzo (€)':  0,
            })

    # ── Caso 3: Codici attivi non presenti nel file PREZZI ───────────────────
    codes_in_prezzi = set(pz_df[mat_col])
    codes_not_found = sorted(active_codes - codes_in_prezzi)
    n_missing       = len(codes_not_found)
    if n_missing > 0:
        has_problems = True
        preview = (', '.join(codes_not_found[:8])
                   + (f' ... (+{n_missing-8} altri)' if n_missing > 8 else ''))
        print(f"  [ERRORE] Codici attivi non trovati in PREZZI: {n_missing}")
        print(f"           Codici: {preview}")
        for code in codes_not_found:
            issues.append({
                'Severita':    'ERRORE',
                'Problema':    'Codice non trovato in PREZZI.xlsx',
                'Codice':      code,
                'Descrizione': '',
                'Prezzo (€)':  None,
            })

    if not has_problems:
        print("  [OK] Tutti i componenti attivi hanno un prezzo valido in PREZZI.xlsx.")
    print("  " + "-" * 55)
    return issues


def _check_ct_cv_traceability(bom_data: pd.DataFrame,
                               condition_col: str,
                               parser,
                               bom_file: str = '') -> Dict:
    """
    Scansiona tutte le condizioni BOM filtrate e verifica che ogni CT e CV
    presente nelle dipendenze sia tracciato nel mapping (Mappatura Schema).

    Restituisce un dict con tre liste di dict (usato da _write_quality_report):
        'ct_missing'     — CT non trovati nel mapping (ERRORE)
        'cv_missing'     — CV non trovati per CT noti (WARN)
        'ct_system'      — CT di sistema non espansi in matrice (INFO)

    Args:
        bom_data: DataFrame BOM
        condition_col: colonna condizioni
        parser: ConditionParser
        bom_file: nome file BOM per tracciabilità (es: BOM_150V21E.xlsx)
    """
    SYSTEM_CT = {'ZPRODGRP1', 'ZPRODGRP', 'ZCOUNTRY', 'ZCOL_RUOTE',
                 'ZCOL', 'ZRIMS', 'ZMATERIAL', 'ZPLT', 'ZMYEAR', 'ZSUPM'}

    print("\n  --- TRACCIABILITA' CT/CV: condizioni vs Mappatura Schema ---")

    mapping_ct = set(parser.mapping_dict.keys())
    unknown_ct        = {}   # {ct: {'count': n, 'esempi': [...]}}
    unknown_ct_system = {}   # {ct: n}
    unknown_cv        = {}   # {(ct, cv): {'count': n, 'esempi': [...]}}

    conditions = bom_data[condition_col].dropna()
    n_conditions = len(conditions)

    for cond in conditions:
        cond_str = str(cond)

        ct_found = set(m.group(1) for m in re.finditer(r'\$root\.(\w+)', cond_str))
        for ct in ct_found:
            if ct not in mapping_ct:
                if ct.upper() in SYSTEM_CT or not ct.startswith('CT'):
                    unknown_ct_system[ct] = unknown_ct_system.get(ct, 0) + 1
                else:
                    if ct not in unknown_ct:
                        unknown_ct[ct] = {'count': 0, 'esempi': []}
                    unknown_ct[ct]['count'] += 1
                    if len(unknown_ct[ct]['esempi']) < 3:
                        unknown_ct[ct]['esempi'].append(cond_str[:120])

        for m in re.finditer(
            r'\$root\.(\w+)\s+(?:eq|ne)\s+[\'"]([^\'"]+)[\'"]',
            cond_str, re.IGNORECASE
        ):
            ct, cv = m.group(1), m.group(2)
            if ct in mapping_ct and cv not in parser.mapping_dict.get(ct, {}):
                key = (ct, cv)
                if key not in unknown_cv:
                    unknown_cv[key] = {'count': 0, 'esempi': []}
                unknown_cv[key]['count'] += 1
                if len(unknown_cv[key]['esempi']) < 3:
                    unknown_cv[key]['esempi'].append(cond_str[:120])

        for m in re.finditer(
            r'\$root\.(\w+)\s+IN\s*\(([^)]+)\)', cond_str, re.IGNORECASE
        ):
            ct = m.group(1)
            cv_list = re.findall(r"['\"]([^'\"]+)['\"]", m.group(2))
            if ct in mapping_ct:
                for cv in cv_list:
                    if cv not in parser.mapping_dict.get(ct, {}):
                        key = (ct, cv)
                        if key not in unknown_cv:
                            unknown_cv[key] = {'count': 0, 'esempi': []}
                        unknown_cv[key]['count'] += 1
                        if len(unknown_cv[key]['esempi']) < 3:
                            unknown_cv[key]['esempi'].append(cond_str[:120])

    print(f"  Condizioni analizzate: {n_conditions}")

    if not unknown_ct and not unknown_cv:
        print("  [OK] Tutti i CT/CV nelle condizioni sono tracciati nel mapping.")
    else:
        if unknown_ct:
            print(f"  [ERRORE] CT non trovati nel mapping ({len(unknown_ct)} distinti):")
            for ct, d in sorted(unknown_ct.items(), key=lambda x: -x[1]['count']):
                print(f"    - {ct:<25s}  in {d['count']} condizioni")
        if unknown_cv:
            print(f"  [ERRORE] CV non trovati nel mapping ({len(unknown_cv)} distinti):")
            for (ct, cv), d in sorted(unknown_cv.items(), key=lambda x: -x[1]['count']):
                print(f"    - {ct:<20s} / {cv:<20s}  in {d['count']} condizioni")

    if unknown_ct_system:
        print(f"  [INFO] CT di sistema/filtro: {sorted(unknown_ct_system.keys())}")
    print("  " + "-" * 55)

    # ── Serializza per Excel ───────────────────────────────────────────────────
    ct_rows = [
        {'Severita': 'ERRORE', 'CT': ct, 'BOM': bom_file,
         'N condizioni': d['count'],
         'Esempio condizione 1': d['esempi'][0] if len(d['esempi']) > 0 else '',
         'Esempio condizione 2': d['esempi'][1] if len(d['esempi']) > 1 else '',
         'Esempio condizione 3': d['esempi'][2] if len(d['esempi']) > 2 else ''}
        for ct, d in sorted(unknown_ct.items(), key=lambda x: -x[1]['count'])
    ]
    cv_rows = [
        {'Severita': 'ERRORE', 'CT': ct, 'CV': cv, 'BOM': bom_file,
         'N condizioni': d['count'],
         'Esempio condizione 1': d['esempi'][0] if len(d['esempi']) > 0 else '',
         'Esempio condizione 2': d['esempi'][1] if len(d['esempi']) > 1 else '',
         'Esempio condizione 3': d['esempi'][2] if len(d['esempi']) > 2 else ''}
        for (ct, cv), d in sorted(unknown_cv.items(), key=lambda x: -x[1]['count'])
    ]
    sys_rows = [
        {'Severita': 'INFO', 'CT': ct, 'N condizioni': n,
         'Nota': 'CT di sistema — usato per filtraggio, non espanso in colonne matrice'}
        for ct, n in sorted(unknown_ct_system.items(), key=lambda x: -x[1])
    ]
    return {
        'ct_missing':    ct_rows,
        'cv_missing':    cv_rows,
        'ct_system':     sys_rows,
        # set grezzi per il filtro pipeline
        'unknown_ct_set': set(unknown_ct.keys()),
        'unknown_cv_set': set(unknown_cv.keys()),   # {(ct, cv), ...}
    }


def _write_quality_report(residual_issues: List[Dict],
                           traceability: Dict,
                           price_issues: List[Dict] = None,
                           bom_missing: List[Dict] = None,
                           output_path: str = '.\\Output\\quality_check_report.xlsx') -> None:
    """
    Genera un file Excel con il report di qualità degli input.

    Fogli:
        1. Residual & Quantity  — componenti con problemi nel file residual
        2. BOM Senza Residual   — componenti BOM assenti da quantity e residual.xlsx
        3. CT non tracciati     — CT nelle dipendenze assenti dalla Mappatura
        4. CV non tracciati     — CV nelle dipendenze non mappati per il CT noto
        5. CT di sistema        — CT di filtro (INFO, non sono errori)
        6. Prezzi               — componenti attivi con prezzo mancante o = 0 (opzionale)
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    SEV_COLORS = {'ERRORE': 'FFC7CE', 'WARN': 'FFEB9C', 'INFO': 'D9EAD3'}
    SEV_FONTS  = {'ERRORE': 'C00000', 'WARN': '7F6000', 'INFO': '274E13'}
    H_FILL  = PatternFill('solid', fgColor='1F3864')
    H_FONT  = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    N_FONT  = Font(name='Calibri', size=10)
    B_FONT  = Font(name='Calibri', bold=True, size=10)
    CENTER  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT    = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    def _border():
        s = Side(style='thin', color='BFBFBF')
        return Border(left=s, right=s, top=s, bottom=s)

    def _write_sheet(ws, title: str, subtitle: str,
                     headers: List[str], rows: List[Dict],
                     col_widths: List[int], sev_col: str = 'Severita'):
        ws.sheet_view.showGridLines = False

        # Titolo
        ws.merge_cells(f'A1:{chr(64+len(headers))}1')
        ws['A1'] = title
        ws['A1'].font      = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
        ws['A1'].fill      = H_FILL
        ws['A1'].alignment = CENTER
        ws.row_dimensions[1].height = 28

        # Sottotitolo
        ws.merge_cells(f'A2:{chr(64+len(headers))}2')
        ws['A2'] = subtitle
        ws['A2'].font      = Font(name='Calibri', italic=True, size=10, color='FFFFFF')
        ws['A2'].fill      = PatternFill('solid', fgColor='2E4F8A')
        ws['A2'].alignment = CENTER
        ws.row_dimensions[2].height = 18

        # Intestazioni
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=c, value=h)
            cell.font = H_FONT; cell.fill = H_FILL
            cell.alignment = CENTER; cell.border = _border()
        ws.row_dimensions[3].height = 36

        if not rows:
            ws.merge_cells(f'A4:{chr(64+len(headers))}4')
            ok = ws['A4']
            ok.value = 'Nessun problema rilevato'
            ok.font  = Font(name='Calibri', italic=True, size=10, color='27AE60')
            ok.alignment = CENTER
            ok.fill = PatternFill('solid', fgColor='D9EAD3')
        else:
            for r, row_data in enumerate(rows, 4):
                sev = str(row_data.get(sev_col, 'INFO'))
                bg  = SEV_COLORS.get(sev, 'FFFFFF')
                fc  = SEV_FONTS.get(sev, '000000')
                for c, h in enumerate(headers, 1):
                    val  = row_data.get(h, '')
                    cell = ws.cell(row=r, column=c, value=val)
                    cell.border    = _border()
                    cell.alignment = CENTER if c <= 3 else LEFT
                    if h == sev_col:
                        cell.fill = PatternFill('solid', fgColor=bg)
                        cell.font = Font(name='Calibri', bold=True,
                                         size=10, color=fc)
                    else:
                        cell.fill = PatternFill('solid', fgColor='F9F9F9'
                                                if r % 2 == 0 else 'FFFFFF')
                        cell.font = N_FONT
                ws.row_dimensions[r].height = 18

        from openpyxl.utils import get_column_letter
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
    wb = Workbook()

    # ── Foglio 1: Residual & Quantity ────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Residual & Quantity'
    _write_sheet(
        ws1,
        title    = 'QUALITA\' INPUT — Residual & Quantity',
        subtitle = ('ERRORE = da correggere urgentemente  |  WARN = verificare  |  '
                    'INFO = residual=0 (componente incluso in BOM ma senza residual attivo)'),
        headers  = ['Severita', 'Problema', 'Codice', 'Descrizione',
                    'Residual (mesi)', 'Quantity'],
        rows     = residual_issues,
        col_widths = [10, 28, 16, 36, 16, 12],
    )

    # ── Foglio 2: BOM Senza Residual ─────────────────────────────────────────
    ws_bom = wb.create_sheet('BOM Senza Residual')
    _write_sheet(
        ws_bom,
        title    = 'BOM SENZA RESIDUAL — componenti BOM assenti da quantity e residual.xlsx',
        subtitle = ('Questi componenti sono presenti nella BOM con una dipendenza ma non hanno '
                    'dati di lead time/residual. Safety Stock = 0. Aggiungere al file residual.'),
        headers  = ['Numero componenti', 'Descrizione', 'Blocco dipendenze'],
        rows     = (bom_missing or []),
        col_widths = [20, 40, 80],
        sev_col  = 'Numero componenti',   # nessuna colonna severità, usa codice
    )

    # ── Foglio 3: CT non tracciati ───────────────────────────────────────────
    ws2 = wb.create_sheet('CT non tracciati')
    _write_sheet(
        ws2,
        title    = 'CT NON TRACCIATI — presenti in dipendenze BOM ma assenti dalla Mappatura',
        subtitle = ('Questi CT non vengono espansi in colonne della matrice. '
                    'Aggiungere alla Mappatura Schema oppure verificare se sono obsoleti.'),
        headers  = ['Severita', 'CT', 'BOM', 'N condizioni',
                    'Esempio condizione 1', 'Esempio condizione 2', 'Esempio condizione 3'],
        rows     = traceability.get('ct_missing', []),
        col_widths = [10, 14, 20, 14, 55, 55, 55],
    )

    # ── Foglio 4: CV non tracciati ───────────────────────────────────────────
    ws3 = wb.create_sheet('CV non tracciati')
    _write_sheet(
        ws3,
        title    = 'CV NON TRACCIATI — valori presenti in dipendenze ma non mappati per il CT',
        subtitle = ('Il CT esiste nella Mappatura ma questo specifico valore (CV) non e\' definito. '
                    'Aggiungere l\'opzione alla riga CT corrispondente nella Mappatura Schema.'),
        headers  = ['Severita', 'CT', 'CV', 'BOM', 'N condizioni',
                    'Esempio condizione 1', 'Esempio condizione 2', 'Esempio condizione 3'],
        rows     = traceability.get('cv_missing', []),
        col_widths = [10, 14, 16, 20, 14, 50, 50, 50],
    )

    # ── Foglio 5: CT di sistema ──────────────────────────────────────────────
    ws4 = wb.create_sheet('CT di sistema')
    _write_sheet(
        ws4,
        title    = 'CT DI SISTEMA — usati per filtraggio, non espansi in matrice',
        subtitle = 'Questi codici sono attesi e non richiedono azione.',
        headers  = ['Severita', 'CT', 'N condizioni', 'Nota'],
        rows     = traceability.get('ct_system', []),
        col_widths = [10, 18, 14, 60],
    )

    # ── Foglio 6: Prezzi (opzionale) ─────────────────────────────────────────
    if price_issues is not None:
        ws5 = wb.create_sheet('Prezzi')
        _write_sheet(
            ws5,
            title    = 'QUALITA\' PREZZI — Componenti attivi con prezzo mancante o non valido',
            subtitle = ('ERRORE = prezzo NULL, prezzo = 0, o codice non trovato in PREZZI.xlsx. '
                        'Correggere prima di eseguire l\'analisi Cycle Stock.'),
            headers  = ['Severita', 'Problema', 'Codice', 'Descrizione', 'Prezzo (€)'],
            rows     = price_issues,
            col_widths = [10, 32, 16, 40, 12],
        )

    wb.save(output_path)
    print(f"\n  [OK] Report qualita' salvato: {output_path}")


# ============================================================================
# ENRICHMENT BOM GREZZO → BOM_150VER.xlsx  (integrato da bom_enrich.py)
# ============================================================================

_DEP_COL_ENRICH = "Blocco dati per archivio dipendenze"


def _bom_to_str(v) -> str:
    """NaN/None → ''. Float interi (30.0) → '30'. Altrimenti str.strip()."""
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _find_raw_file(folder: Path, pattern: str) -> Path:
    """Cerca il primo file che matcha il glob pattern (case-insensitive)."""
    matches = list(folder.glob(pattern))
    if not matches:
        matches = [f for f in folder.iterdir()
                   if pattern.lower().replace("*", "") in f.name.lower()]
    if not matches:
        raise FileNotFoundError(f"Nessun file '{pattern}' in {folder}")
    return matches[0]


def enrich_bom_version(ver: str, in_dir, save: bool = True) -> pd.DataFrame:
    """
    Arricchisce il BOM grezzo per la versione VER in un unico step.

    Legge (da in_dir):
      - BOM150_DVER*.XLSX           (BOM grezzo SAP)
      - DATI_AGGIUNTI*_DVER*.XLSX   (Car.MRP, Tipo approvv., Approvv. speciale)
      - ZIBP_BOM_125*_DVER*.XLSX    (BOM125 con colonna dipendenze)

    Produce:
      - DataFrame arricchito con "Blocco dati per archivio dipendenze" propagato,
        filtrato su Check Generale = Mostra
      - Salva BOM_150VER.xlsx in in_dir se save=True
    """
    in_dir = Path(in_dir)
    print(f"\n  [ENRICH] Versione: {ver}")

    bom_file  = _find_raw_file(in_dir, f"BOM*150*D{ver}*.XLSX")         # BOM150_DVER*.XLSX o BOM_150DVER*.XLSX
    dag_file  = _find_raw_file(in_dir, f"DATI*AGGIUNT*D{ver}*.XLSX")   # DATI_AGGIUNTI*_DVER*.XLSX o DATI AGGIUNTIVI*DVER*.XLSX
    b125_file = _find_raw_file(in_dir, f"ZIBP*BOM*125*D{ver}*.XLSX")   # ZIBP_BOM_125*_DVER*.XLSX o ZIBP BOM 125 DVER*.XLSX
    print(f"  [ENRICH]   BOM150 : {bom_file.name}")
    print(f"  [ENRICH]   DAG    : {dag_file.name}")
    print(f"  [ENRICH]   BOM125 : {b125_file.name}")

    bom = pd.read_excel(bom_file)
    dag = pd.read_excel(dag_file)

    # — Join Dati Aggiuntivi ———————————————————————————————————————————————
    dag_sel = (dag[["Materiale", "Car. MRP", "Tipo approvv.",
                    "Approvv. speciale", "Merce sfusa"]]
               .rename(columns={"Merce sfusa": "Merce sfusa (DA)"}))
    bom = (bom.merge(dag_sel, left_on="Numero componenti",
                     right_on="Materiale", how="left")
              .drop(columns=["Materiale"]))

    sfusa_bom = [_bom_to_str(v) for v in bom["Merce sfusa"]]
    bom = bom.drop(columns=["Merce sfusa"])
    bom.insert(0, "ID", range(1, len(bom) + 1))
    bom["Livello numerico"] = bom["Liv. esplosione"].apply(
        lambda v: str(v).count(".") if pd.notna(v) else 0
    )

    # — Formule H-L (Check Generale) ————————————————————————————————————————
    n      = len(bom)
    H      = [0] * n;  I_chk = [""] * n
    J      = [0] * n;  K_chk = [""] * n;  L_chk = [""] * n
    lev    = bom["Livello numerico"].tolist()
    tipo   = [_bom_to_str(v) for v in bom["Tipo approvv."]]
    approv = [_bom_to_str(v) for v in bom["Approvv. speciale"]]
    numpos = ([_bom_to_str(v) for v in bom["Numero posizione"]]
              if "Numero posizione" in bom.columns else [""] * n)

    for i in range(n):
        G_i = lev[i];  C_i = tipo[i];  D_i = approv[i]
        O_i = sfusa_bom[i];  R_i = numpos[i]
        Hp  = H[i - 1] if i > 0 else 0
        Jp  = J[i - 1] if i > 0 else 0

        if C_i == "F" and D_i != "30" and O_i != "X" and (Hp >= G_i or Hp == 0):
            H[i] = G_i
        elif G_i == Hp and C_i == "F" and D_i != "30" and R_i != "X":
            H[i] = 0
        elif G_i <= Hp:
            H[i] = 0
        else:
            H[i] = Hp

        if C_i == "" and G_i == 2 and O_i == "":
            I_chk[i] = "Mostra"
        elif C_i in ("E", "") and D_i != "30" and O_i != "X":
            I_chk[i] = "Nascondi"
        elif C_i == "F" and D_i != "30" and O_i != "X" and (H[i] == 0 or H[i] == G_i):
            I_chk[i] = "Mostra"
        elif Hp > 0 and G_i > Hp:
            I_chk[i] = "Nascondi"
        else:
            I_chk[i] = "Mostra"

        J[i] = G_i if O_i == "X" else (Jp if Jp > 0 and G_i > Jp else 0)
        K_chk[i] = "Nascondi" if J[i] > 0 else "Mostra"
        if G_i == 1:
            L_chk[i] = "Mostra"
        elif K_chk[i] == "Nascondi" or I_chk[i] == "Nascondi":
            L_chk[i] = "Nascondi"
        else:
            L_chk[i] = "Mostra"

    bom["Livello Segnalibro Merce Non Sfusa"] = H
    bom["Check Merce non sfusa"]              = I_chk
    bom["Livello Segnalibro Merce Sfusa"]     = J
    bom["Check Merce sfusa"]                  = K_chk
    bom["Check Generale"]                     = L_chk

    # — BOM125: aggrega colonna dipendenze per componente ———————————————————
    bom125 = pd.read_excel(b125_file, header=0)
    bom125_dep: Dict[str, str] = {}
    for _, row in bom125.iterrows():
        comp    = _bom_to_str(row["Componente critico"])
        dep_raw = _bom_to_str(row.get(_DEP_COL_ENRICH, ""))
        dep_val = re.sub(r'\s*\|\s*', ' OR ', dep_raw).strip() if dep_raw else ""
        if not comp:
            continue
        if comp not in bom125_dep:
            bom125_dep[comp] = dep_val
        elif dep_val and not bom125_dep[comp]:
            bom125_dep[comp] = dep_val
        elif dep_val:
            bom125_dep[comp] += " OR " + dep_val

    bom125_df = pd.DataFrame(list(bom125_dep.items()),
                             columns=["Componente critico", _DEP_COL_ENRICH])

    # — Join dipendenze su tutte le righe (prima del filtro) ————————————————
    bom = bom.merge(bom125_df, left_on="Numero componenti",
                    right_on="Componente critico", how="left")
    bom = bom.drop(columns=["Componente critico"])
    dep_vals = bom.pop(_DEP_COL_ENRICH)
    bom.insert(3, _DEP_COL_ENRICH, dep_vals)
    bom[_DEP_COL_ENRICH] = bom[_DEP_COL_ENRICH].fillna("")

    # — Propagazione dipendenze (prima del filtro) ——————————————————————————
    dep_arr    = bom[_DEP_COL_ENRICH].tolist()
    levels     = bom["Livello numerico"].tolist()
    anchor_dep = "";  anchor_set = False;  propagated = 0
    for i in range(len(bom)):
        lvl = levels[i];  row_dep = dep_arr[i]
        if lvl == 2:
            anchor_dep = row_dep if row_dep else ""; anchor_set = True
        elif lvl == 1:
            pass
        elif anchor_set and not row_dep and anchor_dep:
            dep_arr[i] = anchor_dep; propagated += 1
    bom[_DEP_COL_ENRICH] = dep_arr
    print(f"  [ENRICH]   Dipendenze propagate: {propagated}")

    # — Filtro Check Generale = Mostra ——————————————————————————————————————
    bom = bom[bom["Check Generale"] == "Mostra"].copy()
    bom["ID"] = range(1, len(bom) + 1)
    print(f"  [ENRICH]   Righe dopo filtro   : {len(bom)}")

    # — Salvataggio ————————————————————————————————————————————————————————
    if save:
        out_file = in_dir / f"BOM_150{ver}.xlsx"
        bom.to_excel(out_file, index=False)
        print(f"  [ENRICH]   Salvato             : {out_file.name}")

    return bom


# ============================================================================
# FUNZIONE PRINCIPALE: ELABORAZIONE BOM -> MATRICE (V2)
# ============================================================================

def process_excel_to_matrix_v2(excel_path: str = None,
                                mapping_file: str = '.\\Input\\V21E\\Mappatura_V21E.xlsx',
                                tr_file: str = '.\\Input\\TR TOTALV21E - Copia.xlsx',
                                data_sheet: str = None,
                                output_file: str = '.\\Input\\matrix_output_V2.xlsx',
                                states_file: str = '.\\Input\\States.xlsx',
                                residual_file: str = '.\\Input\\quantity e residual.xlsx',
                                prezzi_file: str = None,
                                filter_untraced: bool = True,
                                df_input: pd.DataFrame = None,
                                bom_file: str = ''):
    """
    Funzione principale V2: converte BOM150 in matrice espansa senza hardcoding.

    DIFFERENZE RISPETTO A V1:
    -------------------------
    - calculated_cols derivato dal TR (non hardcoded)
    - rider_seat_order e pillion_seat_order derivati dal TR
    - Pattern negativi gestiti dal Bundle "Not XYZ" in Schema
    - Logica bundle unificata (no if/elif speciali)
    - Sostituzione RIMS configurabile via foglio Excel

    Args:
        excel_path:    Path al file BOM150 (opzionale se df_input è fornito)
        df_input:      DataFrame BOM già arricchito (se fornito, salta la lettura da file).
                       Prodotto da enrich_bom_version() per lo step integrato.
        mapping_file:  Path al file Mappatura
        tr_file:       Path al file TR
        data_sheet:    Nome foglio BOM150 (solo se excel_path punta a un xlsm legacy)
        output_file:   Path file output
        states_file:   Path al file States (filtra righe market-specific)
        residual_file: Path al file quantity e residual
    """
    print(f"\n=== GESTIONE DIPENDENZE V2 — CARICAMENTO ===")
    print(f"BOM:      {excel_path if df_input is None else '(DataFrame in memoria)'}")
    print(f"Mapping:  {mapping_file}")
    print(f"TR:       {tr_file}")
    print(f"Output:   {output_file}")
    print(f"Residual: {residual_file}")

    # -------------------------------------------------------------------------
    # FASE 1: Caricamento dati
    # -------------------------------------------------------------------------
    mapping_df = load_mapping_from_file(mapping_file)
    states_set = _load_states_set(states_file)  # es. {'CDN', 'ARG', 'AUS', ...}
    active_cv_sets = _build_active_cv_sets(tr_file, mapping_df)  # CV con TR > 0
    if df_input is not None:
        # Path integrato: DataFrame già arricchito passato direttamente
        df = df_input.copy()
        df = df[df['Blocco dati per archivio dipendenze'].notna() &
                (df['Blocco dati per archivio dipendenze'] != "")]
    elif data_sheet:
        df = pd.read_excel(excel_path, sheet_name=data_sheet, header=2)
        df = df[df['Blocco dati per archivio dipendenze'].notnull()]
    else:
        df = pd.read_excel(excel_path, header=0)
        df = df[df['Blocco dati per archivio dipendenze'].notnull()]

    start_col = 'Blocco dati per archivio dipendenze'
    if start_col not in df.columns:
        raise ValueError(f"Colonna '{start_col}' non trovata nel BOM150")

    # ── Residual: solo quality check, NON filtra le righe BOM ───────────────
    # La fonte primaria delle SKU da elaborare sono le tre BOM (V21E/V22E/V24T).
    # Il file quantity e residual.xlsx viene usato esclusivamente per:
    #   1. Quality check (componenti con residual=0 / null / mancanti)
    #   2. Arricchimento lead time nel matching SKU (sku_matching_V2.py)
    # NON esclude più componenti dalla matrice.
    _RESIDUAL_COL = ('Residual Purchasing Time (months), '
                     'quanto rimane di mesi rispetto al frozen period')
    _COMPONENT_COL_BOM = 'Numero componenti'
    residual_issues:    List[Dict] = []
    price_issues:       List[Dict] = []
    bom_missing_issues: List[Dict] = []
    try:
        res_df = pd.read_excel(residual_file)
        if 'Codice componenti' in res_df.columns:
            res_df = res_df.rename(columns={'Codice componenti': _COMPONENT_COL_BOM})
        res_df[_RESIDUAL_COL] = pd.to_numeric(res_df[_RESIDUAL_COL], errors='coerce')

        # Quality check residual (solo report, non filtra)
        residual_issues = _check_residual_quality(res_df, _RESIDUAL_COL, _COMPONENT_COL_BOM)

        # Quality check prezzi (opzionale)
        if prezzi_file:
            try:
                price_issues = _check_price_quality(
                    res_df, _RESIDUAL_COL, _COMPONENT_COL_BOM, prezzi_file
                )
            except Exception as _pe:
                print(f"  [WARN] Verifica prezzi non eseguita: {_pe}")

        # Info riepilogativa + costruzione lista BOM senza residual
        bom_codes = set(df[_COMPONENT_COL_BOM].dropna()) if _COMPONENT_COL_BOM in df.columns else set()
        res_codes = set(res_df[_COMPONENT_COL_BOM].dropna())
        missing_in_res = bom_codes - res_codes
        print(f"  [Residual] Componenti BOM: {len(bom_codes)} | "
              f"Presenti in residual: {len(bom_codes & res_codes)} | "
              f"Assenti da residual (inclusi con LT=0): {len(missing_in_res)}")

        # Costruisci lista dettagliata per il foglio "BOM Senza Residual"
        _desc_col_bom = next((c for c in df.columns
                              if 'testo' in c.lower() or 'descr' in c.lower()
                              or 'text' in c.lower() or 'benennung' in c.lower()), None)
        _dep_col_bom  = 'Blocco dati per archivio dipendenze'
        bom_missing_issues = []
        if missing_in_res and _COMPONENT_COL_BOM in df.columns:
            _sub = df[df[_COMPONENT_COL_BOM].isin(missing_in_res)].drop_duplicates(
                subset=[_COMPONENT_COL_BOM])
            for _, _row in _sub.iterrows():
                bom_missing_issues.append({
                    'Numero componenti': str(_row.get(_COMPONENT_COL_BOM, '')),
                    'Descrizione':       str(_row[_desc_col_bom]) if _desc_col_bom else '',
                    'Blocco dipendenze': str(_row.get(_dep_col_bom, '')) if _dep_col_bom in df.columns else '',
                })
    except Exception as e:
        print(f"  [INFO] File residual non caricato ({e}) — tutti i componenti BOM inclusi")
        bom_missing_issues = []
    # ─────────────────────────────────────────────────────────────────────────

    bom_data = df.copy()

    # Parser iniziale (senza filtro) usato per la verifica tracciabilità
    parser = ConditionParser(mapping_df, active_cv_sets)

    # ── Verifica tracciabilità CT/CV ──────────────────────────────────────────
    traceability = _check_ct_cv_traceability(bom_data, start_col, parser, bom_file=bom_file)
    # ─────────────────────────────────────────────────────────────────────────

    # ── Genera report qualità Excel ───────────────────────────────────────────
    try:
        _write_quality_report(residual_issues, traceability, price_issues,
                              bom_missing=bom_missing_issues)
    except Exception as _qe:
        print(f"  [WARN] Impossibile generare quality_check_report.xlsx: {_qe}")

    # ── Ri-crea parser con filtro CT/CV non tracciati (se abilitato) ─────────
    if filter_untraced:
        _unk_ct = traceability.get('unknown_ct_set', set())
        _unk_cv = traceability.get('unknown_cv_set', set())
        if _unk_ct or _unk_cv:
            print(f"  [Filtro] filter_untraced=True: "
                  f"{len(_unk_ct)} CT e {len(_unk_cv)} coppie CT/CV non tracciati "
                  f"verranno esclusi dalle condizioni BOM.")
            parser = ConditionParser(mapping_df, active_cv_sets,
                                     filter_untraced=True,
                                     unknown_ct=_unk_ct,
                                     unknown_cv=_unk_cv)
    # ─────────────────────────────────────────────────────────────────────────
    # ─────────────────────────────────────────────────────────────────────────

    # -------------------------------------------------------------------------
    # FASE 2: Caricamento dinamico colonne calcolate e ordini di priorità dal TR
    # -------------------------------------------------------------------------
    print(f"\n=== CARICAMENTO DINAMICO DAL TR ===")

    # Colonne calcolate derivate dal TR (sostituisce la lista hardcoded V1)
    calculated_cols = load_calculated_cols_from_tr(tr_file)

    # Set dei nomi bundle positivi (usato nella logica di normalizzazione)
    bundle_set = {c[2] for c in calculated_cols if c[3] is None}

    # Ordini di priorità derivati dal TR.
    # SOLO per caratteristiche dove un singolo componente BOM può avere più opzioni
    # e serve scegliere la più "importante" (es: RIDER SEAT, PILLION SEAT).
    # Per le altre caratteristiche fisiche (FAIRING COLOR, RIMS, ecc.) non serve
    # l'ordinamento: si prende il primo valore trovato nella riga, come in V1.
    # L'ordine del TR viene usato come priority_order SOLO se la caratteristica
    # ha effettivamente più valori possibili per lo stesso componente BOM.
    priority_orders = {}
    # Caratteristiche che usano esplicitamente la priorità (come rider_seat_order in V1)
    priority_chars = {'RIDER SEAT', 'PILLION SEAT'}
    for calc_name, search_key in [(c[2], c[3]) for c in calculated_cols]:
        if search_key is not None and search_key in priority_chars:
            order = load_priority_order_from_tr(tr_file, search_key)
            if order:
                priority_orders[calc_name] = order

    # Sostituzioni RIMS (da foglio Excel opzionale)
    rims_substitutions = load_rims_substitutions(mapping_file)

    # Pre-calcola mappa CT/CV -> bundle(s) dallo Schema
    # Usata per la logica split/discriminante sui CT/CV condivisi
    cv_bundle_map = build_shared_cv_map(mapping_df)

    # -------------------------------------------------------------------------
    # FASE 3: Costruzione struttura colonne
    # -------------------------------------------------------------------------
    col_structure = [
        ('', '', 'Numero componenti'),
        ('', '', 'Condizione_Originale'),
        ('', '', 'Condizione_Espansa'),
        (start_col, '', '')
    ]

    for row in mapping_df.itertuples(index=False):
        col_tuple = (row.Bundle, row.Descrizione, row.Readable)
        col_structure.append(col_tuple)

    print(f"[OK] Colonne dopo mapping: {len(col_structure)}")

    # Aggiungi colonne calcolate
    for col_tuple in calculated_cols:
        col_structure.append((col_tuple[0], col_tuple[1], col_tuple[2]))

    print(f"[OK] Colonne totali (incl. calcolate): {len(col_structure)}")

    # -------------------------------------------------------------------------
    # FASE 4: Pre-calcolo indici (O(1) lookup)
    # -------------------------------------------------------------------------
    calc_indices = {}
    mapping_cols_count = len(col_structure) - len(calculated_cols) - 4

    for calc_name, search_key in [(c[2], c[3]) for c in calculated_cols]:
        indices = []

        if search_key is not None and calc_name in priority_orders:
            # Caratteristica con ordine di priorità (es: RIDER SEAT, PILLION SEAT)
            # Costruisce indices nell'ordine di priorità definito nel TR
            # CORREZIONE V2: confronto case-insensitive per gestire divergenze TR/Schema
            for readable_name in priority_orders[calc_name]:
                for idx in range(4, mapping_cols_count + 4):
                    b, d, r = col_structure[idx]
                    r_norm = _remove_parentheses_variant(r)
                    if (r.lower() == readable_name.lower() or
                            r_norm.lower() == readable_name.lower()):
                        indices.append(idx)

        elif search_key is None:
            # Bundle: cerca per bundle name (inclusi "Not NomePack")
            #
            # STRATEGIA DI MATCH (in ordine di priorità):
            # 1. Exact match case-insensitive — copre la maggior parte dei casi
            # 2. Match con rimozione parentesi — gestisce "Touring pack (color variant)" -> "Touring Pack"
            #
            # NOTA IMPORTANTE — perché NON si usa il match parziale a N parole:
            # Il match alle prime 2 parole causava falsi positivi gravi:
            #   • "Touring pack Rally" -> ["touring","pack"] == "Touring Pack" -> SBAGLIATO
            #   • "Enduro pack - Black" -> ["enduro","pack"] == "Enduro pack - Silver" -> SBAGLIATO
            # La rimozione delle parentesi è sicura: "Touring pack (color variant)" diventa
            # "Touring pack" che corrisponde esattamente a "Touring Pack" (case-insensitive),
            # mentre "Touring pack Rally" rimane "Touring pack Rally" -> nessun match spurio.
            not_name = _get_not_bundle_name(calc_name)
            calc_name_lower = calc_name.lower()
            not_name_lower = not_name.lower()
            # Versioni senza parentesi (es: "Touring pack (color variant)" -> "Touring pack")
            calc_name_stripped = re.sub(r'\s*\([^)]*\)', '', calc_name).strip().lower()
            not_name_stripped  = re.sub(r'\s*\([^)]*\)', '', not_name).strip().lower()

            for idx in range(4, mapping_cols_count + 4):
                b, d, r = col_structure[idx]
                b_lower = b.lower()

                # 1. Exact match case-insensitive
                if b_lower == calc_name_lower or b_lower == not_name_lower:
                    indices.append(idx)
                    continue

                # 2. Match con parentesi rimosse dallo Schema e/o dal TR
                b_stripped = re.sub(r'\s*\([^)]*\)', '', b).strip().lower()
                if b_stripped == calc_name_stripped or b_stripped == not_name_stripped:
                    indices.append(idx)

            if not indices:
                print(f"[V2] INFO: bundle '{calc_name}' non trovato in Schema — colonna sempre 'No'")

        else:
            # Caratteristica senza ordine specifico: cerca per descrizione
            # CORREZIONE V2: confronto case-insensitive su desc
            # L'ordine degli indici rispecchia l'ordine del mapping (identico a V1)
            for idx in range(4, mapping_cols_count + 4):
                b, d, r = col_structure[idx]
                if d.lower() == search_key.lower():
                    indices.append(idx)

        calc_indices[calc_name] = indices

    col_index = _build_column_index(col_structure)

    # -------------------------------------------------------------------------
    # FASE 5: Espansione condizioni e popolamento matrice
    # -------------------------------------------------------------------------
    print(f"\n=== ELABORAZIONE ===")
    print(f"Colonne totali: {len(col_structure)}")
    print(f"Righe BOM150:   {len(bom_data)}")

    def expand_all_conditions(condition_text, original_condition):
        """
        Espande le condizioni OR e IN in righe separate.
        Gestisce anche lo split per CT/CV condivisi tra più bundle.
        """
        if not condition_text or pd.isna(condition_text):
            return [(condition_text, original_condition)]
        or_parts = parser.split_or_conditions(condition_text)
        expanded_conditions = []
        for or_part in or_parts:
            if not or_part or pd.isna(or_part):
                expanded_conditions.append((or_part, original_condition))
                continue
            # ── Priorità 1: NOT ($root.VAR IN (lista)) ───────────────────────
            # Semantica: VAR può essere qualsiasi CV del CT NON presente nella
            # lista -> espandi in una riga eq per ogni valore NON escluso.
            # Es: NOT ($root.CT000478 IN ('CV000517','CV000183',...))
            #     + Mappatura ha CV000001, CV000002, CV000517 per CT000478
            #     -> genera: $root.CT000478 eq 'CV000001'
            #               $root.CT000478 eq 'CV000002'
            #     (CV000517 è nella lista -> escluso)
            not_in_match = re.search(
                r'NOT\s*\(\s*\$root\.(\w+)\s+IN\s*\(([^)]+)\)\s*\)',
                or_part, re.IGNORECASE
            )
            if not_in_match:
                var, values_str = not_in_match.groups()
                excluded = {
                    v.strip().upper()
                    for v in re.findall(r"['\"]([^'\"]+)['\"]", values_str)
                }
                all_values = parser.get_all_values_for_char(var)
                expanded_any = False
                for val in all_values:
                    if val.strip().upper() not in excluded:
                        new_cond = (
                            or_part[:not_in_match.start()]
                            + f"$root.{var} eq '{val}'"
                            + or_part[not_in_match.end():]
                        )
                        expanded_conditions.extend(
                            expand_all_conditions(new_cond, original_condition)
                        )
                        expanded_any = True
                if not expanded_any:
                    print(f"[WARN] NOT IN: nessun valore trovato per '{var}' "
                          f"dopo esclusione di {len(excluded)} valori — riga ignorata")

            else:
                # ── Priorità 2: IN normale ────────────────────────────────────
                in_matches = list(re.finditer(r'\$root\.(\w+)\s+IN\s*\(([^)]+)\)', or_part, re.IGNORECASE))
                if in_matches:
                    # Espansione IN: genera un eq per ogni valore
                    first_match = in_matches[0]
                    var, values_str = first_match.groups()
                    values = re.findall(r"['\"]([^'\"]+)['\"]", values_str)
                    for value in values:
                        new_condition = or_part[:first_match.start()] + f"$root.{var} eq '{value}'" + or_part[first_match.end():]
                        recursive_expansions = expand_all_conditions(new_condition, original_condition)
                        expanded_conditions.extend(recursive_expansions)
                else:
                    # ── Priorità 3: ne su ZPRODGRP ───────────────────────────
                    # $root.ZPRODGRPx ne 'VAL' -> una riga eq per ogni modello in
                    # mappatura per ZPRODGRPx, escluso VAL.
                    ne_zprodgrp = re.search(
                        r'\$root\.(ZPRODGRP\w*)\s+ne\s+["\']([^"\']+)["\']',
                        or_part, re.IGNORECASE
                    )
                    if ne_zprodgrp:
                        var, excluded = ne_zprodgrp.groups()
                        all_values = parser.get_all_values_for_char(var)
                        expanded_any = False
                        for model_value in all_values:
                            if model_value.strip().upper() != excluded.strip().upper():
                                new_cond = (
                                    or_part[:ne_zprodgrp.start()]
                                    + f"$root.{var} eq '{model_value}'"
                                    + or_part[ne_zprodgrp.end():]
                                )
                                expanded_conditions.extend(
                                    expand_all_conditions(new_cond, original_condition)
                                )
                                expanded_any = True
                        if not expanded_any:
                            print(f"[WARN] ne ZPRODGRP: nessun modello trovato per '{var}' "
                                  f"dopo esclusione di '{excluded}' — riga ignorata")
                    else:
                        expanded_conditions.append((or_part, original_condition))
        return expanded_conditions

    def resolve_bundle_split(condition_text):
        """
        Determina come gestire i CT/CV condivisi tra più bundle in una condizione.

        LOGICA CORRETTA:
        ================
        1. Estrae tutti i (CHAR, VALUE) dalla condizione (solo op eq, non NOT)
        2. Per ogni CT/CV, recupera i bundle associati dallo Schema (solo positivi)
        3. Identifica i CT/CV CONDIVISI: quelli che mappano su 2+ bundle
           -> Questi sono i soli responsabili di un vero conflitto tra bundle
        4. Se NON ci sono CT/CV condivisi -> return None
           (bundle diversi da CT/CV esclusivi sono bundle INDIPENDENTI, non in conflitto)
        5. competing_bundles = unione dei bundle dai soli CT/CV condivisi
        6. Cerca discriminante: qualsiasi CT/CV (condiviso o esclusivo) che mappa
           su ESATTAMENTE 1 bundle tra competing_bundles
        7. Se discriminante trovato -> ('discriminant', winner, competing_bundles)
        8. Se nessun discriminante -> ('split', competing_bundles)

        ESEMPIO CORRETTO:
        -----------------
        Condizione: CT000418=TRUE (Touring+Adventure condiviso) AND CT001979=TRUE (Radar esclusivo)
        -> shared_cvs = [(CT000418, TRUE, [Touring, Adventure])]
        -> competing_bundles = {Touring, Adventure}
        -> CT001979 mappa su [Radar], Radar NON è in competing_bundles -> non è discriminante
        -> Nessun discriminante -> ('split', [Adventure, Touring])
        -> Radar Pack viene attivato INDIPENDENTEMENTE in entrambe le split rows ✓

        CASO SBAGLIATO (bug precedente):
        ---------------------------------
        Condizione: CT000076=CV000004 (Tech Pack esclusivo) AND CT001979=TRUE (Radar esclusivo)
        -> shared_cvs = [] -> return None ✓ -> Tech Pack e Radar Pack attivati entrambi

        Returns:
            None                                -> nessun conflitto, path normale
            ('discriminant', winner, candidates) -> un bundle vince tra i competitori
            ('split', [b1, b2, ...])            -> split necessario tra bundle competitori
        """
        if not condition_text or pd.isna(condition_text):
            return None

        # Estrai tutti i (CHAR, VALUE) con op eq dalla condizione
        # Ignora NOT (...) perché non discriminano positivamente
        eq_pairs = re.findall(
            r'(?<!NOT\s)(?<!\()\$root\.(\w+)\s+eq\s+["\']([^"\']+)["\']',
            condition_text, re.IGNORECASE
        )

        if not eq_pairs:
            return None

        # Per ogni CT/CV, recupera i bundle positivi dalla mappa
        all_bundles_per_cv = []
        for char, value in eq_pairs:
            bundles = cv_bundle_map.get((char, value), [])
            if bundles:
                all_bundles_per_cv.append((char, value, bundles))

        if not all_bundles_per_cv:
            return None

        # STEP CRITICO: identifica i soli CT/CV CONDIVISI (2+ bundle)
        # Solo questi creano un vero conflitto — bundle da CT/CV esclusivi
        # sono bundle INDIPENDENTI che devono coesistere, non competere.
        shared_cvs = [(char, value, bundles)
                      for char, value, bundles in all_bundles_per_cv
                      if len(bundles) > 1]

        if not shared_cvs:
            return None  # Tutti CT/CV esclusivi -> bundle indipendenti, nessun conflitto

        # competing_bundles: solo bundle dai CT/CV condivisi
        competing_bundles = set()
        for _, _, bundles in shared_cvs:
            competing_bundles.update(bundles)

        # Cerca discriminante: qualsiasi CT/CV che mappa su ESATTAMENTE 1 bundle
        # tra quelli competitori (può essere shared o esclusivo)
        for char, value, bundles in all_bundles_per_cv:
            bundles_in_competing = [b for b in bundles if b in competing_bundles]
            if len(bundles_in_competing) == 1:
                # Discriminante trovato: identifica univocamente un bundle competitore
                winner = bundles_in_competing[0]
                return ('discriminant', winner, competing_bundles)

        # Nessun discriminante trovato: tutti i CT/CV condivisi non discriminano
        # -> split necessario su tutti i bundle competitori
        return ('split', sorted(competing_bundles))

    matrix_rows = []

    for bom_row in bom_data.itertuples():
        bom_index = bom_row.Index
        condition = getattr(bom_row, start_col.replace(' ', '_').replace('/', '_'), None)
        if condition is None:
            condition = bom_data.loc[bom_index, start_col]

        numero_componenti = getattr(bom_row, 'Numero_componenti', '') if hasattr(bom_row, 'Numero_componenti') else ''
        if not numero_componenti:
            numero_componenti = bom_data.loc[bom_index, 'Numero componenti'] if 'Numero componenti' in bom_data.columns else ''

        if not condition or pd.isna(condition):
            row = ['' for _ in col_structure]
            row[0] = numero_componenti
            row[1] = condition if condition else ''
            row[2] = condition if condition else ''
            row[3] = condition if condition else ''
            matrix_rows.append(row)
        else:
            all_expanded_conditions = expand_all_conditions(condition, condition)
            for expanded_cond, original_cond in all_expanded_conditions:

                # ── Filtro stati market-specific ─────────────────────────────
                # Scarta le righe la cui condizione espansa contiene
                # $root.ZCOUNTRY eq 'STATE' per uno stato in States.xlsx.
                # Es: "$root.ZCOUNTRY eq 'CDN'" -> riga scartata (solo Canada).
                # "$root.ZCOUNTRY ne 'CDN'" -> riga mantenuta (tutti tranne CDN).
                if _is_state_specific(expanded_cond, states_set):
                    continue
                # ─────────────────────────────────────────────────────────────

                # Verifica se serve uno split per CT/CV condivisi
                split_result = resolve_bundle_split(expanded_cond)

                if split_result is None:
                    # -------------------------------------------------------
                    # CASO NORMALE: nessun conflitto tra bundle
                    # -------------------------------------------------------
                    row = ['' for _ in col_structure]
                    row[0] = numero_componenti
                    row[1] = original_cond
                    row[2] = expanded_cond
                    row[3] = expanded_cond

                    included = _populate_row_from_condition_v2(
                        row, expanded_cond, parser, col_structure,
                        calc_indices, calculated_cols, col_index, bundle_set
                    )
                    if included is not False:
                        matrix_rows.append(row)

                elif split_result[0] == 'split':
                    # -------------------------------------------------------
                    # CASO SPLIT: tutti i CT/CV condivisi, nessun discriminante
                    # Crea una riga per ogni bundle candidato
                    # Ogni riga: bundle target attivo, tutti gli altri azzerati
                    # -------------------------------------------------------
                    _, split_bundles = split_result
                    calc_offset = len(calculated_cols)
                    split_bundles_lower = {b.lower() for b in split_bundles}

                    for target_bundle in split_bundles:
                        row = ['' for _ in col_structure]
                        row[0] = numero_componenti
                        row[1] = original_cond
                        row[2] = expanded_cond
                        row[3] = expanded_cond

                        included = _populate_row_from_condition_v2(
                            row, expanded_cond, parser, col_structure,
                            calc_indices, calculated_cols, col_index, bundle_set
                        )
                        if included is False:
                            continue

                        # Azzera i bundle candidati NON target
                        target_bundle_lower = target_bundle.lower()
                        for i, col_tuple in enumerate(calculated_cols):
                            calc_name = col_tuple[2]
                            search_key = col_tuple[3]
                            if search_key is not None:
                                continue  # caratteristica fisica, non bundle
                            calc_name_lower = calc_name.lower()
                            if calc_name_lower in split_bundles_lower and calc_name_lower != target_bundle_lower:
                                row[-(calc_offset - i)] = 'No'

                        matrix_rows.append(row)

                elif split_result[0] == 'discriminant':
                    # -------------------------------------------------------
                    # CASO DISCRIMINANTE: esiste un CT/CV che identifica univocamente
                    # il bundle corretto -> una sola riga, attiva solo il winner
                    # e azzera tutti gli altri candidati
                    # -------------------------------------------------------
                    _, winner, candidates = split_result
                    calc_offset = len(calculated_cols)
                    winner_lower = winner.lower()
                    candidates_lower = {c.lower() for c in candidates}

                    row = ['' for _ in col_structure]
                    row[0] = numero_componenti
                    row[1] = original_cond
                    row[2] = expanded_cond
                    row[3] = expanded_cond

                    included = _populate_row_from_condition_v2(
                        row, expanded_cond, parser, col_structure,
                        calc_indices, calculated_cols, col_index, bundle_set
                    )
                    if included is False:
                        continue

                    # Azzera i candidati che non sono il winner
                    for i, col_tuple in enumerate(calculated_cols):
                        calc_name = col_tuple[2]
                        search_key = col_tuple[3]
                        if search_key is not None:
                            continue  # caratteristica fisica, non bundle
                        calc_name_lower = calc_name.lower()
                        if calc_name_lower in candidates_lower and calc_name_lower != winner_lower:
                            row[-(calc_offset - i)] = 'No'

                    matrix_rows.append(row)

    print(f"[OK] Righe totali matrice: {len(matrix_rows)}")

    # -------------------------------------------------------------------------
    # FASE 6: Post-processing RIMS (configurabile da foglio Excel)
    # -------------------------------------------------------------------------
    if rims_substitutions:
        # Trova indici colonne coinvolte
        col_names = [c[2] if c[2] else c[0] for c in col_structure]
        rims_idx = None
        model_indices = {}

        for idx, readable in enumerate(col_names):
            if readable == 'Rims' or readable == 'RIMS':
                rims_idx = idx
            for base_val, subs in rims_substitutions.items():
                for model_code, _ in subs:
                    if readable == model_code:
                        model_indices[model_code] = idx

        if rims_idx is not None and model_indices:
            substituted_count = 0
            for row in matrix_rows:
                current_rims = row[rims_idx]
                if current_rims in rims_substitutions:
                    for model_code, substitution in rims_substitutions[current_rims]:
                        if model_code in model_indices and row[model_indices[model_code]]:
                            row[rims_idx] = substitution
                            substituted_count += 1
                            break
            print(f"[OK] Sostituzione Rims: {substituted_count} righe modificate")
    else:
        # Fallback: comportamento identico a V1 se il foglio RIMS_Substitutions non esiste
        col_names = [c[2] if c[2] else c[0] for c in col_structure]
        rims_idx = None
        msv4pp_idx = None
        msv4s_idx = None
        for idx, readable in enumerate(col_names):
            if readable == 'Rims': rims_idx = idx
            elif readable == 'MSV4PP': msv4pp_idx = idx
            elif readable == 'MSV4S': msv4s_idx = idx

        if all(idx is not None for idx in [rims_idx, msv4pp_idx, msv4s_idx]):
            substituted_count = 0
            for row in matrix_rows:
                if row[rims_idx] == 'Forged':
                    if row[msv4pp_idx]:
                        row[rims_idx] = 'Forged(17"/17") - Pikes Peak'
                        substituted_count += 1
                    elif row[msv4s_idx]:
                        row[rims_idx] = 'Forged(19"/17") - Black'
                        substituted_count += 1
            print(f"[OK] Sostituzione Rims (fallback V1): {substituted_count} righe modificate")

    # -------------------------------------------------------------------------
    # FASE 6.5: Allineamento valori caratteristiche con nomi opzione TR
    # -------------------------------------------------------------------------
    # Problema: la Mappatura usa nomi come "Viola Blast", il TR usa
    # "Viola Blast gloss". Questo causa mismatch nel matching perché
    # il simulatore genera valori basati sui nomi opzione esatti del TR.
    # Soluzione: per ogni cella caratteristica nella matrice, cercare il
    # nome opzione TR che corrisponde (exact o prefix-match non ambiguo)
    # e sostituire il valore con il nome esatto del TR.
    # -------------------------------------------------------------------------
    tr_char_options = _build_tr_char_options(tr_file)
    if tr_char_options:
        calc_offset = len(calculated_cols)
        corrected_count = 0
        for row in matrix_rows:
            for i, col_tuple in enumerate(calculated_cols):
                calc_name = col_tuple[2]   # es. "FAIRING COLOR"
                search_key = col_tuple[3]  # None = bundle, str = caratteristica
                if search_key is None:
                    continue  # Salta i bundle, processa solo le caratteristiche
                col_pos = -(calc_offset - i)
                current_value = row[col_pos]
                if not current_value or current_value == 'No':
                    continue
                aligned = _align_to_tr_option(current_value, calc_name, tr_char_options)
                if aligned != current_value:
                    row[col_pos] = aligned
                    corrected_count += 1
        print(f"[OK] Allineamento TR: {corrected_count} valori corretti")

    # -------------------------------------------------------------------------
    # FASE 7: Salvataggio matrice in Excel
    # -------------------------------------------------------------------------
    print(f"\n=== SALVATAGGIO ===")

    with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
        pd.DataFrame(matrix_rows).to_excel(writer, sheet_name='Matrix', index=False, header=False, startrow=3)
        ws = writer.sheets['Matrix']

        for col_idx, (bundle, desc, readable) in enumerate(col_structure, start=1):
            ws.cell(1, col_idx, bundle if bundle else '')
            ws.cell(2, col_idx, desc if desc else '')
            ws.cell(3, col_idx, readable if readable else col_structure[col_idx-1][0])

        last_col = get_column_letter(len(col_structure))
        tab = Table(displayName="MatrixTable_V2", ref=f"A3:{last_col}{len(matrix_rows)+3}")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tab)

    print(f"\n=== RISULTATO V2 ===")
    print(f"Salvato: {output_file}")
    print(f"Dimensioni: {len(matrix_rows)} righe × {len(col_structure)} colonne")

    return matrix_rows


# ============================================================================
# ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    excel_path = r'.\Input\V21E\BOM_150V21E.xlsx'

    process_excel_to_matrix_v2(
        excel_path=excel_path,
        mapping_file='.\\Input\\V21E\\Mappatura_V21E.xlsx',
        tr_file='.\\Input\\TR TOTALV21E - Copia.xlsx',
        output_file='.\\Input\\matrix_output_V2.xlsx'
    )
