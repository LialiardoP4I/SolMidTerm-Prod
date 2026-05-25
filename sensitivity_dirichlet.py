"""
sensitivity_dirichlet.py
========================
Analisi di sensibilità sui moltiplicatori dei parametri Dirichlet.

Genera un file Excel esplicativo che mostra, per diversi valori di TR e
per ciascun livello di affidabilità (NULLA / BASSA / MEDIA / ALTA), come
varia l'incertezza nella simulazione Monte Carlo.

Output: Output/sensitivity_dirichlet.xlsx
"""

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
import warnings, os
# Restringi soppressione warning a categorie note (FutureWarning/Deprecation)
# per non nascondere bug futuri.
warnings.filterwarnings('ignore', category=FutureWarning)
warnings.filterwarnings('ignore', category=DeprecationWarning)

# ─────────────────────────────────────────────────────────────────────────────
# PARAMETRI CONFIGURABILI
# ─────────────────────────────────────────────────────────────────────────────

# Valori di Take Rate da analizzare (rappresentativi di vari componenti)
TR_VALUES = [0.01, 0.02, 0.05, 0.10, 0.15, 0.20, 0.25, 0.30, 0.40, 0.50, 0.60, 0.70, 0.80, 0.90, 0.95]

# Livelli di confidenza TR (VSS = Virtual Sample Size).
# Ogni valore rappresenta il numero equivalente di ordini storici
# che sostengono la stima del Take Rate.
# Calibrati per Ducati: volumi 2.000-5.000 unità/anno per variante,
# TR stimati a priori da expert judgment sales/product marketing.
# Ref: Morita, Thall, Müller 2008 — regola 10%: ESS <= 10% dati attesi.
MULTIPLIERS = {
    'NULLA': 50,
    'BASSA': 200,
    'MEDIA': 1_000,
    'ALTA':  5_000,
}

# Domanda mensile assunta (moto/mese) per tradurre σ_TR in unità fisiche
MONTHLY_BIKES = 500

# Quantile P99 (z-score distribuzione normale)
Z_P99 = 2.326

# ─────────────────────────────────────────────────────────────────────────────
# STILI EXCEL
# ─────────────────────────────────────────────────────────────────────────────

COL_VERDE    = "C6EFCE"   # verde chiaro
COL_GIALLO   = "FFEB9C"   # giallo chiaro
COL_ARANCIO  = "FFCC99"   # arancio chiaro
COL_ROSSO    = "FFC7CE"   # rosso chiaro
COL_HEADER   = "1F3864"   # blu scuro intestazioni
COL_TR_BG    = "D9E1F2"   # azzurro per riga TR
COL_BIANCO   = "FFFFFF"
COL_GRIGIO   = "F2F2F2"

FONT_WHITE   = Font(name='Calibri', bold=True, color="FFFFFF", size=10)
FONT_HEADER  = Font(name='Calibri', bold=True, color="FFFFFF", size=11)
FONT_TR      = Font(name='Calibri', bold=True, color="1F3864", size=10)
FONT_NORMAL  = Font(name='Calibri', size=10)
FONT_BOLD    = Font(name='Calibri', bold=True, size=10)

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
ALIGN_RIGHT  = Alignment(horizontal='right',  vertical='center')

def thin_border():
    s = Side(style='thin', color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

# ─────────────────────────────────────────────────────────────────────────────
# CALCOLO METRICHE
# ─────────────────────────────────────────────────────────────────────────────

def compute_metrics(tr: float, M: int, monthly_demand: int = MONTHLY_BIKES) -> dict:
    """
    Calcola le metriche di sensibilità Dirichlet per un dato TR e moltiplicatore M.

    Formule:
        alpha   = TR × M           (parametro Dirichlet)
        sigma   = sqrt(TR*(1-TR)/M) (deviazione standard del TR campionato)
        CV      = sigma / TR        (coefficiente di variazione)
        IC90%   = TR ± 1.645*sigma  (intervallo di confidenza al 90%)

    L'IC90% mostra il range entro cui cade il 90% dei take rate
    generati dalla simulazione in ogni singolo run.
    """
    sigma_tr      = np.sqrt(tr * (1.0 - tr) / M) if tr < 1.0 else 0.0
    cv            = (sigma_tr / tr * 100) if tr > 0 else 0.0
    ic_low        = max(0.0, (tr - 1.645 * sigma_tr)) * 100
    ic_high       = min(1.0, (tr + 1.645 * sigma_tr)) * 100
    ic_range      = ic_high - ic_low
    mean_demand   = monthly_demand * tr
    sigma_demand  = monthly_demand * sigma_tr
    ss_p99        = Z_P99 * sigma_demand        # safety stock aggiuntivo da incertezza TR

    # Etichetta qualitativa basata su CV
    if   cv <  3.0:  label, color = "Stabile",  COL_VERDE
    elif cv < 10.0:  label, color = "Moderato", COL_GIALLO
    elif cv < 25.0:  label, color = "Alto",      COL_ARANCIO
    else:            label, color = "Critico",   COL_ROSSO

    return dict(
        sigma_tr     = sigma_tr * 100,   # in %
        cv           = cv,
        ic_low       = ic_low,
        ic_high      = ic_high,
        ic_range     = ic_range,
        mean_demand  = mean_demand,
        sigma_demand = sigma_demand,
        ss_p99       = ss_p99,
        label        = label,
        color        = color,
    )

# ─────────────────────────────────────────────────────────────────────────────
# FOGLIO 1 — MAPPA SENSIBILITÀ
# ─────────────────────────────────────────────────────────────────────────────

def build_sheet_mappa(wb: Workbook):
    ws = wb.active
    ws.title = "Mappa Sensibilità"
    ws.sheet_view.showGridLines = False

    # ── Titolo ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:M1")
    ws["A1"] = "ANALISI DI SENSIBILITÀ — MOLTIPLICATORI DIRICHLET"
    ws["A1"].font      = Font(name='Calibri', bold=True, size=14, color="FFFFFF")
    ws["A1"].fill      = fill(COL_HEADER)
    ws["A1"].alignment = ALIGN_CENTER

    ws.merge_cells("A2:M2")
    ws["A2"] = (f"Domanda mensile assunta: {MONTHLY_BIKES} moto/mese  |  "
                f"IC al 90%  |  Safety Stock calcolato a P99 (z={Z_P99})")
    ws["A2"].font      = Font(name='Calibri', italic=True, size=10, color="FFFFFF")
    ws["A2"].fill      = fill("2E4F8A")
    ws["A2"].alignment = ALIGN_CENTER

    # ── Intestazioni colonne ─────────────────────────────────────────────────
    headers = [
        "TR\nNominale",
        "Livello\nAffidabilità",
        "Moltiplicatore\n(M)",
        "Osservazioni\nequivalenti",
        "σ TR\n(%)",
        "CV\n(%)",
        "IC 90%\nMinimo (%)",
        "IC 90%\nMassimo (%)",
        "Ampiezza\nIC 90% (pp)",
        f"Domanda media\n({MONTHLY_BIKES} moto)",
        "σ Domanda\n(unità)",
        "Safety Stock\nP99 (unità)",
        "Livello\nIncertezza",
    ]
    ROW_H = 4
    for col_i, h in enumerate(headers, start=1):
        cell = ws.cell(row=ROW_H, column=col_i, value=h)
        cell.font      = FONT_HEADER
        cell.fill      = fill(COL_HEADER)
        cell.alignment = ALIGN_CENTER
        cell.border    = thin_border()

    # ── Dati ─────────────────────────────────────────────────────────────────
    row = ROW_H + 1
    for tr in TR_VALUES:
        tr_label = f"{tr*100:.0f}%"
        first_row_for_tr = row

        for level, M in MULTIPLIERS.items():
            m = compute_metrics(tr, M)

            # Colori alternati per righe
            row_bg = COL_GRIGIO if (TR_VALUES.index(tr) % 2 == 0) else COL_BIANCO

            values = [
                tr_label,
                level,
                M,
                f"{M:,}",           # osservazioni equivalenti = M
                round(m['sigma_tr'], 3),
                round(m['cv'], 1),
                round(m['ic_low'],  2),
                round(m['ic_high'], 2),
                round(m['ic_range'], 2),
                round(m['mean_demand'], 1),
                round(m['sigma_demand'], 1),
                round(m['ss_p99'], 1),
                m['label'],
            ]

            for col_i, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_i, value=val)
                cell.border    = thin_border()
                cell.alignment = ALIGN_CENTER

                # Colore cella incertezza (colonna 13)
                if col_i == 13:
                    cell.fill = fill(m['color'])
                    cell.font = FONT_BOLD
                elif col_i == 1:
                    # TR label: sfondo azzurro, solo sulla prima riga del gruppo
                    if level == 'NULLA':
                        cell.fill = fill(COL_TR_BG)
                        cell.font = FONT_TR
                    else:
                        cell.fill = fill(row_bg)
                        cell.value = ""   # non ripetere il TR
                        cell.font  = FONT_NORMAL
                elif col_i == 2:
                    # Colore livello
                    lv_colors = {'NULLA': COL_ROSSO, 'BASSA': COL_ARANCIO,
                                 'MEDIA': COL_GIALLO, 'ALTA': COL_VERDE}
                    cell.fill = fill(lv_colors[level])
                    cell.font = FONT_BOLD
                else:
                    cell.fill = fill(row_bg)
                    cell.font = FONT_NORMAL

                # Grassetto per colonne chiave
                if col_i in (5, 6, 9, 12):
                    cell.font = Font(name='Calibri', bold=True, size=10)

            row += 1

        # Linea di separazione più spessa tra gruppi TR
        for col_i in range(1, len(headers) + 1):
            top_cell = ws.cell(row=first_row_for_tr, column=col_i)
            s_thick  = Side(style='medium', color="7F7F7F")
            s_thin   = Side(style='thin',   color="BFBFBF")
            top_cell.border = Border(
                left=s_thin, right=s_thin,
                top=s_thick, bottom=s_thin
            )

    # ── Larghezze colonne ─────────────────────────────────────────────────────
    col_widths = [9, 12, 16, 18, 10, 9, 14, 14, 16, 17, 16, 17, 14]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Altezze righe
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[ROW_H].height = 40
    for r in range(ROW_H + 1, row):
        ws.row_dimensions[r].height = 18

# ─────────────────────────────────────────────────────────────────────────────
# FOGLIO 2 — EFFETTO AMPLIFICAZIONE
# ─────────────────────────────────────────────────────────────────────────────

def build_sheet_amplificazione(wb: Workbook):
    """
    Mostra quanto il Safety Stock cambia tra ALTA e gli altri livelli,
    per ogni TR. Risponde alla domanda: 'Se uso BASSA invece di ALTA,
    di quanto si gonfia il safety stock?'
    """
    ws = wb.create_sheet("Effetto Amplificazione")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    ws["A1"] = "EFFETTO AMPLIFICAZIONE — Safety Stock: quanto cresce rispetto al livello ALTA?"
    ws["A1"].font      = Font(name='Calibri', bold=True, size=13, color="FFFFFF")
    ws["A1"].fill      = fill(COL_HEADER)
    ws["A1"].alignment = ALIGN_CENTER

    ws.merge_cells("A2:J2")
    ws["A2"] = ("Il rapporto SS_livello / SS_ALTA misura quante volte è più grande il safety stock "
                "al variare del moltiplicatore. È indipendente da TR.")
    ws["A2"].font      = Font(name='Calibri', italic=True, size=10, color="FFFFFF")
    ws["A2"].fill      = fill("2E4F8A")
    ws["A2"].alignment = ALIGN_CENTER

    # Header
    headers = ["TR Nominale", "Domanda media", "SS ALTA", "SS MEDIA",
               "SS BASSA", "SS NULLA",
               "MEDIA/ALTA (×)", "BASSA/ALTA (×)", "NULLA/ALTA (×)",
               "Range SS\n(NULLA-ALTA)"]
    ROW_H = 4
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=ROW_H, column=c, value=h)
        cell.font      = FONT_HEADER
        cell.fill      = fill(COL_HEADER)
        cell.alignment = ALIGN_CENTER
        cell.border    = thin_border()

    # Rapporto teorico (uguale per tutti i TR):
    # sqrt(M_alta / M_livello) = amplificazione su sigma, quindi su SS
    ratios = {lv: np.sqrt(MULTIPLIERS['ALTA'] / M) for lv, M in MULTIPLIERS.items()}

    row = ROW_H + 1
    for i, tr in enumerate(TR_VALUES):
        ss = {lv: compute_metrics(tr, M)['ss_p99'] for lv, M in MULTIPLIERS.items()}
        mean_d = MONTHLY_BIKES * tr

        row_bg = COL_GRIGIO if i % 2 == 0 else COL_BIANCO

        vals = [
            f"{tr*100:.0f}%",
            round(mean_d, 1),
            round(ss['ALTA'],  1),
            round(ss['MEDIA'], 1),
            round(ss['BASSA'], 1),
            round(ss['NULLA'], 1),
            round(ratios['MEDIA'], 2),
            round(ratios['BASSA'], 2),
            round(ratios['NULLA'], 2),
            round(ss['NULLA'] - ss['ALTA'], 1),
        ]

        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border    = thin_border()
            cell.alignment = ALIGN_CENTER
            cell.font      = FONT_NORMAL

            if c == 1:
                cell.fill = fill(COL_TR_BG)
                cell.font = FONT_TR
            elif c in (7, 8, 9):
                # Ratio columns — color by severity
                cell.fill = fill(COL_GIALLO if c == 7 else
                                 COL_ARANCIO if c == 8 else COL_ROSSO)
                cell.font = FONT_BOLD
            elif c == 10:
                cell.fill = fill(COL_ROSSO)
                cell.font = FONT_BOLD
            else:
                cell.fill = fill(row_bg)

        row += 1

    # Nota teorica
    row += 1
    ws.merge_cells(f"A{row}:J{row}")
    nota = (
        "Nota: i rapporti di amplificazione (×) sono indipendenti dal TR e dipendono solo dai moltiplicatori. "
        f"MEDIA/ALTA = ×{ratios['MEDIA']:.1f},  BASSA/ALTA = ×{ratios['BASSA']:.1f},  "
        f"NULLA/ALTA = ×{ratios['NULLA']:.1f}  "
        f"[Formula: sqrt(M_alta / M_livello) = sqrt({MULTIPLIERS['ALTA']:,} / M)]"
    )
    ws.merge_cells(f"A{row}:J{row}")
    cell = ws.cell(row=row, column=1, value=nota)
    cell.font      = Font(name='Calibri', italic=True, size=9, color="595959")
    cell.alignment = ALIGN_LEFT

    col_widths = [13, 16, 12, 12, 12, 12, 15, 15, 15, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[ROW_H].height = 40

# ─────────────────────────────────────────────────────────────────────────────
# FOGLIO 3 — GUIDA ALLA SCELTA DEL LIVELLO
# ─────────────────────────────────────────────────────────────────────────────

def build_sheet_guida(wb: Workbook):
    ws = wb.create_sheet("Guida alla Scelta")
    ws.sheet_view.showGridLines = False

    # Titolo
    ws.merge_cells("A1:G1")
    ws["A1"] = "GUIDA ALLA SCELTA DEL LIVELLO DI AFFIDABILITÀ"
    ws["A1"].font      = Font(name='Calibri', bold=True, size=13, color="FFFFFF")
    ws["A1"].fill      = fill(COL_HEADER)
    ws["A1"].alignment = ALIGN_CENTER

    # ── Tabella livelli ───────────────────────────────────────────────────────
    row = 3
    ws.merge_cells(f"A{row}:G{row}")
    ws.cell(row=row, column=1).value = "1. COSA SIGNIFICA OGNI LIVELLO"
    ws.cell(row=row, column=1).font  = Font(name='Calibri', bold=True, size=11, color=COL_HEADER)

    row += 1
    livelli_headers = ["Livello", "M", "Osservazioni equiv.", "σ su TR=10%",
                       "Amplif. SS vs ALTA", "Condizione tipica", "Esempio pratico"]
    for c, h in enumerate(livelli_headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font      = FONT_HEADER
        cell.fill      = fill(COL_HEADER)
        cell.alignment = ALIGN_CENTER
        cell.border    = thin_border()

    # Bugfix: deriva valori dinamicamente da MULTIPLIERS per evitare incoerenza
    # con il resto del report (mappa M, σ, amplificazione).
    _ref_tr = 0.10  # TR di riferimento per la colonna σ
    _M_alta = MULTIPLIERS['ALTA']
    _sigma_alta_ref = (_ref_tr * (1 - _ref_tr) / _M_alta) ** 0.5
    _descrizioni = {
        'ALTA':  ("Storico >3 anni, dati consolidati",
                  "Ducati Red su MSV4: TR stabile da anni"),
        'MEDIA': ("Forecast commerciale, 1-2 stagioni",
                  "Colori standard nuova generazione"),
        'BASSA': ("Pochi mesi di dati, mercato volatile",
                  "Optional su modello appena lanciato"),
        'NULLA': ("Stima/ipotesi senza dati veri",
                  "Colori speciali, serie limitate"),
    }
    # Ordina: ALTA, MEDIA, BASSA, NULLA (decrescente per M)
    _ordered_levels = sorted(MULTIPLIERS.keys(), key=lambda k: -MULTIPLIERS[k])
    livelli_data = []
    for _lv in _ordered_levels:
        _M = MULTIPLIERS[_lv]
        _sigma = (_ref_tr * (1 - _ref_tr) / _M) ** 0.5
        _ampl = _sigma / _sigma_alta_ref if _sigma_alta_ref > 0 else 0.0
        _cond, _es = _descrizioni.get(_lv, ("", ""))
        livelli_data.append((
            _lv, _M, f"~{_M:,}".replace(",", "."),
            f"{_sigma*100:.2f}%",
            f"×{_ampl:.1f}",
            _cond, _es,
        ))
    lv_colors = {'ALTA': COL_VERDE, 'MEDIA': COL_GIALLO,
                 'BASSA': COL_ARANCIO, 'NULLA': COL_ROSSO}

    row += 1
    for ld in livelli_data:
        for c, v in enumerate(ld, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border    = thin_border()
            cell.alignment = ALIGN_CENTER if c <= 5 else ALIGN_LEFT
            cell.font      = FONT_NORMAL
            if c == 1:
                cell.fill = fill(lv_colors[ld[0]])
                cell.font = FONT_BOLD
            else:
                cell.fill = fill(COL_GRIGIO if row % 2 == 0 else COL_BIANCO)
        row += 1

    # ── Regola pratica TR basso ───────────────────────────────────────────────
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    ws.cell(row=row, column=1).value = "2. REGOLA PRATICA: COMPONENTI CON TR BASSO"
    ws.cell(row=row, column=1).font  = Font(name='Calibri', bold=True, size=11, color=COL_HEADER)

    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    testo = ("Per TR < 5%, la scelta del moltiplicatore ha un impatto enorme (CV può superare il 50% con NULLA). "
             "Per questi componenti si raccomanda di usare almeno MEDIA, oppure di raccogliere più dati "
             "prima di fissare il TR. Con TR > 30%, invece, anche BASSA produce risultati ragionevolmente stabili.")
    cell = ws.cell(row=row, column=1, value=testo)
    cell.font      = Font(name='Calibri', size=10)
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    cell.fill      = fill("EBF3FB")
    cell.border    = thin_border()
    ws.row_dimensions[row].height = 50

    # ── CV soglie ─────────────────────────────────────────────────────────────
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws.cell(row=row, column=1).value = "3. LETTURA DEL SEMAFORO (CV = Coefficiente di Variazione)"
    ws.cell(row=row, column=1).font  = Font(name='Calibri', bold=True, size=11, color=COL_HEADER)

    row += 1
    semaforo_data = [
        (COL_VERDE,   "STABILE",  "CV < 3%",
         "Il TR simulato oscilla poco attorno al valore nominale. "
         "Il safety stock è affidabile."),
        (COL_GIALLO,  "MODERATO", "3% ≤ CV < 10%",
         "Variabilità contenuta. I risultati sono utili ma tieni un margine di sicurezza."),
        (COL_ARANCIO, "ALTO",     "10% ≤ CV < 25%",
         "Il TR può discostarsi significativamente. "
         "Valuta se aumentare il moltiplicatore o raccogliere più dati."),
        (COL_ROSSO,   "CRITICO",  "CV ≥ 25%",
         "Altissima incertezza. Il safety stock calcolato potrebbe essere "
         "molto diverso dalla realtà. Necessario rivedere il livello di affidabilità."),
    ]
    for sd in semaforo_data:
        col_hex, etichetta, soglia, spiegazione = sd
        ws.cell(row=row, column=1).value = etichetta
        ws.cell(row=row, column=1).fill  = fill(col_hex)
        ws.cell(row=row, column=1).font  = FONT_BOLD
        ws.cell(row=row, column=1).alignment = ALIGN_CENTER
        ws.cell(row=row, column=1).border    = thin_border()

        ws.cell(row=row, column=2).value = soglia
        ws.cell(row=row, column=2).font  = FONT_BOLD
        ws.cell(row=row, column=2).fill  = fill(col_hex)
        ws.cell(row=row, column=2).alignment = ALIGN_CENTER
        ws.cell(row=row, column=2).border    = thin_border()

        ws.merge_cells(f"C{row}:G{row}")
        cell = ws.cell(row=row, column=3, value=spiegazione)
        cell.font      = FONT_NORMAL
        cell.fill      = fill(COL_GRIGIO if row % 2 == 0 else COL_BIANCO)
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        cell.border    = thin_border()
        ws.row_dimensions[row].height = 38
        row += 1

    col_widths = [14, 18, 14, 14, 18, 28, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30

# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("=== Analisi di Sensibilità Dirichlet ===")
    print(f"TR analizzati:     {[f'{t*100:.0f}%' for t in TR_VALUES]}")
    print(f"Livelli:           {list(MULTIPLIERS.keys())}")
    print(f"Domanda mensile:   {MONTHLY_BIKES} moto/mese")
    print(f"Safety Stock:      P99 (z={Z_P99})")
    print()

    # Stampa preview a console
    print(f"{'TR':>5} {'Livello':>8}  {'M':>8}  {'s_TR':>7}  {'CV':>7}  "
          f"{'IC90% min':>10}  {'IC90% max':>10}  {'SS P99':>8}  Incertezza")
    print("-" * 95)
    for tr in TR_VALUES:
        for level, M in MULTIPLIERS.items():
            m = compute_metrics(tr, M)
            print(f"{tr*100:>4.0f}% {level:>8}  {M:>8,}  "
                  f"{m['sigma_tr']:>6.3f}%  {m['cv']:>6.1f}%  "
                  f"{m['ic_low']:>9.2f}%  {m['ic_high']:>9.2f}%  "
                  f"{m['ss_p99']:>7.1f}  {m['label']}")
        print()

    # Genera Excel
    wb = Workbook()
    build_sheet_mappa(wb)
    build_sheet_amplificazione(wb)
    build_sheet_guida(wb)

    os.makedirs("Output", exist_ok=True)
    out_path = r"Output\sensitivity_dirichlet.xlsx"
    wb.save(out_path)
    print(f"\n[OK] File salvato: {out_path}")
    print(f"     Fogli: 'Mappa Sensibilità', 'Effetto Amplificazione', 'Guida alla Scelta'")

if __name__ == "__main__":
    main()
