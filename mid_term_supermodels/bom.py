# -*- coding: utf-8 -*-
"""Modulo BOM (Distinta Base): costruisce l'elenco componenti per ogni moto.

In parole semplici: parte dalle distinte base grezze (cosa serve per costruire
ogni configurazione di moto) e produce un catalogo pulito di tutte le
configurazioni valide e dei loro componenti, applicando le regole di visibilita'
(Mostra/Nascondi) e le esclusioni tra optional incompatibili.

Consolida tre responsabilita':
- parser BOM e espansione delle dipendenze OR/IN/bundle a partire
  da Mappatura.xlsx, TR e dei BOM grezzi (ex gestione_dipendenze_V2);
- generazione delle configurazioni univoche per file (ex skubundle_OPTIMIZED_V2);
- orchestrazione multi-BOM con quality checks e consolidamento finale
  (ex process_all_bom_combinations).

Le fonti di verita' sono il foglio Schema della Mappatura (mapping
CHAR/VALUE/Bundle) e la colonna B (CONCAT) del TR (colonne calcolate/bundle).
"""

import os
import pandas as pd
import numpy as np
import re
from pathlib import Path
from typing import Dict, Any, List, Tuple, Optional, Set
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from mid_term_supermodels._logging import get_logger
from mid_term_supermodels.exceptions import BOMParseError, InputFileError
log = get_logger("bom")

# ============================================================================
# CONFIGURAZIONE
# ============================================================================

DEBUG_MODE = False


# ============================================================================
# FUNZIONI DI CARICAMENTO DINAMICO DAL TR
# ============================================================================

def load_calculated_cols_from_tr(tr_file: str) -> List[Tuple[str, str, str, Optional[str]]]:
    """
    Deriva dinamicamente le colonne calcolate dal file TR (colonna B - CONCAT).

    LOGICA:
    - Legge il foglio TR_Optional più vecchio disponibile (fonte di riferimento globale)
    - Parte sinistra di " - " -> gruppi caratteristiche (es: FAIRING COLOR, RIDER SEAT)
    - "BUNDLE - NomePack" (senza NOT) -> bundle positivi (es: Radar Pack, Tech Pack)

    Il quarto elemento della tupla:
    - Per caratteristiche: uguale al nome gruppo (chiave di ricerca su DESC)
    - Per bundle: None (ricerca per Bundle name)

    Args:
        tr_file: Path al file TR Excel

    Returns:
        Lista di tuple (bundle, desc, nome_colonna, chiave_ricerca)
    """
    # Individua il foglio TR_Optional più vecchio disponibile
    xl = pd.ExcelFile(tr_file)
    sheets = xl.sheet_names

    from mid_term_supermodels.tr import _parse_month_str as _pms
    opt_sheets = sorted(
        [s for s in sheets if s.lower().startswith('tr_optional_')],
        key=lambda s: _pms(s[len('tr_optional_'):])  # ordine CRONOLOGICO (coerente con tr_parser_V2)
    )

    if not opt_sheets:
        raise ValueError(f"Nessun foglio TR_Optional_* trovato in {tr_file}")

    first_opt = opt_sheets[0]
    log.info("[V2] TR fonte di verità: '%s'", first_opt)

    tr_df = pd.read_excel(tr_file, sheet_name=first_opt, header=0)

    # Individua colonna CONCAT (seconda colonna dati)
    concat_col = None
    for col in tr_df.columns:
        if col and 'concat' in str(col).lower():
            concat_col = col
            break
    if concat_col is None:
        # fallback: prendi seconda colonna
        concat_col = tr_df.columns[1]

    calculated = []
    seen_chars = set()
    seen_bundles = set()

    for raw_val in tr_df[concat_col]:
        if not raw_val or pd.isna(raw_val):
            continue
        val_str = str(raw_val).strip()

        # Salta la riga di header interna
        if 'concat' in val_str.lower() or 'nome opzione' in val_str.lower():
            continue

        parts = val_str.split(' - ', 1)
        if len(parts) < 2:
            continue

        group = parts[0].strip()
        rest = parts[1].strip()

        if group == 'BUNDLE':
            # Bundle: scarta quelli con NOT
            if rest.upper().startswith('NOT '):
                continue
            bundle_name = rest  # già il nome completo del bundle
            if bundle_name not in seen_bundles:
                # (bundle_intestazione, desc_intestazione, nome_colonna, chiave_ricerca)
                calculated.append(('', '', bundle_name, None))
                seen_bundles.add(bundle_name)
        else:
            # Caratteristica: usa il gruppo come nome colonna e chiave ricerca
            if group not in seen_chars:
                calculated.append(('', '', group, group))
                seen_chars.add(group)

    log.info('[V2] Colonne calcolate derivate dal TR: %s', len(calculated))
    for c in calculated:
        log.info("     - '%s' (chiave: %s)", c[2], c[3])

    return calculated


def load_priority_order_from_tr(tr_file: str, group_prefix: str) -> List[str]:
    """
    Estrae l'ordine di priorità per un gruppo dal TR.

    L'ordine di apparizione nel TR definisce la priorità:
    la prima opzione trovata per un componente viene usata.

    Args:
        tr_file: Path al file TR Excel
        group_prefix: Prefisso del gruppo (es: 'RIDER SEAT', 'PILLION SEAT')

    Returns:
        Lista di nomi opzioni in ordine di priorità
    """
    xl = pd.ExcelFile(tr_file)
    sheets = xl.sheet_names
    from mid_term_supermodels.tr import _parse_month_str as _pms
    opt_sheets = sorted(
        [s for s in sheets if s.lower().startswith('tr_optional_')],
        key=lambda s: _pms(s[len('tr_optional_'):])  # ordine CRONOLOGICO (coerente con tr_parser_V2)
    )

    if not opt_sheets:
        return []

    tr_df = pd.read_excel(tr_file, sheet_name=opt_sheets[0], header=0)

    concat_col = None
    for col in tr_df.columns:
        if col and 'concat' in str(col).lower():
            concat_col = col
            break
    if concat_col is None:
        concat_col = tr_df.columns[1]

    prefix_search = group_prefix + ' - '
    order = []
    seen = set()

    for raw_val in tr_df[concat_col]:
        if not raw_val or pd.isna(raw_val):
            continue
        val_str = str(raw_val).strip()
        if val_str.startswith(prefix_search):
            option_name = val_str[len(prefix_search):]
            if option_name not in seen:
                order.append(option_name)
                seen.add(option_name)

    log.info("[V2] Ordine priorità '%s': %s", group_prefix, order)
    return order


def load_rims_substitutions(mapping_file: str) -> Dict[str, List[Tuple[str, str]]]:
    """
    Carica le sostituzioni RIMS dal foglio RIMS_Substitutions in Mappatura.

    Il foglio deve avere queste colonne:
    | Base_Value | Model_Code | Substitution |
    |------------|-----------|--------------|
    | Forged     | MSV4PP    | Forged(17"/17") - Pikes Peak |
    | Forged     | MSV4S     | Forged(19"/17") - Black |

    Se il foglio non esiste, ritorna dizionario vuoto (nessuna sostituzione).

    Args:
        mapping_file: Path al file Mappatura Excel

    Returns:
        Dict: {base_value: [(model_code, substitution), ...]}
    """
    try:
        xl = pd.ExcelFile(mapping_file)
        if 'RIMS_Substitutions' not in xl.sheet_names:
            log.info('[V2] Foglio RIMS_Substitutions non trovato — sostituzione RIMS disabilitata')
            return {}

        config_df = pd.read_excel(mapping_file, sheet_name='RIMS_Substitutions', header=0)
        substitutions = {}

        for row in config_df.itertuples(index=False):
            base = str(row[0]).strip()
            model = str(row[1]).strip()
            sub = str(row[2]).strip()

            if base and model and sub:
                if base not in substitutions:
                    substitutions[base] = []
                substitutions[base].append((model, sub))

        log.info('[V2] Sostituzioni RIMS caricate: %s regole', sum(len(v) for v in substitutions.values()))
        return substitutions

    except Exception as e:
        log.info('[V2] RIMS_Substitutions non disponibile: %s', e)
        return {}


# ============================================================================
# FUNZIONI DI UTILITÀ
# ============================================================================

def _remove_parentheses_variant(text: str) -> str:
    """Rimuove la parte tra parentesi da un testo per normalizzazione."""
    if not text:
        return text
    return re.sub(r'\s*\([^)]*\)', '', text).strip()


def _build_column_index(col_structure: List[Tuple]) -> Dict[Tuple, List[int]]:
    """
    Pre-calcola indice per ricerca veloce delle colonne nella matrice.
    Struttura: (bundle, desc, readable) -> [indici]
    """
    index = {}
    for idx, col_data in enumerate(col_structure):
        if len(col_data) >= 3:
            b, d, r = col_data[0], col_data[1], col_data[2]
            key = (b, d, r)
            if key not in index:
                index[key] = []
            index[key].append(idx)

            r_norm = _remove_parentheses_variant(r)
            if r_norm != r:
                key_norm = (b, d, r_norm)
                if key_norm not in index:
                    index[key_norm] = []
                index[key_norm].append(idx)

    return index



def _get_not_bundle_name(bundle_name: str) -> str:
    """
    Ritorna il nome del bundle negativo corrispondente.
    Es: 'Tech Pack' -> 'Not Tech Pack'
    """
    return f"Not {bundle_name}"


# ============================================================================
# CLASSE PRINCIPALE: ConditionParser
# ============================================================================

class ConditionParser:
    """
    Parser per condizioni logiche presenti nel BOM150.
    Identica alla V1 — nessuna modifica necessaria.
    """

    def __init__(self, mapping_df=None, active_cv_sets: Dict[str, Set[str]] = None,
                 filter_untraced: bool = False,
                 unknown_ct: Set[str] = None,
                 unknown_cv: Set[tuple] = None):
        """Parser delle condizioni BOM. Costruisce il lookup CHAR/VALUE/Bundle
        dalla Mappatura e conserva i set di CV attivi (TR>0) e i CT/CV non
        tracciati, usati per filtrare ed espandere le condizioni."""
        self.mapping_dict = self._build_mapping_dict(mapping_df) if mapping_df is not None else {}
        # active_cv_sets: {CHAR_code: {VALUE_code, ...}} — solo CV con TR > 0.
        # Dict vuoto = nessun filtro (tutti i CV vengono restituiti).
        self.active_cv_sets: Dict[str, Set[str]] = active_cv_sets or {}
        # Filtro CT/CV non tracciati
        self.filter_untraced: bool = filter_untraced
        self.unknown_ct: Set[str]    = unknown_ct or set()
        self.unknown_cv: Set[tuple]  = unknown_cv or set()  # {(ct, cv), ...}

    def _build_mapping_dict(self, mapping_df: pd.DataFrame) -> Dict[str, Dict[str, List[Tuple[str, str, str]]]]:
        """Costruisce il lookup {CHAR: {VALUE: [(Bundle, Descrizione, Readable)]}}
        dalla Mappatura, base per risolvere codici caratteristica/valore."""
        mapping = {}
        for row in mapping_df.itertuples(index=False):
            char = str(getattr(row, 'CHAR', '')).strip()
            value = str(getattr(row, 'VALUE', '')).strip()
            bundle = str(getattr(row, 'Bundle', '')).strip()
            descrizione = str(getattr(row, 'Descrizione', '')).strip()
            readable = str(getattr(row, 'Readable', '')).strip()

            if not char or not value:
                continue
            if char not in mapping:
                mapping[char] = {}
            if value not in mapping[char]:
                mapping[char][value] = []
            mapping[char][value].append((bundle, descrizione, readable))

        log.info('\\n[OK] Mapping: %s CHAR codes, %s totali', len(mapping), sum(len(v) for v in mapping.values()))

        if DEBUG_MODE:
            debug_data = []
            for c, vals in mapping.items():
                for v, entries in vals.items():
                    for b, d, r in entries:
                        debug_data.append({'CHAR': c, 'VALUE': v, 'Bundle': b, 'Descrizione': d, 'Readable': r})
            pd.DataFrame(debug_data).to_excel('mapping_debug_v2.xlsx', index=False)

        return mapping

    def get_column_info(self, char_code: str, value_code: str) -> List[Tuple[str, str, str]]:
        """Ritorna le tuple (Bundle, Descrizione, Readable) mappate per la coppia
        CHAR/VALUE; lista vuota se non presenti."""
        return self.mapping_dict.get(char_code, {}).get(value_code, [])

    def get_all_values_for_char(self, char_code: str) -> List[str]:
        """
        Restituisce i VALUE definiti in mappatura per un dato CHAR.

        Filtri applicati (in ordine):
        1. Esclude marker interni 'NE ...' (es: 'NE MSV4PP')
        2. Se active_cv_sets è disponibile per questo CHAR, filtra solo i CV
           con TR > 0 nel TR file. Evita di espandere NOT(IN) con CV "morti"
           che la simulazione non può mai generare (TR=0 -> match impossibile).

        Usato per espandere:
        - ZPRODGRP ne 'VAL'          -> una riga eq per ogni modello eccetto VAL
        - NOT ($root.CT IN (lista))  -> una riga eq per ogni CV non nella lista
        """
        all_values = [
            v for v in self.mapping_dict.get(char_code, {}).keys()
            if not str(v).upper().startswith('NE ')
        ]
        if char_code in self.active_cv_sets:
            active = self.active_cv_sets[char_code]
            filtered = [v for v in all_values if v in active]
            if len(filtered) < len(all_values):
                log.info('  [V2] Filtro TR attivo: %s %s -> %s CV (scartati %s con TR=0)', char_code, len(all_values), len(filtered), len(all_values)-len(filtered))
            return filtered
        return all_values

    def split_or_conditions(self, condition: str) -> List[str]:
        """Divide una condizione sui separatori OR (case-insensitive) nelle sue
        sotto-condizioni."""
        if not condition or pd.isna(condition):
            return [condition]
        return [p.strip() for p in re.split(r'\s+OR\s+', condition, flags=re.IGNORECASE)]

    def parse_condition(self, condition: str) -> Dict[str, Any]:
        """Interpreta una condizione BOM ($root.VAR eq/ne 'x', IN(...), NOT(...))
        e ritorna {VAR: {op, value/values}}. Gestisce NOT(eq)->ne, NOT(IN) e
        l'eventuale filtro dei CT/CV non tracciati."""
        if not condition or pd.isna(condition):
            return {}
        result = {}

        # ── STEP 1: NOT ($root.VAR eq 'VALUE') -> tratta come 'ne' ──────────
        # Registra anche gli span di questi blocchi per escluderli dall'eq scan.
        # Es: NOT ($root.ZCOUNTRY eq 'CDN') -> ZCOUNTRY: {op='ne', value='CDN'}
        not_spans = []
        for m in re.finditer(
            r'NOT\s*\(\s*\$root\.(\w+)\s+eq\s+["\']([^"\']+)["\']\s*\)',
            condition, re.IGNORECASE
        ):
            var, value = m.groups()
            result[var] = {'op': 'ne', 'value': value}
            not_spans.append((m.start(), m.end()))

        def _in_not_span(pos: int) -> bool:
            """True se la posizione cade dentro un blocco NOT(eq) gia' gestito."""
            return any(s <= pos < e for s, e in not_spans)

        # ── STEP 2: eq normali (escludi quelli coperti da NOT) ──────────────
        for match in re.finditer(r'\$root\.(\w+)\s+eq\s+["\']([^"\']+)["\']', condition):
            if not _in_not_span(match.start()):
                var, value = match.groups()
                result[var] = {'op': 'eq', 'value': value}

        # ── STEP 3: ne espliciti ─────────────────────────────────────────────
        for match in re.finditer(r'\$root\.(\w+)\s+ne\s+["\']([^"\']+)["\']', condition, re.IGNORECASE):
            var, value = match.groups()
            result[var] = {'op': 'ne', 'value': value}

        # ── STEP 4: IN normali (escludi quelli dentro NOT) ──────────────────
        # Raccogli gli span dei blocchi NOT (IN) per non riprocessarli come IN.
        # Se expand_all_conditions funziona correttamente questa condizione non
        # dovrebbe mai arrivare qui, ma la protezione evita bug silenziosi.
        not_in_spans = [
            (m.start(), m.end())
            for m in re.finditer(
                r'NOT\s*\(\s*\$root\.\w+\s+IN\s*\([^)]+\)\s*\)',
                condition, re.IGNORECASE
            )
        ]

        def _in_not_in_span(pos: int) -> bool:
            """True se la posizione cade dentro un blocco NOT(IN) gia' gestito."""
            return any(s <= pos < e for s, e in not_in_spans)

        for match in re.finditer(r'\$root\.(\w+)\s+IN\s*\(([^)]+)\)', condition, re.IGNORECASE):
            if not _in_not_in_span(match.start()):
                var, values_str = match.groups()
                values = re.findall(r"['\"]([^'\"]+)['\"]", values_str)
                result[var] = {'op': 'IN', 'values': values}

        # ── Filtro CT/CV non tracciati ────────────────────────────────────────
        # Comportamento (filter_untraced=True):
        # - CT sconosciuto intero   : scarta il predicato, valuta il resto (AND)
        # - eq/ne con CV sconosciuto: scarta il predicato, valuta il resto (AND)
        # - IN con CV sconosciuto   : rimuovi quel CV dalla lista; se lista vuota scarta predicato
        # - Tutti i predicati scartati: elimina intera dipendenza (return None)
        if self.filter_untraced and result:
            filtered = {}
            for ct, info in result.items():
                if ct in self.unknown_ct:
                    continue                          # CT intero non tracciato -> scarta predicato
                op = info['op']
                if op in ('eq', 'ne'):
                    if (ct, info['value']) in self.unknown_cv:
                        continue                      # CV sconosciuto -> scarta solo questo predicato
                    filtered[ct] = info
                elif op == 'IN':
                    known_values = [v for v in info['values']
                                    if (ct, v) not in self.unknown_cv]
                    if not known_values:
                        continue                      # lista svuotata -> scarta predicato
                    filtered[ct] = {**info, 'values': known_values}
                else:
                    filtered[ct] = info
            # Se tutti i predicati erano sconosciuti -> escludi intera dipendenza
            if not filtered:
                return None
            result = filtered

        return result


# ============================================================================
# CARICAMENTO MAPPING
# ============================================================================

def build_shared_cv_map(mapping_df: pd.DataFrame) -> Dict[Tuple[str, str], List[str]]:
    """
    Pre-calcola da Schema quali (CHAR, VALUE) sono condivisi tra più bundle.

    Restituisce un dizionario:
        (CHAR, VALUE) -> [bundle1, bundle2, ...]

    Se la lista ha 1 solo elemento -> CT/CV esclusivo di quel bundle.
    Se la lista ha N elementi -> CT/CV condiviso tra N bundle.

    Usato per la logica di split/discriminante:
    - Condiviso + nessun discriminante esclusivo -> split in N righe
    - Condiviso + almeno un discriminante esclusivo -> attiva solo il bundle esclusivo

    NOTA: considera solo i bundle non vuoti e non negativi ("Not ..."),
    perché i bundle negativi non discriminano la variante positiva.
    """
    cv_to_bundles: Dict[Tuple[str, str], List[str]] = {}

    for row in mapping_df.itertuples(index=False):
        char  = str(row.CHAR).strip()
        value = str(row.VALUE).strip()
        bundle = str(row.Bundle).strip()

        if not char or not value or not bundle:
            continue
        if bundle.lower().startswith('not '):
            continue  # i bundle negativi non contano come discriminanti

        key = (char, value)
        if key not in cv_to_bundles:
            cv_to_bundles[key] = []
        if bundle not in cv_to_bundles[key]:
            cv_to_bundles[key].append(bundle)

    shared = {k: v for k, v in cv_to_bundles.items() if len(v) > 1}
    exclusive = {k: v for k, v in cv_to_bundles.items() if len(v) == 1}

    log.info('[V2] CT/CV condivisi tra più bundle: %s', len(shared))
    log.info('[V2] CT/CV esclusivi di un solo bundle: %s', len(exclusive))
    if shared:
        for (c, v), bundles in list(shared.items())[:5]:
            log.info('     (%s, %s) -> %s', c, v, bundles)

    # Restituisce il mapping completo (sia shared che exclusive)
    return cv_to_bundles


def load_mapping_from_file(mapping_file: str = '.\\Input\\Mappatura.xlsx') -> pd.DataFrame:
    """
    Carica il file di mappatura e crea DataFrame normalizzato.
    Tenta prima il formato long (sheet 'Schema_Long', colonne Bundle/CT/CT_Desc/CV/CV_Desc),
    altrimenti cade in fallback sul formato wide (sheet 'Schema', colonne Opz1/Desc1...).
    """
    xl = pd.ExcelFile(mapping_file)

    # ── Formato long (Schema_Long) ────────────────────────────────────────────
    if 'Schema_Long' in xl.sheet_names:
        df = xl.parse('Schema_Long', header=0)
        required = {'Bundle', 'CT', 'CT_Desc', 'CV', 'CV_Desc'}
        if required.issubset(set(df.columns)):
            mapping_rows = []
            # Perf: to_dict('records') ~10x più veloce di iterrows()
            for row in df.to_dict('records'):
                bundle    = str(row.get('Bundle', '')).strip()
                char_code = str(row.get('CT', '')).strip()
                char_desc = str(row.get('CT_Desc', '')).strip()
                value     = str(row.get('CV', '')).strip()
                readable  = str(row.get('CV_Desc', '')).strip()

                if not char_code or char_code.lower() in ('nan', ''):
                    continue
                if not value or value.lower() in ('nan', ''):
                    continue

                if value.upper() in ('TRUE', 'FALSE', 'NE TRUE'):
                    if not readable or readable.upper() in ('TRUE', 'FALSE', 'NE TRUE', 'NAN'):
                        readable = value

                mapping_rows.append({
                    'Bundle':      bundle if bundle and bundle.lower() not in ('nan', 'no') else '',
                    'Descrizione': char_desc,
                    'CHAR':        char_code,
                    'VALUE':       value,
                    'Readable':    readable if readable and readable.lower() not in ('nan', '') else value,
                })

            result_df = pd.DataFrame(mapping_rows)
            log.info('[OK] Mappatura Long: %s entries da %s righe Schema_Long', len(result_df), len(df))
            return result_df

    # ── Fallback formato wide (Schema) ────────────────────────────────────────
    df = xl.parse('Schema', header=0)
    mapping_rows = []

    for row in df.itertuples(index=False):
        bundle = str(getattr(row, 'Bundle', '')).strip() if hasattr(row, 'Bundle') else ''
        char_code = str(row[1]).strip() if len(row) > 1 else ''
        char_desc = str(row[2]).strip() if len(row) > 2 else ''

        if not char_code or pd.isna(row[1]):
            continue

        col_idx = 3
        row_list = list(row)
        while col_idx < len(row_list):
            opz_col_val  = row_list[col_idx]     if col_idx     < len(row_list) else None
            desc_col_val = row_list[col_idx + 1] if col_idx + 1 < len(row_list) else None

            value_code = str(opz_col_val).strip()  if opz_col_val  is not None else ''
            readable   = str(desc_col_val).strip() if desc_col_val is not None else ''

            if value_code and value_code.lower() not in ('', 'nan', 'none'):
                if value_code.upper() in ('TRUE', 'FALSE', 'NE TRUE'):
                    if not readable or readable.upper() in ('TRUE', 'FALSE', 'NE TRUE', 'NAN'):
                        readable = value_code
                else:
                    readable = readable if readable and readable.lower() not in ('', 'nan') else value_code

                mapping_rows.append({
                    'Bundle':      bundle if bundle and bundle.lower() not in ('nan', 'no') else '',
                    'Descrizione': char_desc,
                    'CHAR':        char_code,
                    'VALUE':       value_code,
                    'Readable':    readable,
                })

            col_idx += 2

    result_df = pd.DataFrame(mapping_rows)
    log.info('[OK] Mappatura Wide: %s entries da %s righe Schema', len(result_df), len(df))
    return result_df




# ============================================================================
# FUNZIONE DI POPOLAMENTO RIGA — VERSIONE 2 (SENZA HARDCODING)
# ============================================================================

def _populate_row_from_condition_v2(row, condition, parser, col_structure,
                                    calc_indices, calculated_cols, col_index,
                                    bundle_set: set):
    """
    Popola una riga della matrice in base alla condizione parsata.

    DIFFERENZE RISPETTO A V1:
    -------------------------
    1. Nessun if/elif hardcoded per Smoked Windshield, Enduro Pack, ecc.
    2. Nessun pattern negativo hardcoded (gestito dal Bundle "Not XYZ" in Schema)
    3. Logica "Yes" basata su bundle_set (non su keyword)
    4. Logica unificata: per bundle -> valore = nome bundle (o "Yes"); per caratteristiche -> readable

    Args:
        row: Lista rappresentante la riga da popolare
        condition: Stringa condizione da parsare
        parser: Istanza ConditionParser
        col_structure: Lista struttura colonne
        calc_indices: Dizionario pre-calcolato per colonne calcolate
        calculated_cols: Lista definizione colonne calcolate
        col_index: Indice pre-calcolato per lookup veloce
        bundle_set: Set dei nomi bundle positivi (da TR)
    """
    if condition and not pd.isna(condition):
        parsed = parser.parse_condition(condition)

        # parse_condition restituisce None quando filter_untraced=True
        # e tutti i predicati erano non tracciati -> componente da escludere
        if parsed is None:
            return False

        for var, info in parsed.items():
            if info['op'] == 'eq':
                all_matches = parser.get_column_info(var, info['value'])
                for bundle, desc, readable in all_matches:
                    target_indices = []
                    key = (bundle if bundle else '', desc, readable)
                    if key in col_index:
                        target_indices.extend(col_index[key])
                    readable_norm = _remove_parentheses_variant(readable)
                    if readable_norm != readable:
                        key_norm = (bundle if bundle else '', desc, readable_norm)
                        if key_norm in col_index:
                            target_indices.extend(col_index[key_norm])
                    if not target_indices:
                        for idx, col_data in enumerate(col_structure):
                            if len(col_data) == 3:
                                b, d, r = col_data
                                readable_match = (r == readable or _remove_parentheses_variant(r) == readable)
                                context_match = (b == bundle) if bundle else (d == desc)
                                if readable_match and context_match:
                                    target_indices.append(idx)
                    for idx in target_indices:
                        row[idx] = f"eq '{info['value']}'"

            elif info['op'] == 'IN':
                for val in info['values']:
                    all_matches = parser.get_column_info(var, val)
                    for bundle, desc, readable in all_matches:
                        target_indices = []
                        for idx, col_data in enumerate(col_structure):
                            if len(col_data) == 3:
                                b, d, r = col_data
                                readable_match = (r == readable or _remove_parentheses_variant(r) == readable)
                                context_match = (b == bundle) if bundle else (d == desc)
                                if readable_match and context_match:
                                    target_indices.append(idx)
                        for idx in target_indices:
                            row[idx] = f"eq '{val}'"

            elif info['op'] == 'ne':
                # Le variabili ZPRODGRP con ne sono già state espanse in condizioni
                # eq separate da expand_all_conditions -> skip per evitare doppio processing.
                if 'ZPRODGRP' in var.upper():
                    continue
                ne_value = f"NE {info['value']}"
                all_matches = parser.get_column_info(var, ne_value)
                if not all_matches:
                    all_matches = parser.get_column_info(var, info['value'])
                for bundle, desc, readable in all_matches:
                    target_indices = []
                    for idx, col_data in enumerate(col_structure):
                        if len(col_data) == 3:
                            b, d, r = col_data
                            readable_match = (r == readable or _remove_parentheses_variant(r) == readable)
                            context_match = (b == bundle) if bundle else (d == desc)
                            if readable_match and context_match:
                                target_indices.append(idx)
                    for idx in target_indices:
                        row[idx] = f"ne '{info['value']}'"

    # =========================================================================
    # CALCOLA COLONNE CALCOLATE — LOGICA UNIFICATA V2 (ZERO HARDCODING)
    # =========================================================================
    calc_offset = len(calculated_cols)

    for i, c in enumerate(calculated_cols):
        calc_name, search_key = c[2], c[3]
        value = "No"
        is_bundle = (search_key is None)  # True = bundle, False = caratteristica

        for col_idx in calc_indices.get(calc_name, []):
            if row[col_idx]:
                bundle_col, desc_col, readable_col = col_structure[col_idx]

                if is_bundle:
                    # -----------------------------------------------------------
                    # LOGICA BUNDLE (V2):
                    # - Se il Bundle inizia con "Not " -> "Not" (provvisorio, non break)
                    # - Altrimenti -> nome del bundle (definitivo, break)
                    #
                    # REGOLA: il match positivo vince sul match negativo.
                    # Un bundle può avere più indicatori (es: Radar Pack = Cruise
                    # Control OR Blind Spot Detection). Se la condizione attiva uno
                    # positivamente e un altro negativamente, la presenza del
                    # positivo deve vincere.
                    # Esempio: $root.CT001979 eq 'TRUE' AND $root.CT001007 ne 'CV001013'
                    # -> Blind Spot Detection presente -> Radar Pack = 'Yes' (non 'Not')
                    # -----------------------------------------------------------
                    if bundle_col.lower().startswith('not '):
                        value = "Not"
                        # NON fare break: continua a cercare un match positivo
                    else:
                        # Match positivo: definitivo.
                        # Usa bundle_col (nome dal Schema) invece di calc_name (nome dal TR)
                        # per preservare la capitalizzazione originale del Schema.
                        # Es: TR "Enduro pack - Black" (lowercase 'p') vs
                        #     Schema "Enduro Pack - Black" (uppercase 'P') -> usa Schema.
                        # Per bundle senza varianti (es: "Touring Pack") il risultato
                        # è lo stesso perché verrà normalizzato a "Yes" dalla logica sotto.
                        value = bundle_col
                        break
                else:
                    # -----------------------------------------------------------
                    # LOGICA CARATTERISTICA (V2):
                    # - Usa il readable INTEGRALE (con eventuali parentesi).
                    #
                    # NOTA: NON applicare _remove_parentheses_variant qui.
                    # Il simulatore MC usa i nomi opzione COMPLETI dal TR
                    # (es: "SUSPENSION - Ohlins Smart EC (electronic)").
                    # Dopo normalizzazione nel matching, il valore diventa
                    # "ohlins smart ec(electronic)". Se togliessimo le
                    # parentesi, il valore in unique_rows sarebbe
                    # "ohlins smart ec" -> MISMATCH -> 0 risultati per SKU
                    # con opzioni del tipo "Ohlins Smart EC (electronic)".
                    # -----------------------------------------------------------
                    value = readable_col
                    break  # primo match valido vince

        # ---------------------------------------------------------------------
        # NORMALIZZAZIONE FINALE (V2):
        # Per i bundle, se il valore non è "No" o "Not" -> normalizza a "Yes"
        # Nessuna keyword hardcoded: usa is_bundle
        #
        # REGOLA: TUTTI i bundle diventano "Yes" se attivi, indipendentemente
        # dal nome (anche quelli con varianti come "Enduro pack - Black").
        # Le caratteristiche categoriali (Fairing Color, Subframe, ecc.) hanno
        # is_bundle=False e NON vengono normalizzate -> mantengono i valori specifici.
        # ---------------------------------------------------------------------
        if is_bundle and value not in ['No', 'Not']:
            value = 'Yes'

        row[-(calc_offset - i)] = value

    return True  # riga inclusa nella matrice


# ============================================================================
# ALLINEAMENTO VALORI CARATTERISTICHE CON NOMI OPZIONE TR
# ============================================================================

def _build_tr_char_options(tr_file: str) -> Dict[str, List[str]]:
    """
    Legge dal TR i nomi opzione per ogni gruppo caratteristica.

    Ritorna: {group_name: [option_name, ...]}
    Es: {"FAIRING COLOR": ["Viola Blast gloss", "Rosso Ducati", ...]}

    Usato per allineare i valori scritti nella matrice con i nomi
    esatti del TR (che coincidono con quelli generati dalla simulazione).
    """
    try:
        xl = pd.ExcelFile(tr_file)
        sheets = xl.sheet_names

        from mid_term_supermodels.tr import _parse_month_str as _pms
        opt_sheets = sorted(
            [s for s in sheets if s.lower().startswith('tr_optional_')],
            key=lambda s: _pms(s[len('tr_optional_'):])
        )
        if not opt_sheets:
            return {}

        tr_df = pd.read_excel(tr_file, sheet_name=opt_sheets[0], header=0)

        concat_col = None
        for col in tr_df.columns:
            if col and 'concat' in str(col).lower():
                concat_col = col
                break
        if concat_col is None:
            concat_col = tr_df.columns[1]

        options: Dict[str, List[str]] = {}
        for val in tr_df[concat_col]:
            if not val or pd.isna(val):
                continue
            val_str = str(val).strip()
            parts = val_str.split(' - ', 1)
            if len(parts) < 2:
                continue
            group = parts[0].strip()
            if group.upper() == 'BUNDLE':
                continue
            option = parts[1].strip()          # es. "Viola Blast gloss"
            if group not in options:
                options[group] = []
            if option not in options[group]:
                options[group].append(option)

        total = sum(len(v) for v in options.values())
        log.info('[V2] TR char options: %s gruppi, %s opzioni totali', len(options), total)
        return options

    except Exception as e:
        log.warning('[WARN] _build_tr_char_options: %s -> allineamento TR disabilitato', e)
        return {}


def _align_to_tr_option(value: str, group: str,
                         tr_char_options: Dict[str, List[str]]) -> str:
    """
    Cerca il nome opzione TR che meglio corrisponde al readable della Mappatura.

    LOGICA (in ordine):
    1. Match esatto case-insensitive  -> ritorna il nome TR con la sua capitalizzazione
    2. Prefix match (1 candidato)     -> es. "Viola Blast" -> "Viola Blast gloss"
       Sicuro SOLO se esiste UN SOLO candidato: evita falsi positivi quando
       esistono sia "Blue gloss" che "Blue matte" (in quel caso si usa il valore originale).
    3. Nessun match / ambiguo         -> ritorna il valore originale invariato
    """
    options = tr_char_options.get(group, [])
    if not options:
        return value

    value_norm = value.lower().strip()

    # 1. Exact match
    for opt in options:
        if opt.lower().strip() == value_norm:
            return opt

    # 2. Prefix match (unambiguous)
    prefix = value_norm + ' '
    matches = [opt for opt in options if opt.lower().strip().startswith(prefix)]
    if len(matches) == 1:
        return matches[0]

    # 3. Fallback: valore originale
    return value


# ============================================================================
# FILTRO CV ATTIVI (TR > 0)
# ============================================================================

def _build_active_cv_sets(tr_file: str, mapping_df: pd.DataFrame) -> Dict[str, Set[str]]:
    """
    Costruisce {CHAR_code: {VALUE_code, ...}} con solo i CV la cui opzione TR
    ha TR > 0 in almeno un modello.

    Usa parse_tr_file() da tr_parser_V2, che:
    - Salta già interi gruppi con _all_tr_zero (TR=0 su tutti i modelli/opzioni)
    - Mantiene però le singole opzioni TR=0 dentro un gruppo altrimenti attivo

    Questo filtro agisce al livello di singola opzione all'interno del gruppo:
    se un'opzione ha TR=0 su TUTTI i modelli -> esclusa dalla espansione.

    Caratteristiche non presenti nel TR -> tutti i CV inclusi (conservativo).
    In caso di errore ritorna {} -> filtro disabilitato, comportamento invariato.
    """
    try:
        from mid_term_supermodels.tr import parse_tr_file

        _, characteristics = parse_tr_file(tr_file)

        # Per ogni CharacteristicGroup già parsato, raccoglie i nomi delle opzioni
        # con TR > 0 in almeno un modello.
        # active_options: {group_name_lower: {option_name_lower, ...}}
        active_options: Dict[str, Set[str]] = {}
        n_zero = 0

        for char_name, char_group in characteristics.items():
            g_lower = char_name.lower()
            active_options[g_lower] = set()

            for i, concat_val in enumerate(char_group.values):
                # concat_val = "FAIRING COLOR - Viola Blast gloss"
                parts = str(concat_val).split(' - ', 1)
                if len(parts) < 2:
                    continue
                option_name = parts[1].strip()

                # TR > 0 in almeno un modello?
                has_tr = any(
                    float(tr_arr[i]) > 0
                    for tr_arr in char_group.tr.values()
                    if i < len(tr_arr)
                )
                if has_tr:
                    active_options[g_lower].add(option_name.lower())
                else:
                    n_zero += 1

        n_active_opts = sum(len(v) for v in active_options.values())
        log.info('[V2] Opzioni TR attive: %s | scartate TR=0: %s', n_active_opts, n_zero)

        # Perf: itera mapping_df in un solo passaggio con to_dict('records')
        # e costruisce sia chars_in_tr sia active_cv_sets in un loop.
        chars_in_tr: Set[str] = set()
        active_cv_sets: Dict[str, Set[str]] = {}
        mapping_records = mapping_df.to_dict('records')
        for mrow in mapping_records:
            if str(mrow['Descrizione']).strip().lower() in active_options:
                chars_in_tr.add(str(mrow['CHAR']).strip())

        for mrow in mapping_records:
            char_code  = str(mrow['CHAR']).strip()
            value_code = str(mrow['VALUE']).strip()
            char_desc  = str(mrow['Descrizione']).strip().lower()
            readable   = str(mrow['Readable']).strip().lower()

            if char_code not in active_cv_sets:
                active_cv_sets[char_code] = set()

            if char_code not in chars_in_tr:
                # Caratteristica non nel TR -> includi per sicurezza (conservativo)
                active_cv_sets[char_code].add(value_code)
                continue

            if readable in active_options.get(char_desc, set()):
                active_cv_sets[char_code].add(value_code)
            # else: opzione con TR=0 su tutti i modelli -> esclusa

        total_active = sum(len(v) for v in active_cv_sets.values())
        log.info('[V2] CV attivi (TR > 0): %s/%s', total_active, len(mapping_df))
        return active_cv_sets

    except Exception as e:
        log.warning('[WARN] _build_active_cv_sets: %s -> filtro TR disabilitato', e)
        return {}


# ============================================================================
# FILTRO STATI (market-specific)
# ============================================================================

def _load_states_set(states_file: str) -> set:
    """
    Carica l'elenco degli stati/mercati da States.xlsx.
    Ritorna un set di codici (es: {'ARG', 'AUS', 'BRA', 'CAL', 'CDN', ...}).
    Se il file non esiste o non è leggibile, ritorna un set vuoto (filtro disabilitato).
    """
    if not states_file:
        log.info('[V2] States file non specificato -> filtro stati disabilitato')
        return set()
    try:
        df = pd.read_excel(states_file, header=0)
        col = df.columns[0]
        states = {str(v).strip().upper() for v in df[col] if v and not pd.isna(v)}
        log.info('[V2] States caricati: %s', sorted(states))
        return states
    except Exception as e:
        log.warning('[WARN] _load_states_set: %s -> filtro stati disabilitato', e)
        return set()


def _is_state_specific(condition_text: str, states_set: set) -> bool:
    """
    Ritorna True se la condizione espansa contiene $root.ZCOUNTRY eq 'STATE'
    per almeno uno STATE presente in states_set.
    Queste righe sono market-specific e vanno escluse dalla matrice globale.

    Esempi che ritornano True (riga scartata):
        $root.ZCOUNTRY eq 'CDN'
        $root.ZPRODGRP1 eq 'MSV4' AND $root.ZCOUNTRY eq 'ARG'

    Esempi che ritornano False (riga mantenuta):
        NOT ($root.ZCOUNTRY eq 'CDN')   ← vale per tutti tranne CDN
        $root.ZCOUNTRY ne 'CDN'         ← idem
        $root.ZPRODGRP1 eq 'MSV4'       ← nessuna clausola paese
    """
    if not condition_text or not states_set:
        return False
    # Bugfix: rimuovi i blocchi NOT(...) prima del match, altrimenti
    # "NOT ($root.ZCOUNTRY eq 'CDN')" verrebbe erroneamente classificato come state-specific.
    # Strip ricorsivo di NOT( ... ) bilanciato sui paren.
    _cleaned = condition_text
    _prev = None
    while _cleaned != _prev:
        _prev = _cleaned
        _cleaned = re.sub(
            r'NOT\s*\([^()]*\)', '', _cleaned, flags=re.IGNORECASE
        )
    for match in re.finditer(
        r'\$root\.ZCOUNTRY\s+eq\s+["\']([^"\']+)["\']',
        _cleaned, re.IGNORECASE
    ):
        if match.group(1).strip().upper() in states_set:
            return True
    return False


def _check_residual_quality(res_df: pd.DataFrame,
                             residual_col: str,
                             component_col: str) -> List[Dict]:
    """
    Diagnostica la qualità del file 'quantity e residual'.
    Stampa un report a console e restituisce una lista di dict con i problemi
    rilevati (usata da _write_quality_report per generare l'Excel).

    Severità:
        ERRORE — residual/quantity null, negativi
        INFO   — residual = 0 (scartato dal filtro, atteso)
        WARN   — quantity null o zero
    """
    log.info("\\n  --- QUALITA' INPUT: file residual/quantity ---")
    log.info('  Totale componenti nel file: %s', len(res_df))

    # ── Colonne ───────────────────────────────────────────────────────────────
    res_series = res_df[residual_col]
    mask_res_null = res_series.isna()
    mask_res_zero = (~mask_res_null) & (res_series == 0)
    mask_res_neg  = (~mask_res_null) & (res_series < 0)

    qty_col = next((c for c in res_df.columns
                    if c != component_col and c != residual_col and (
                    'quant' in c.lower() or 'qty' in c.lower()
                    or 'qt' in c.lower())), None)
    desc_col = next((c for c in res_df.columns
                     if 'testo' in c.lower() or 'descr' in c.lower()
                     or 'text' in c.lower()), None)

    if qty_col:
        qty_series = pd.to_numeric(res_df[qty_col], errors='coerce')
        mask_qty_null = qty_series.isna()
        mask_qty_zero = (~mask_qty_null) & (qty_series == 0)
        mask_qty_neg  = (~mask_qty_null) & (qty_series < 0)
    else:
        mask_qty_null = pd.Series([False] * len(res_df), index=res_df.index)
        mask_qty_zero = mask_qty_null.copy()
        mask_qty_neg  = mask_qty_null.copy()
        qty_series    = pd.Series([None] * len(res_df), index=res_df.index)

    # ── Raccolta dati per Excel ────────────────────────────────────────────────
    checks = [
        (mask_res_null, "Residual NULL/non numerico", "ERRORE"),
        (mask_res_neg,  "Residual < 0 (anomalia)",   "ERRORE"),
        (mask_qty_null, "Quantity NULL/non numerico", "WARN"),
        (mask_qty_zero, "Quantity = 0",               "WARN"),
        (mask_qty_neg,  "Quantity < 0 (anomalia)",    "ERRORE"),
        (mask_res_zero, "Residual = 0 (scartato)",    "INFO"),
    ]

    issues = []
    has_problems = False

    for mask, label, severity in checks:
        n = int(mask.sum())
        if n == 0:
            continue
        if severity in ('ERRORE', 'WARN'):
            has_problems = True

        codes = res_df.loc[mask, component_col].dropna().astype(str).tolist()
        preview = ', '.join(codes[:8]) + (f' ... (+{len(codes)-8} altri)' if len(codes) > 8 else '')
        log.info('  [%s] %s: %s componenti', severity, label, n)
        log.info('           Codici: %s', preview)

        for idx in res_df.index[mask]:
            row = res_df.loc[idx]
            issues.append({
                'Severita':    severity,
                'Problema':    label,
                'Codice':      str(row.get(component_col, '')),
                'Descrizione': str(row[desc_col]) if desc_col else '',
                'Residual':    row.get(residual_col, None),
                'Quantity':    qty_series.loc[idx] if qty_col else None,
            })

    if not has_problems:
        log.info('  [OK] Nessun problema ERRORE/WARN nel file residual/quantity.')
    log.info("  " + "-" * 55)
    return issues


def _check_price_quality(res_df: pd.DataFrame,
                          residual_col: str,
                          component_col: str,
                          prezzi_file: str,
                          prezzi_sheet: str = 'Purch Docs 26-2',
                          price_col_hint: str = 'Net price for unit') -> List[Dict]:
    """
    Verifica i prezzi dei componenti attivi (residual > 0) nel file PREZZI.xlsx.

    Per ogni codice componente attivo, controlla che esista un prezzo valido
    nel listino acquisti.

    Severità:
        ERRORE — prezzo NULL/mancante, prezzo = 0, o codice non trovato in PREZZI.xlsx
    """
    import os as _os
    log.info("\\n  --- QUALITA' INPUT: verifica prezzi componenti ---")

    if not _os.path.isfile(prezzi_file):
        log.info('  [SKIP] File prezzi non trovato: %s', prezzi_file)
        return []

    # Codici attivi: residual > 0
    _res = res_df.copy()
    _res[residual_col] = pd.to_numeric(_res[residual_col], errors='coerce')
    active_codes = set(
        _res.loc[_res[residual_col] > 0, component_col].dropna().astype(str)
    )
    log.info('  Componenti attivi (residual > 0): %s', len(active_codes))

    if not active_codes:
        log.info('  [SKIP] Nessun componente attivo — skip verifica prezzi.')
        return []

    # Carica PREZZI.xlsx
    try:
        pz_df = pd.read_excel(prezzi_file, sheet_name=prezzi_sheet, header=2)
    except Exception as e:
        log.warning('  [WARN] Impossibile caricare %s: %s', prezzi_file, e)
        return []

    # Individua colonna codice Materiale
    mat_col = next((c for c in pz_df.columns
                    if 'materiale' in str(c).lower()), None)
    if mat_col is None:
        log.warning("  [WARN] Colonna 'Materiale' non trovata in PREZZI.xlsx")
        return []

    # Individua colonna prezzo
    price_col = next((c for c in pz_df.columns
                      if price_col_hint.lower() in str(c).lower()), None)
    if price_col is None:
        log.warning("  [WARN] Colonna prezzo ('%s...') non trovata in PREZZI.xlsx", price_col_hint)
        return []

    # Colonna descrizione (opzionale)
    desc_col = next((c for c in pz_df.columns
                     if any(k in str(c).lower()
                            for k in ('testo', 'descr', 'text', 'breve'))), None)

    pz_df[mat_col]   = pz_df[mat_col].astype(str).str.strip()
    pz_df[price_col] = pd.to_numeric(pz_df[price_col], errors='coerce')

    # Filtra solo righe relative ai codici attivi
    pz_active = pz_df[pz_df[mat_col].isin(active_codes)].copy()

    issues = []
    has_problems = False

    # ── Caso 1: Prezzo NULL ───────────────────────────────────────────────────
    mask_null = pz_active[price_col].isna()
    n_null    = int(mask_null.sum())
    if n_null > 0:
        has_problems = True
        codes_null = pz_active.loc[mask_null, mat_col].tolist()
        preview = (', '.join(codes_null[:8])
                   + (f' ... (+{len(codes_null)-8} altri)' if len(codes_null) > 8 else ''))
        log.error('  [ERRORE] Prezzo NULL/assente: %s componenti', n_null)
        log.info('           Codici: %s', preview)
        for idx in pz_active.index[mask_null]:
            row = pz_active.loc[idx]
            issues.append({
                'Severita':    'ERRORE',
                'Problema':    'Prezzo assente (NULL)',
                'Codice':      str(row[mat_col]),
                'Descrizione': str(row[desc_col]) if desc_col else '',
                'Prezzo (€)':  None,
            })

    # ── Caso 2: Prezzo = 0 ───────────────────────────────────────────────────
    mask_zero = (~mask_null) & (pz_active[price_col] == 0)
    n_zero    = int(mask_zero.sum())
    if n_zero > 0:
        has_problems = True
        codes_zero = pz_active.loc[mask_zero, mat_col].tolist()
        preview = (', '.join(codes_zero[:8])
                   + (f' ... (+{len(codes_zero)-8} altri)' if len(codes_zero) > 8 else ''))
        log.error('  [ERRORE] Prezzo = 0: %s componenti', n_zero)
        log.info('           Codici: %s', preview)
        for idx in pz_active.index[mask_zero]:
            row = pz_active.loc[idx]
            issues.append({
                'Severita':    'ERRORE',
                'Problema':    'Prezzo = 0',
                'Codice':      str(row[mat_col]),
                'Descrizione': str(row[desc_col]) if desc_col else '',
                'Prezzo (€)':  0,
            })

    # ── Caso 3: Codici attivi non presenti nel file PREZZI ───────────────────
    codes_in_prezzi = set(pz_df[mat_col])
    codes_not_found = sorted(active_codes - codes_in_prezzi)
    n_missing       = len(codes_not_found)
    if n_missing > 0:
        has_problems = True
        preview = (', '.join(codes_not_found[:8])
                   + (f' ... (+{n_missing-8} altri)' if n_missing > 8 else ''))
        log.error('  [ERRORE] Codici attivi non trovati in PREZZI: %s', n_missing)
        log.info('           Codici: %s', preview)
        for code in codes_not_found:
            issues.append({
                'Severita':    'ERRORE',
                'Problema':    'Codice non trovato in PREZZI.xlsx',
                'Codice':      code,
                'Descrizione': '',
                'Prezzo (€)':  None,
            })

    if not has_problems:
        log.info('  [OK] Tutti i componenti attivi hanno un prezzo valido in PREZZI.xlsx.')
    log.info("  " + "-" * 55)
    return issues


def _check_ct_cv_traceability(bom_data: pd.DataFrame,
                               condition_col: str,
                               parser,
                               bom_file: str = '') -> Dict:
    """
    Scansiona tutte le condizioni BOM filtrate e verifica che ogni CT e CV
    presente nelle dipendenze sia tracciato nel mapping (Mappatura Schema).

    Restituisce un dict con tre liste di dict (usato da _write_quality_report):
        'ct_missing'     — CT non trovati nel mapping (ERRORE)
        'cv_missing'     — CV non trovati per CT noti (WARN)
        'ct_system'      — CT di sistema non espansi in matrice (INFO)

    Args:
        bom_data: DataFrame BOM
        condition_col: colonna condizioni
        parser: ConditionParser
        bom_file: nome file BOM per tracciabilità (es: BOM_150V21E.xlsx)
    """
    SYSTEM_CT = {'ZPRODGRP1', 'ZPRODGRP', 'ZCOUNTRY', 'ZCOL_RUOTE',
                 'ZCOL', 'ZRIMS', 'ZMATERIAL', 'ZPLT', 'ZMYEAR', 'ZSUPM'}

    log.info("\\n  --- TRACCIABILITA' CT/CV: condizioni vs Mappatura Schema ---")

    mapping_ct = set(parser.mapping_dict.keys())
    unknown_ct        = {}   # {ct: {'count': n, 'esempi': [...]}}
    unknown_ct_system = {}   # {ct: n}
    unknown_cv        = {}   # {(ct, cv): {'count': n, 'esempi': [...]}}

    conditions = bom_data[condition_col].dropna()
    n_conditions = len(conditions)

    for cond in conditions:
        cond_str = str(cond)

        ct_found = set(m.group(1) for m in re.finditer(r'\$root\.(\w+)', cond_str))
        for ct in ct_found:
            if ct not in mapping_ct:
                if ct.upper() in SYSTEM_CT or not ct.startswith('CT'):
                    unknown_ct_system[ct] = unknown_ct_system.get(ct, 0) + 1
                else:
                    if ct not in unknown_ct:
                        unknown_ct[ct] = {'count': 0, 'esempi': []}
                    unknown_ct[ct]['count'] += 1
                    if len(unknown_ct[ct]['esempi']) < 3:
                        unknown_ct[ct]['esempi'].append(cond_str[:120])

        for m in re.finditer(
            r'\$root\.(\w+)\s+(?:eq|ne)\s+[\'"]([^\'"]+)[\'"]',
            cond_str, re.IGNORECASE
        ):
            ct, cv = m.group(1), m.group(2)
            if ct in mapping_ct and cv not in parser.mapping_dict.get(ct, {}):
                key = (ct, cv)
                if key not in unknown_cv:
                    unknown_cv[key] = {'count': 0, 'esempi': []}
                unknown_cv[key]['count'] += 1
                if len(unknown_cv[key]['esempi']) < 3:
                    unknown_cv[key]['esempi'].append(cond_str[:120])

        for m in re.finditer(
            r'\$root\.(\w+)\s+IN\s*\(([^)]+)\)', cond_str, re.IGNORECASE
        ):
            ct = m.group(1)
            cv_list = re.findall(r"['\"]([^'\"]+)['\"]", m.group(2))
            if ct in mapping_ct:
                for cv in cv_list:
                    if cv not in parser.mapping_dict.get(ct, {}):
                        key = (ct, cv)
                        if key not in unknown_cv:
                            unknown_cv[key] = {'count': 0, 'esempi': []}
                        unknown_cv[key]['count'] += 1
                        if len(unknown_cv[key]['esempi']) < 3:
                            unknown_cv[key]['esempi'].append(cond_str[:120])

    log.info('  Condizioni analizzate: %s', n_conditions)

    if not unknown_ct and not unknown_cv:
        log.info('  [OK] Tutti i CT/CV nelle condizioni sono tracciati nel mapping.')
    else:
        if unknown_ct:
            log.error('  [ERRORE] CT non trovati nel mapping (%s distinti):', len(unknown_ct))
            for ct, d in sorted(unknown_ct.items(), key=lambda x: -x[1]['count']):
                log.info('    - %-25s  in %s condizioni', ct, d['count'])
        if unknown_cv:
            log.error('  [ERRORE] CV non trovati nel mapping (%s distinti):', len(unknown_cv))
            for (ct, cv), d in sorted(unknown_cv.items(), key=lambda x: -x[1]['count']):
                log.info('    - %-20s / %-20s  in %s condizioni', ct, cv, d['count'])

    if unknown_ct_system:
        log.info('  [INFO] CT di sistema/filtro: %s', sorted(unknown_ct_system.keys()))
    log.info("  " + "-" * 55)

    # ── Serializza per Excel ───────────────────────────────────────────────────
    ct_rows = [
        {'Severita': 'ERRORE', 'CT': ct, 'BOM': bom_file,
         'N condizioni': d['count'],
         'Esempio condizione 1': d['esempi'][0] if len(d['esempi']) > 0 else '',
         'Esempio condizione 2': d['esempi'][1] if len(d['esempi']) > 1 else '',
         'Esempio condizione 3': d['esempi'][2] if len(d['esempi']) > 2 else ''}
        for ct, d in sorted(unknown_ct.items(), key=lambda x: -x[1]['count'])
    ]
    cv_rows = [
        {'Severita': 'ERRORE', 'CT': ct, 'CV': cv, 'BOM': bom_file,
         'N condizioni': d['count'],
         'Esempio condizione 1': d['esempi'][0] if len(d['esempi']) > 0 else '',
         'Esempio condizione 2': d['esempi'][1] if len(d['esempi']) > 1 else '',
         'Esempio condizione 3': d['esempi'][2] if len(d['esempi']) > 2 else ''}
        for (ct, cv), d in sorted(unknown_cv.items(), key=lambda x: -x[1]['count'])
    ]
    sys_rows = [
        {'Severita': 'INFO', 'CT': ct, 'N condizioni': n,
         'Nota': 'CT di sistema — usato per filtraggio, non espanso in colonne matrice'}
        for ct, n in sorted(unknown_ct_system.items(), key=lambda x: -x[1])
    ]
    return {
        'ct_missing':    ct_rows,
        'cv_missing':    cv_rows,
        'ct_system':     sys_rows,
        # set grezzi per il filtro pipeline
        'unknown_ct_set': set(unknown_ct.keys()),
        'unknown_cv_set': set(unknown_cv.keys()),   # {(ct, cv), ...}
    }


def _write_quality_report(residual_issues: List[Dict],
                           traceability: Dict,
                           price_issues: List[Dict] = None,
                           bom_missing: List[Dict] = None,
                           output_path: str = '.\\Output\\quality_check_report.xlsx') -> None:
    """
    Genera un file Excel con il report di qualità degli input.

    Fogli:
        1. Residual & Quantity  — componenti con problemi nel file residual
        2. BOM Senza Residual   — componenti BOM assenti da quantity e residual.xlsx
        3. CT non tracciati     — CT nelle dipendenze assenti dalla Mappatura
        4. CV non tracciati     — CV nelle dipendenze non mappati per il CT noto
        5. CT di sistema        — CT di filtro (INFO, non sono errori)
        6. Prezzi               — componenti attivi con prezzo mancante o = 0 (opzionale)
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    SEV_COLORS = {'ERRORE': 'FFC7CE', 'WARN': 'FFEB9C', 'INFO': 'D9EAD3'}
    SEV_FONTS  = {'ERRORE': 'C00000', 'WARN': '7F6000', 'INFO': '274E13'}
    H_FILL  = PatternFill('solid', fgColor='1F3864')
    H_FONT  = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    N_FONT  = Font(name='Calibri', size=10)
    B_FONT  = Font(name='Calibri', bold=True, size=10)
    CENTER  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT    = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    def _border():
        """Bordo sottile grigio per le celle del report Excel qualita'."""
        s = Side(style='thin', color='BFBFBF')
        return Border(left=s, right=s, top=s, bottom=s)

    def _write_sheet(ws, title: str, subtitle: str,
                     headers: List[str], rows: List[Dict],
                     col_widths: List[int], sev_col: str = 'Severita'):
        """Scrive un foglio formattato del report qualita' (titolo, intestazioni,
        righe colorate per severita', larghezze colonne)."""
        ws.sheet_view.showGridLines = False

        # Titolo
        ws.merge_cells(f'A1:{chr(64+len(headers))}1')
        ws['A1'] = title
        ws['A1'].font      = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
        ws['A1'].fill      = H_FILL
        ws['A1'].alignment = CENTER
        ws.row_dimensions[1].height = 28

        # Sottotitolo
        ws.merge_cells(f'A2:{chr(64+len(headers))}2')
        ws['A2'] = subtitle
        ws['A2'].font      = Font(name='Calibri', italic=True, size=10, color='FFFFFF')
        ws['A2'].fill      = PatternFill('solid', fgColor='2E4F8A')
        ws['A2'].alignment = CENTER
        ws.row_dimensions[2].height = 18

        # Intestazioni
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=c, value=h)
            cell.font = H_FONT; cell.fill = H_FILL
            cell.alignment = CENTER; cell.border = _border()
        ws.row_dimensions[3].height = 36

        if not rows:
            ws.merge_cells(f'A4:{chr(64+len(headers))}4')
            ok = ws['A4']
            ok.value = 'Nessun problema rilevato'
            ok.font  = Font(name='Calibri', italic=True, size=10, color='27AE60')
            ok.alignment = CENTER
            ok.fill = PatternFill('solid', fgColor='D9EAD3')
        else:
            for r, row_data in enumerate(rows, 4):
                sev = str(row_data.get(sev_col, 'INFO'))
                bg  = SEV_COLORS.get(sev, 'FFFFFF')
                fc  = SEV_FONTS.get(sev, '000000')
                for c, h in enumerate(headers, 1):
                    val  = row_data.get(h, '')
                    cell = ws.cell(row=r, column=c, value=val)
                    cell.border    = _border()
                    cell.alignment = CENTER if c <= 3 else LEFT
                    if h == sev_col:
                        cell.fill = PatternFill('solid', fgColor=bg)
                        cell.font = Font(name='Calibri', bold=True,
                                         size=10, color=fc)
                    else:
                        cell.fill = PatternFill('solid', fgColor='F9F9F9'
                                                if r % 2 == 0 else 'FFFFFF')
                        cell.font = N_FONT
                ws.row_dimensions[r].height = 18

        from openpyxl.utils import get_column_letter
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
    wb = Workbook()

    # ── Foglio 1: Residual & Quantity ────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Residual & Quantity'
    _write_sheet(
        ws1,
        title    = 'QUALITA\' INPUT — Residual & Quantity',
        subtitle = ('ERRORE = da correggere urgentemente  |  WARN = verificare  |  '
                    'INFO = residual=0 (componente incluso in BOM ma senza residual attivo)'),
        headers  = ['Severita', 'Problema', 'Codice', 'Descrizione',
                    'Residual (mesi)', 'Quantity'],
        rows     = residual_issues,
        col_widths = [10, 28, 16, 36, 16, 12],
    )

    # ── Foglio 2: BOM Senza Residual ─────────────────────────────────────────
    ws_bom = wb.create_sheet('BOM Senza Residual')
    _write_sheet(
        ws_bom,
        title    = 'BOM SENZA RESIDUAL — componenti BOM assenti da quantity e residual.xlsx',
        subtitle = ('Questi componenti sono presenti nella BOM con una dipendenza ma non hanno '
                    'dati di lead time/residual. Safety Stock = 0. Aggiungere al file residual.'),
        headers  = ['Numero componenti', 'Descrizione', 'Blocco dipendenze'],
        rows     = (bom_missing or []),
        col_widths = [20, 40, 80],
        sev_col  = 'Numero componenti',   # nessuna colonna severità, usa codice
    )

    # ── Foglio 3: CT non tracciati ───────────────────────────────────────────
    ws2 = wb.create_sheet('CT non tracciati')
    _write_sheet(
        ws2,
        title    = 'CT NON TRACCIATI — presenti in dipendenze BOM ma assenti dalla Mappatura',
        subtitle = ('Questi CT non vengono espansi in colonne della matrice. '
                    'Aggiungere alla Mappatura Schema oppure verificare se sono obsoleti.'),
        headers  = ['Severita', 'CT', 'BOM', 'N condizioni',
                    'Esempio condizione 1', 'Esempio condizione 2', 'Esempio condizione 3'],
        rows     = traceability.get('ct_missing', []),
        col_widths = [10, 14, 20, 14, 55, 55, 55],
    )

    # ── Foglio 4: CV non tracciati ───────────────────────────────────────────
    ws3 = wb.create_sheet('CV non tracciati')
    _write_sheet(
        ws3,
        title    = 'CV NON TRACCIATI — valori presenti in dipendenze ma non mappati per il CT',
        subtitle = ('Il CT esiste nella Mappatura ma questo specifico valore (CV) non e\' definito. '
                    'Aggiungere l\'opzione alla riga CT corrispondente nella Mappatura Schema.'),
        headers  = ['Severita', 'CT', 'CV', 'BOM', 'N condizioni',
                    'Esempio condizione 1', 'Esempio condizione 2', 'Esempio condizione 3'],
        rows     = traceability.get('cv_missing', []),
        col_widths = [10, 14, 16, 20, 14, 50, 50, 50],
    )

    # ── Foglio 5: CT di sistema ──────────────────────────────────────────────
    ws4 = wb.create_sheet('CT di sistema')
    _write_sheet(
        ws4,
        title    = 'CT DI SISTEMA — usati per filtraggio, non espansi in matrice',
        subtitle = 'Questi codici sono attesi e non richiedono azione.',
        headers  = ['Severita', 'CT', 'N condizioni', 'Nota'],
        rows     = traceability.get('ct_system', []),
        col_widths = [10, 18, 14, 60],
    )

    # ── Foglio 6: Prezzi (opzionale) ─────────────────────────────────────────
    if price_issues is not None:
        ws5 = wb.create_sheet('Prezzi')
        _write_sheet(
            ws5,
            title    = 'QUALITA\' PREZZI — Componenti attivi con prezzo mancante o non valido',
            subtitle = ('ERRORE = prezzo NULL, prezzo = 0, o codice non trovato in PREZZI.xlsx. '
                        'Correggere prima di eseguire l\'analisi Cycle Stock.'),
            headers  = ['Severita', 'Problema', 'Codice', 'Descrizione', 'Prezzo (€)'],
            rows     = price_issues,
            col_widths = [10, 32, 16, 40, 12],
        )

    wb.save(output_path)
    log.info("\\n  [OK] Report qualita' salvato: %s", output_path)


# ============================================================================
# ENRICHMENT BOM GREZZO → BOM_150VER.xlsx  (integrato da bom_enrich.py)
# ============================================================================

_DEP_COL_ENRICH = "Blocco dati per archivio dipendenze"


def _bom_to_str(v) -> str:
    """NaN/None → ''. Float interi (30.0) → '30'. Altrimenti str.strip()."""
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _classify_rule(C_i: str, D_i: str, O_i: str) -> str:
    """
    In parole semplici: decide la "categoria" di una riga BOM guardando 3 campi
    (Tipo approvv. = C, Approvv. speciale = AS, Merce sfusa = O). La categoria
    (R0..R4) serve poi a stabilire se la riga e i suoi sotto-componenti vanno
    mostrati o nascosti.

    Classifica riga secondo 5 regole business:
      R0 - sfusa (O=X)
      R1 - F + AS!=30
      R2 - E + AS!=30
      R3 - N.DEF ('') + AS!=30
      R4 - F + AS=30
    """
    if O_i == "X":                return "R0"
    if C_i == "F" and D_i != "30": return "R1"
    if C_i == "E" and D_i != "30": return "R2"
    if C_i == "" and D_i != "30":  return "R3"
    if C_i == "F" and D_i == "30": return "R4"
    return "OTHER"


def _apply_mostra_nascondi(lev, tipo, approv, sfusa_bom):
    """
    In parole semplici: scorre la distinta dall'alto verso il basso e decide per
    ogni componente se "Mostra" (entra nel calcolo) o "Nascondi" (escluso). Una
    riga "merce sfusa" nasconde se stessa e tutto il suo sotto-albero; le altre
    seguono il loro tipo di approvvigionamento.

    Algoritmo Mostra/Nascondi basato su 5 regole business.

    sfusa_lvl - livello del piu' ampio antenato sfusa attivo.
                Mai sovrascritto da discendenti sfusa (fix bug J-overwrite
                del vecchio algoritmo H-L).
    r1_lvl    - livello del piu' ampio antenato R1 attivo.
                Blocca TUTTI i discendenti. Si azzera all'uscita dal livello R1.

    Restituisce (check, sfusa_lvl_arr, r1_lvl_arr).
    """
    n = len(lev)
    check         = [""] * n
    sfusa_lvl_arr = [0]  * n
    r1_lvl_arr    = [0]  * n
    sfusa_lvl = 0
    r1_lvl    = 0

    for i in range(n):
        G_i = lev[i]; C_i = tipo[i]; D_i = approv[i]; O_i = sfusa_bom[i]
        Hp_sfusa = sfusa_lvl; Hp_r1 = r1_lvl
        rule = _classify_rule(C_i, D_i, O_i)

        if   G_i == 1:                          chk = "Mostra"
        elif Hp_sfusa > 0 and G_i > Hp_sfusa:  chk = "Nascondi"
        elif rule == "R0":                       chk = "Nascondi"
        elif Hp_r1 > 0 and G_i > Hp_r1:        chk = "Nascondi"
        elif rule == "R1":                       chk = "Mostra"
        elif rule == "R2":                       chk = "Nascondi"
        elif rule == "R3":                       chk = "Mostra" if G_i == 2 else "Nascondi"
        elif rule == "R4":                       chk = "Mostra"
        else:                                    chk = "Nascondi"

        if   G_i <= Hp_sfusa: sfusa_lvl = G_i if O_i == "X" else 0
        elif O_i == "X":      sfusa_lvl = Hp_sfusa if Hp_sfusa > 0 else G_i
        else:                 sfusa_lvl = Hp_sfusa

        if   G_i <= Hp_r1: r1_lvl = G_i if rule == "R1" else 0
        elif Hp_r1 > 0:    r1_lvl = Hp_r1
        else:              r1_lvl = G_i if rule == "R1" else 0

        check[i]         = chk
        sfusa_lvl_arr[i] = Hp_sfusa
        r1_lvl_arr[i]    = Hp_r1

    return check, sfusa_lvl_arr, r1_lvl_arr


def _find_raw_file(folder: Path, pattern: str) -> Path:
    """Cerca il primo file che matcha il glob pattern (case-insensitive)."""
    matches = list(folder.glob(pattern))
    if not matches:
        matches = [f for f in folder.iterdir()
                   if pattern.lower().replace("*", "") in f.name.lower()]
    if not matches:
        raise FileNotFoundError(f"Nessun file '{pattern}' in {folder}")
    return matches[0]


def enrich_bom_version(ver: str, in_dir, save: bool = True) -> pd.DataFrame:
    """
    Arricchisce il BOM grezzo per la versione VER in un unico step.

    Legge (da in_dir):
      - BOM150_DVER*.XLSX           (BOM grezzo SAP)
      - DATI_AGGIUNTI*_DVER*.XLSX   (Car.MRP, Tipo approvv., Approvv. speciale)
      - ZIBP_BOM_125*_DVER*.XLSX    (BOM125 con colonna dipendenze)

    Produce:
      - DataFrame arricchito con "Blocco dati per archivio dipendenze" propagato,
        filtrato su Check Generale = Mostra
      - Salva BOM_150VER.xlsx in in_dir se save=True
    """
    in_dir = Path(in_dir)
    log.info('\\n  [ENRICH] Versione: %s', ver)

    bom_file  = _find_raw_file(in_dir, f"BOM*150*D{ver}*.XLSX")         # BOM150_DVER*.XLSX o BOM_150DVER*.XLSX
    dag_file  = _find_raw_file(in_dir, f"DATI*AGGIUNT*D{ver}*.XLSX")   # DATI_AGGIUNTI*_DVER*.XLSX o DATI AGGIUNTIVI*DVER*.XLSX
    b125_file = _find_raw_file(in_dir, f"ZIBP*BOM*125*D{ver}*.XLSX")   # ZIBP_BOM_125*_DVER*.XLSX o ZIBP BOM 125 DVER*.XLSX
    log.info('  [ENRICH]   BOM150 : %s', bom_file.name)
    log.info('  [ENRICH]   DAG    : %s', dag_file.name)
    log.info('  [ENRICH]   BOM125 : %s', b125_file.name)

    bom = pd.read_excel(bom_file)
    dag = pd.read_excel(dag_file)

    # — Join Dati Aggiuntivi ———————————————————————————————————————————————
    dag_sel = (dag[["Materiale", "Car. MRP", "Tipo approvv.",
                    "Approvv. speciale", "Merce sfusa"]]
               .rename(columns={"Merce sfusa": "Merce sfusa (DA)"}))
    bom = (bom.merge(dag_sel, left_on="Numero componenti",
                     right_on="Materiale", how="left")
              .drop(columns=["Materiale"]))

    # O_i (sfusa) per Mostra/Nascondi: usa il flag per-materiale dei Dati
    # Aggiuntivi ("Merce sfusa (DA)"), NON quello per-linea del BOM150 (che
    # viene scartato qui sotto). Cosi' la colonna che decide coincide con quella
    # visibile nell'output.
    sfusa_bom = [_bom_to_str(v) for v in bom["Merce sfusa (DA)"]]
    bom = bom.drop(columns=["Merce sfusa"])
    bom.insert(0, "ID", range(1, len(bom) + 1))
    bom["Livello numerico"] = bom["Liv. esplosione"].apply(
        lambda v: str(v).count(".") if pd.notna(v) else 0
    )

    # — Check Generale (5 regole Mostra/Nascondi) ——————————————————————————
    lev    = bom["Livello numerico"].tolist()
    tipo   = [_bom_to_str(v) for v in bom["Tipo approvv."]]
    approv = [_bom_to_str(v) for v in bom["Approvv. speciale"]]

    check_gen, sfusa_lvl_arr, r1_lvl_arr = _apply_mostra_nascondi(
        lev, tipo, approv, sfusa_bom
    )

    bom["Livello Segnalibro Sfusa"] = sfusa_lvl_arr
    bom["Livello Segnalibro R1"]    = r1_lvl_arr
    bom["Check Generale"]           = check_gen

    # — BOM125: aggrega colonna dipendenze per componente ———————————————————
    bom125 = pd.read_excel(b125_file, header=0)
    bom125_dep: Dict[str, str] = {}
    # Perf: to_dict('records') ~10x più veloce di iterrows() per BOM grandi
    for row in bom125.to_dict('records'):
        comp    = _bom_to_str(row.get("Componente critico", ""))
        dep_raw = _bom_to_str(row.get(_DEP_COL_ENRICH, ""))
        dep_val = re.sub(r'\s*\|\s*', ' OR ', dep_raw).strip() if dep_raw else ""
        if not comp:
            continue
        if comp not in bom125_dep:
            bom125_dep[comp] = dep_val
        elif dep_val and not bom125_dep[comp]:
            bom125_dep[comp] = dep_val
        elif dep_val:
            bom125_dep[comp] += " OR " + dep_val

    bom125_df = pd.DataFrame(list(bom125_dep.items()),
                             columns=["Componente critico", _DEP_COL_ENRICH])

    # — Join dipendenze su tutte le righe (prima del filtro) ————————————————
    bom = bom.merge(bom125_df, left_on="Numero componenti",
                    right_on="Componente critico", how="left")
    bom = bom.drop(columns=["Componente critico"])
    dep_vals = bom.pop(_DEP_COL_ENRICH)
    bom.insert(3, _DEP_COL_ENRICH, dep_vals)
    bom[_DEP_COL_ENRICH] = bom[_DEP_COL_ENRICH].fillna("")

    # — Propagazione dipendenze (prima del filtro) ——————————————————————————
    dep_arr    = bom[_DEP_COL_ENRICH].tolist()
    levels     = bom["Livello numerico"].tolist()
    anchor_dep = "";  anchor_set = False;  propagated = 0
    for i in range(len(bom)):
        lvl = levels[i];  row_dep = dep_arr[i]
        if lvl == 2:
            anchor_dep = row_dep if row_dep else ""; anchor_set = True
        elif lvl == 1:
            pass
        elif anchor_set and not row_dep and anchor_dep:
            dep_arr[i] = anchor_dep; propagated += 1
    bom[_DEP_COL_ENRICH] = dep_arr
    log.info('  [ENRICH]   Dipendenze propagate: %s', propagated)

    # — Salvataggio pre-filtro ————————————————————————————————————————————
    if save:
        pre_file = in_dir / f"BOM_150_PreFilter_{ver}.xlsx"
        bom.to_excel(pre_file, index=False)
        log.info('  [ENRICH]   Salvato pre-filtro  : %s', pre_file.name)

    # — Filtro Check Generale = Mostra ——————————————————————————————————————
    bom = bom[bom["Check Generale"] == "Mostra"].copy()
    bom["ID"] = range(1, len(bom) + 1)
    log.info('  [ENRICH]   Righe dopo filtro   : %s', len(bom))

    # — Salvataggio ————————————————————————————————————————————————————————
    if save:
        out_file = in_dir / f"BOM_150{ver}.xlsx"
        bom.to_excel(out_file, index=False)
        log.info('  [ENRICH]   Salvato             : %s', out_file.name)

    return bom


# ============================================================================
# FUNZIONE PRINCIPALE: ELABORAZIONE BOM -> MATRICE (V2)
# ============================================================================

def process_excel_to_matrix_v2(excel_path: str = None,
                                mapping_file: str = '.\\Input\\V21E\\Mappatura_V21E.xlsx',
                                tr_file: str = '.\\Input\\TR TOTALV21E - Copia.xlsx',
                                data_sheet: str = None,
                                output_file: str = '.\\Input\\matrix_output_V2.xlsx',
                                states_file: str = '.\\Input\\States.xlsx',
                                prezzi_file: str = None,
                                filter_untraced: bool = True,
                                df_input: pd.DataFrame = None,
                                bom_file: str = ''):
    """
    Funzione principale V2: converte BOM150 in matrice espansa senza hardcoding.

    DIFFERENZE RISPETTO A V1:
    -------------------------
    - calculated_cols derivato dal TR (non hardcoded)
    - rider_seat_order e pillion_seat_order derivati dal TR
    - Pattern negativi gestiti dal Bundle "Not XYZ" in Schema
    - Logica bundle unificata (no if/elif speciali)
    - Sostituzione RIMS configurabile via foglio Excel

    Args:
        excel_path:    Path al file BOM150 (opzionale se df_input è fornito)
        df_input:      DataFrame BOM già arricchito (se fornito, salta la lettura da file).
                       Prodotto da enrich_bom_version() per lo step integrato.
        mapping_file:  Path al file Mappatura
        tr_file:       Path al file TR
        data_sheet:    Nome foglio BOM150 (solo se excel_path punta a un xlsm legacy)
        output_file:   Path file output
        states_file:   Path al file States (filtra righe market-specific)
    """
    log.info('\\n=== GESTIONE DIPENDENZE V2 — CARICAMENTO ===')
    log.info('BOM:      %s', excel_path if df_input is None else '(DataFrame in memoria)')
    log.info('Mapping:  %s', mapping_file)
    log.info('TR:       %s', tr_file)
    log.info('Output:   %s', output_file)

    # -------------------------------------------------------------------------
    # FASE 1: Caricamento dati
    # -------------------------------------------------------------------------
    mapping_df = load_mapping_from_file(mapping_file)
    states_set = _load_states_set(states_file)  # es. {'CDN', 'ARG', 'AUS', ...}
    active_cv_sets = _build_active_cv_sets(tr_file, mapping_df)  # CV con TR > 0
    if df_input is not None:
        # Path integrato: DataFrame già arricchito passato direttamente
        df = df_input.copy()
        df = df[df['Blocco dati per archivio dipendenze'].notna() &
                (df['Blocco dati per archivio dipendenze'] != "")]
    elif data_sheet:
        df = pd.read_excel(excel_path, sheet_name=data_sheet, header=2)
        df = df[df['Blocco dati per archivio dipendenze'].notnull()]
    else:
        df = pd.read_excel(excel_path, header=0)
        df = df[df['Blocco dati per archivio dipendenze'].notnull()]

    start_col = 'Blocco dati per archivio dipendenze'
    if start_col not in df.columns:
        raise ValueError(f"Colonna '{start_col}' non trovata nel BOM150")

    # ── Quality check residual: rimosso (file quantity e residual.xlsx eliminato) ─
    # In passato veniva fatto un check sui componenti BOM con residual=0/null/mancanti.
    # Con la rimozione del file, residual_issues e bom_missing_issues sono vuoti.
    residual_issues:    List[Dict] = []
    price_issues:       List[Dict] = []
    bom_missing_issues: List[Dict] = []
    # ─────────────────────────────────────────────────────────────────────────

    bom_data = df.copy()

    # Parser iniziale (senza filtro) usato per la verifica tracciabilità
    parser = ConditionParser(mapping_df, active_cv_sets)

    # ── Verifica tracciabilità CT/CV ──────────────────────────────────────────
    traceability = _check_ct_cv_traceability(bom_data, start_col, parser, bom_file=bom_file)
    # ─────────────────────────────────────────────────────────────────────────

    # ── Genera report qualità Excel ───────────────────────────────────────────
    try:
        _write_quality_report(residual_issues, traceability, price_issues,
                              bom_missing=bom_missing_issues)
    except Exception as _qe:
        log.warning('  [WARN] Impossibile generare quality_check_report.xlsx: %s', _qe)

    # ── Ri-crea parser con filtro CT/CV non tracciati (se abilitato) ─────────
    if filter_untraced:
        _unk_ct = traceability.get('unknown_ct_set', set())
        _unk_cv = traceability.get('unknown_cv_set', set())
        if _unk_ct or _unk_cv:
            log.info('  [Filtro] filter_untraced=True: %s CT e %s coppie CT/CV non tracciati verranno esclusi dalle condizioni BOM.', len(_unk_ct), len(_unk_cv))
            parser = ConditionParser(mapping_df, active_cv_sets,
                                     filter_untraced=True,
                                     unknown_ct=_unk_ct,
                                     unknown_cv=_unk_cv)
    # ─────────────────────────────────────────────────────────────────────────
    # ─────────────────────────────────────────────────────────────────────────

    # -------------------------------------------------------------------------
    # FASE 2: Caricamento dinamico colonne calcolate e ordini di priorità dal TR
    # -------------------------------------------------------------------------
    log.info('\\n=== CARICAMENTO DINAMICO DAL TR ===')

    # Colonne calcolate derivate dal TR (sostituisce la lista hardcoded V1)
    calculated_cols = load_calculated_cols_from_tr(tr_file)

    # Set dei nomi bundle positivi (usato nella logica di normalizzazione)
    bundle_set = {c[2] for c in calculated_cols if c[3] is None}

    # Ordini di priorità derivati dal TR.
    # SOLO per caratteristiche dove un singolo componente BOM può avere più opzioni
    # e serve scegliere la più "importante" (es: RIDER SEAT, PILLION SEAT).
    # Per le altre caratteristiche fisiche (FAIRING COLOR, RIMS, ecc.) non serve
    # l'ordinamento: si prende il primo valore trovato nella riga, come in V1.
    # L'ordine del TR viene usato come priority_order SOLO se la caratteristica
    # ha effettivamente più valori possibili per lo stesso componente BOM.
    priority_orders = {}
    # Caratteristiche che usano esplicitamente la priorità (come rider_seat_order in V1)
    priority_chars = {'RIDER SEAT', 'PILLION SEAT'}
    for c in calculated_cols:
        calc_name, search_key = c[2], c[3]
        if search_key is not None and search_key in priority_chars:
            order = load_priority_order_from_tr(tr_file, search_key)
            if order:
                priority_orders[calc_name] = order

    # Sostituzioni RIMS (da foglio Excel opzionale)
    rims_substitutions = load_rims_substitutions(mapping_file)

    # Pre-calcola mappa CT/CV -> bundle(s) dallo Schema
    # Usata per la logica split/discriminante sui CT/CV condivisi
    cv_bundle_map = build_shared_cv_map(mapping_df)

    # -------------------------------------------------------------------------
    # FASE 3: Costruzione struttura colonne
    # -------------------------------------------------------------------------
    col_structure = [
        ('', '', 'Numero componenti'),
        ('', '', 'Condizione_Originale'),
        ('', '', 'Condizione_Espansa'),
        (start_col, '', ''),
        ('', '', 'Qta_comp_UMC'),
    ]

    for row in mapping_df.itertuples(index=False):
        col_tuple = (row.Bundle, row.Descrizione, row.Readable)
        col_structure.append(col_tuple)

    log.info('[OK] Colonne dopo mapping: %s', len(col_structure))

    # Aggiungi colonne calcolate
    for col_tuple in calculated_cols:
        col_structure.append((col_tuple[0], col_tuple[1], col_tuple[2]))

    log.info('[OK] Colonne totali (incl. calcolate): %s', len(col_structure))

    # -------------------------------------------------------------------------
    # FASE 4: Pre-calcolo indici (O(1) lookup)
    # -------------------------------------------------------------------------
    calc_indices = {}
    mapping_cols_count = len(col_structure) - len(calculated_cols) - 5

    for c in calculated_cols:
        calc_name, search_key = c[2], c[3]
        indices = []

        if search_key is not None and calc_name in priority_orders:
            # Caratteristica con ordine di priorità (es: RIDER SEAT, PILLION SEAT)
            # Costruisce indices nell'ordine di priorità definito nel TR
            # CORREZIONE V2: confronto case-insensitive per gestire divergenze TR/Schema
            for readable_name in priority_orders[calc_name]:
                for idx in range(5, mapping_cols_count + 5):
                    b, d, r = col_structure[idx]
                    r_norm = _remove_parentheses_variant(r)
                    if (r.lower() == readable_name.lower() or
                            r_norm.lower() == readable_name.lower()):
                        indices.append(idx)

        elif search_key is None:
            # Bundle: cerca per bundle name (inclusi "Not NomePack")
            #
            # STRATEGIA DI MATCH (in ordine di priorità):
            # 1. Exact match case-insensitive — copre la maggior parte dei casi
            # 2. Match con rimozione parentesi — gestisce "Touring pack (color variant)" -> "Touring Pack"
            #
            # NOTA IMPORTANTE — perché NON si usa il match parziale a N parole:
            # Il match alle prime 2 parole causava falsi positivi gravi:
            #   • "Touring pack Rally" -> ["touring","pack"] == "Touring Pack" -> SBAGLIATO
            #   • "Enduro pack - Black" -> ["enduro","pack"] == "Enduro pack - Silver" -> SBAGLIATO
            # La rimozione delle parentesi è sicura: "Touring pack (color variant)" diventa
            # "Touring pack" che corrisponde esattamente a "Touring Pack" (case-insensitive),
            # mentre "Touring pack Rally" rimane "Touring pack Rally" -> nessun match spurio.
            not_name = _get_not_bundle_name(calc_name)
            calc_name_lower = calc_name.lower()
            not_name_lower = not_name.lower()
            # Versioni senza parentesi (es: "Touring pack (color variant)" -> "Touring pack")
            calc_name_stripped = re.sub(r'\s*\([^)]*\)', '', calc_name).strip().lower()
            not_name_stripped  = re.sub(r'\s*\([^)]*\)', '', not_name).strip().lower()

            for idx in range(5, mapping_cols_count + 5):
                b, d, r = col_structure[idx]
                b_lower = b.lower()

                # 1. Exact match case-insensitive
                if b_lower == calc_name_lower or b_lower == not_name_lower:
                    indices.append(idx)
                    continue

                # 2. Match con parentesi rimosse dallo Schema e/o dal TR
                b_stripped = re.sub(r'\s*\([^)]*\)', '', b).strip().lower()
                if b_stripped == calc_name_stripped or b_stripped == not_name_stripped:
                    indices.append(idx)

            if not indices:
                log.info("[V2] INFO: bundle '%s' non trovato in Schema — colonna sempre 'No'", calc_name)

        else:
            # Caratteristica senza ordine specifico: cerca per descrizione
            # CORREZIONE V2: confronto case-insensitive su desc
            # L'ordine degli indici rispecchia l'ordine del mapping (identico a V1)
            for idx in range(5, mapping_cols_count + 5):
                b, d, r = col_structure[idx]
                if d.lower() == search_key.lower():
                    indices.append(idx)

        calc_indices[calc_name] = indices

    col_index = _build_column_index(col_structure)

    # -------------------------------------------------------------------------
    # FASE 5: Espansione condizioni e popolamento matrice
    # -------------------------------------------------------------------------
    log.info('\\n=== ELABORAZIONE ===')
    log.info('Colonne totali: %s', len(col_structure))
    log.info('Righe BOM150:   %s', len(bom_data))

    def expand_all_conditions(condition_text, original_condition):
        """
        Espande le condizioni OR e IN in righe separate.
        Gestisce anche lo split per CT/CV condivisi tra più bundle.
        """
        if not condition_text or pd.isna(condition_text):
            return [(condition_text, original_condition)]
        or_parts = parser.split_or_conditions(condition_text)
        expanded_conditions = []
        for or_part in or_parts:
            if not or_part or pd.isna(or_part):
                expanded_conditions.append((or_part, original_condition))
                continue
            # ── Priorità 1: NOT ($root.VAR IN (lista)) ───────────────────────
            # Semantica: VAR può essere qualsiasi CV del CT NON presente nella
            # lista -> espandi in una riga eq per ogni valore NON escluso.
            # Es: NOT ($root.CT000478 IN ('CV000517','CV000183',...))
            #     + Mappatura ha CV000001, CV000002, CV000517 per CT000478
            #     -> genera: $root.CT000478 eq 'CV000001'
            #               $root.CT000478 eq 'CV000002'
            #     (CV000517 è nella lista -> escluso)
            not_in_match = re.search(
                r'NOT\s*\(\s*\$root\.(\w+)\s+IN\s*\(([^)]+)\)\s*\)',
                or_part, re.IGNORECASE
            )
            if not_in_match:
                var, values_str = not_in_match.groups()
                excluded = {
                    v.strip().upper()
                    for v in re.findall(r"['\"]([^'\"]+)['\"]", values_str)
                }
                all_values = parser.get_all_values_for_char(var)
                expanded_any = False
                for val in all_values:
                    if val.strip().upper() not in excluded:
                        new_cond = (
                            or_part[:not_in_match.start()]
                            + f"$root.{var} eq '{val}'"
                            + or_part[not_in_match.end():]
                        )
                        expanded_conditions.extend(
                            expand_all_conditions(new_cond, original_condition)
                        )
                        expanded_any = True
                if not expanded_any:
                    log.warning("[WARN] NOT IN: nessun valore trovato per '%s' dopo esclusione di %s valori — riga ignorata", var, len(excluded))

            else:
                # ── Priorità 2: IN normale ────────────────────────────────────
                in_matches = list(re.finditer(r'\$root\.(\w+)\s+IN\s*\(([^)]+)\)', or_part, re.IGNORECASE))
                if in_matches:
                    # Espansione IN: genera un eq per ogni valore
                    first_match = in_matches[0]
                    var, values_str = first_match.groups()
                    values = re.findall(r"['\"]([^'\"]+)['\"]", values_str)
                    for value in values:
                        new_condition = or_part[:first_match.start()] + f"$root.{var} eq '{value}'" + or_part[first_match.end():]
                        recursive_expansions = expand_all_conditions(new_condition, original_condition)
                        expanded_conditions.extend(recursive_expansions)
                else:
                    # ── Priorità 3: ne su ZPRODGRP ───────────────────────────
                    # $root.ZPRODGRPx ne 'VAL' -> una riga eq per ogni modello in
                    # mappatura per ZPRODGRPx, escluso VAL.
                    ne_zprodgrp = re.search(
                        r'\$root\.(ZPRODGRP\w*)\s+ne\s+["\']([^"\']+)["\']',
                        or_part, re.IGNORECASE
                    )
                    if ne_zprodgrp:
                        var, excluded = ne_zprodgrp.groups()
                        all_values = parser.get_all_values_for_char(var)
                        expanded_any = False
                        for model_value in all_values:
                            if model_value.strip().upper() != excluded.strip().upper():
                                new_cond = (
                                    or_part[:ne_zprodgrp.start()]
                                    + f"$root.{var} eq '{model_value}'"
                                    + or_part[ne_zprodgrp.end():]
                                )
                                expanded_conditions.extend(
                                    expand_all_conditions(new_cond, original_condition)
                                )
                                expanded_any = True
                        if not expanded_any:
                            log.warning("[WARN] ne ZPRODGRP: nessun modello trovato per '%s' dopo esclusione di '%s' — riga ignorata", var, excluded)
                    else:
                        expanded_conditions.append((or_part, original_condition))
        return expanded_conditions

    def resolve_bundle_split(condition_text):
        """
        Determina come gestire i CT/CV condivisi tra più bundle in una condizione.

        LOGICA CORRETTA:
        ================
        1. Estrae tutti i (CHAR, VALUE) dalla condizione (solo op eq, non NOT)
        2. Per ogni CT/CV, recupera i bundle associati dallo Schema (solo positivi)
        3. Identifica i CT/CV CONDIVISI: quelli che mappano su 2+ bundle
           -> Questi sono i soli responsabili di un vero conflitto tra bundle
        4. Se NON ci sono CT/CV condivisi -> return None
           (bundle diversi da CT/CV esclusivi sono bundle INDIPENDENTI, non in conflitto)
        5. competing_bundles = unione dei bundle dai soli CT/CV condivisi
        6. Cerca discriminante: qualsiasi CT/CV (condiviso o esclusivo) che mappa
           su ESATTAMENTE 1 bundle tra competing_bundles
        7. Se discriminante trovato -> ('discriminant', winner, competing_bundles)
        8. Se nessun discriminante -> ('split', competing_bundles)

        ESEMPIO CORRETTO:
        -----------------
        Condizione: CT000418=TRUE (Touring+Adventure condiviso) AND CT001979=TRUE (Radar esclusivo)
        -> shared_cvs = [(CT000418, TRUE, [Touring, Adventure])]
        -> competing_bundles = {Touring, Adventure}
        -> CT001979 mappa su [Radar], Radar NON è in competing_bundles -> non è discriminante
        -> Nessun discriminante -> ('split', [Adventure, Touring])
        -> Radar Pack viene attivato INDIPENDENTEMENTE in entrambe le split rows ✓

        CASO SBAGLIATO (bug precedente):
        ---------------------------------
        Condizione: CT000076=CV000004 (Tech Pack esclusivo) AND CT001979=TRUE (Radar esclusivo)
        -> shared_cvs = [] -> return None ✓ -> Tech Pack e Radar Pack attivati entrambi

        Returns:
            None                                -> nessun conflitto, path normale
            ('discriminant', winner, candidates) -> un bundle vince tra i competitori
            ('split', [b1, b2, ...])            -> split necessario tra bundle competitori
        """
        if not condition_text or pd.isna(condition_text):
            return None

        # Estrai tutti i (CHAR, VALUE) con op eq dalla condizione
        # Ignora NOT (...) perché non discriminano positivamente
        # Bugfix: il vecchio lookbehind (?<!NOT\s)(?<!\() catturava il SECONDO
        # $root dentro "NOT ($root.A eq 'X' AND $root.B eq 'Y')". Strip esplicito
        # dei blocchi NOT(...) bilanciati prima del findall.
        _cleaned_cond = condition_text
        _prev = None
        while _cleaned_cond != _prev:
            _prev = _cleaned_cond
            _cleaned_cond = re.sub(
                r'NOT\s*\([^()]*\)', '', _cleaned_cond, flags=re.IGNORECASE
            )
        eq_pairs = re.findall(
            r'\$root\.(\w+)\s+eq\s+["\']([^"\']+)["\']',
            _cleaned_cond, re.IGNORECASE
        )

        if not eq_pairs:
            return None

        # Per ogni CT/CV, recupera i bundle positivi dalla mappa
        all_bundles_per_cv = []
        for char, value in eq_pairs:
            bundles = cv_bundle_map.get((char, value), [])
            if bundles:
                all_bundles_per_cv.append((char, value, bundles))

        if not all_bundles_per_cv:
            return None

        # STEP CRITICO: identifica i soli CT/CV CONDIVISI (2+ bundle)
        # Solo questi creano un vero conflitto — bundle da CT/CV esclusivi
        # sono bundle INDIPENDENTI che devono coesistere, non competere.
        shared_cvs = [(char, value, bundles)
                      for char, value, bundles in all_bundles_per_cv
                      if len(bundles) > 1]

        if not shared_cvs:
            return None  # Tutti CT/CV esclusivi -> bundle indipendenti, nessun conflitto

        # competing_bundles: solo bundle dai CT/CV condivisi
        competing_bundles = set()
        for _, _, bundles in shared_cvs:
            competing_bundles.update(bundles)

        # Cerca discriminante: qualsiasi CT/CV che mappa su ESATTAMENTE 1 bundle
        # tra quelli competitori (può essere shared o esclusivo)
        for char, value, bundles in all_bundles_per_cv:
            bundles_in_competing = [b for b in bundles if b in competing_bundles]
            if len(bundles_in_competing) == 1:
                # Discriminante trovato: identifica univocamente un bundle competitore
                winner = bundles_in_competing[0]
                return ('discriminant', winner, competing_bundles)

        # Nessun discriminante trovato: tutti i CT/CV condivisi non discriminano
        # -> split necessario su tutti i bundle competitori
        return ('split', sorted(competing_bundles))

    matrix_rows = []

    for bom_row in bom_data.itertuples():
        bom_index = bom_row.Index
        condition = getattr(bom_row, start_col.replace(' ', '_').replace('/', '_'), None)
        if condition is None:
            condition = bom_data.loc[bom_index, start_col]

        numero_componenti = getattr(bom_row, 'Numero_componenti', '') if hasattr(bom_row, 'Numero_componenti') else ''
        if not numero_componenti:
            numero_componenti = bom_data.loc[bom_index, 'Numero componenti'] if 'Numero componenti' in bom_data.columns else ''

        # Qtà componente (UMC): rilevazione robusta del nome colonna (accento à / encoding)
        qcol = next((c for c in bom_data.columns if str(c).startswith('Qt') and 'UMC' in str(c)), None)
        qta_umc = bom_data.loc[bom_index, qcol] if qcol is not None else 0
        try:
            qta_umc = float(qta_umc)
        except (TypeError, ValueError):
            qta_umc = 0.0

        if not condition or pd.isna(condition):
            row = ['' for _ in col_structure]
            row[0] = numero_componenti
            row[1] = condition if condition else ''
            row[2] = condition if condition else ''
            row[3] = condition if condition else ''
            row[4] = qta_umc
            matrix_rows.append(row)
        else:
            all_expanded_conditions = expand_all_conditions(condition, condition)
            for expanded_cond, original_cond in all_expanded_conditions:

                # ── Filtro stati market-specific ─────────────────────────────
                # Scarta le righe la cui condizione espansa contiene
                # $root.ZCOUNTRY eq 'STATE' per uno stato in States.xlsx.
                # Es: "$root.ZCOUNTRY eq 'CDN'" -> riga scartata (solo Canada).
                # "$root.ZCOUNTRY ne 'CDN'" -> riga mantenuta (tutti tranne CDN).
                if _is_state_specific(expanded_cond, states_set):
                    continue
                # ─────────────────────────────────────────────────────────────

                # Verifica se serve uno split per CT/CV condivisi
                split_result = resolve_bundle_split(expanded_cond)

                if split_result is None:
                    # -------------------------------------------------------
                    # CASO NORMALE: nessun conflitto tra bundle
                    # -------------------------------------------------------
                    row = ['' for _ in col_structure]
                    row[0] = numero_componenti
                    row[1] = original_cond
                    row[2] = expanded_cond
                    row[3] = expanded_cond
                    row[4] = qta_umc

                    included = _populate_row_from_condition_v2(
                        row, expanded_cond, parser, col_structure,
                        calc_indices, calculated_cols, col_index, bundle_set
                    )
                    if included is not False:
                        matrix_rows.append(row)

                elif split_result[0] == 'split':
                    # -------------------------------------------------------
                    # CASO SPLIT: tutti i CT/CV condivisi, nessun discriminante
                    # Crea una riga per ogni bundle candidato
                    # Ogni riga: bundle target attivo, tutti gli altri azzerati
                    # -------------------------------------------------------
                    _, split_bundles = split_result
                    calc_offset = len(calculated_cols)
                    split_bundles_lower = {b.lower() for b in split_bundles}

                    for target_bundle in split_bundles:
                        row = ['' for _ in col_structure]
                        row[0] = numero_componenti
                        row[1] = original_cond
                        row[2] = expanded_cond
                        row[3] = expanded_cond
                        row[4] = qta_umc

                        included = _populate_row_from_condition_v2(
                            row, expanded_cond, parser, col_structure,
                            calc_indices, calculated_cols, col_index, bundle_set
                        )
                        if included is False:
                            continue

                        # Azzera i bundle candidati NON target
                        target_bundle_lower = target_bundle.lower()
                        for i, col_tuple in enumerate(calculated_cols):
                            calc_name = col_tuple[2]
                            search_key = col_tuple[3]
                            if search_key is not None:
                                continue  # caratteristica fisica, non bundle
                            calc_name_lower = calc_name.lower()
                            if calc_name_lower in split_bundles_lower and calc_name_lower != target_bundle_lower:
                                row[-(calc_offset - i)] = 'No'

                        matrix_rows.append(row)

                elif split_result[0] == 'discriminant':
                    # -------------------------------------------------------
                    # CASO DISCRIMINANTE: esiste un CT/CV che identifica univocamente
                    # il bundle corretto -> una sola riga, attiva solo il winner
                    # e azzera tutti gli altri candidati
                    # -------------------------------------------------------
                    _, winner, candidates = split_result
                    calc_offset = len(calculated_cols)
                    winner_lower = winner.lower()
                    candidates_lower = {c.lower() for c in candidates}

                    row = ['' for _ in col_structure]
                    row[0] = numero_componenti
                    row[1] = original_cond
                    row[2] = expanded_cond
                    row[3] = expanded_cond
                    row[4] = qta_umc

                    included = _populate_row_from_condition_v2(
                        row, expanded_cond, parser, col_structure,
                        calc_indices, calculated_cols, col_index, bundle_set
                    )
                    if included is False:
                        continue

                    # Azzera i candidati che non sono il winner
                    for i, col_tuple in enumerate(calculated_cols):
                        calc_name = col_tuple[2]
                        search_key = col_tuple[3]
                        if search_key is not None:
                            continue  # caratteristica fisica, non bundle
                        calc_name_lower = calc_name.lower()
                        if calc_name_lower in candidates_lower and calc_name_lower != winner_lower:
                            row[-(calc_offset - i)] = 'No'

                    matrix_rows.append(row)

    log.info('[OK] Righe totali matrice: %s', len(matrix_rows))

    # -------------------------------------------------------------------------
    # FASE 6: Post-processing RIMS (configurabile da foglio Excel)
    # -------------------------------------------------------------------------
    if rims_substitutions:
        # Trova indici colonne coinvolte
        col_names = [c[2] if c[2] else c[0] for c in col_structure]
        rims_idx = None
        model_indices = {}

        for idx, readable in enumerate(col_names):
            if readable == 'Rims' or readable == 'RIMS':
                rims_idx = idx
            for base_val, subs in rims_substitutions.items():
                for model_code, _ in subs:
                    if readable == model_code:
                        model_indices[model_code] = idx

        if rims_idx is not None and model_indices:
            substituted_count = 0
            for row in matrix_rows:
                current_rims = row[rims_idx]
                if current_rims in rims_substitutions:
                    for model_code, substitution in rims_substitutions[current_rims]:
                        if model_code in model_indices and row[model_indices[model_code]]:
                            row[rims_idx] = substitution
                            substituted_count += 1
                            break
            log.info('[OK] Sostituzione Rims: %s righe modificate', substituted_count)
    else:
        # Fallback: comportamento identico a V1 se il foglio RIMS_Substitutions non esiste
        col_names = [c[2] if c[2] else c[0] for c in col_structure]
        rims_idx = None
        msv4pp_idx = None
        msv4s_idx = None
        for idx, readable in enumerate(col_names):
            if readable == 'Rims': rims_idx = idx
            elif readable == 'MSV4PP': msv4pp_idx = idx
            elif readable == 'MSV4S': msv4s_idx = idx

        if all(idx is not None for idx in [rims_idx, msv4pp_idx, msv4s_idx]):
            substituted_count = 0
            for row in matrix_rows:
                if row[rims_idx] == 'Forged':
                    if row[msv4pp_idx]:
                        row[rims_idx] = 'Forged(17"/17") - Pikes Peak'
                        substituted_count += 1
                    elif row[msv4s_idx]:
                        row[rims_idx] = 'Forged(19"/17") - Black'
                        substituted_count += 1
            log.info('[OK] Sostituzione Rims (fallback V1): %s righe modificate', substituted_count)

    # -------------------------------------------------------------------------
    # FASE 6.5: Allineamento valori caratteristiche con nomi opzione TR
    # -------------------------------------------------------------------------
    # Problema: la Mappatura usa nomi come "Viola Blast", il TR usa
    # "Viola Blast gloss". Questo causa mismatch nel matching perché
    # il simulatore genera valori basati sui nomi opzione esatti del TR.
    # Soluzione: per ogni cella caratteristica nella matrice, cercare il
    # nome opzione TR che corrisponde (exact o prefix-match non ambiguo)
    # e sostituire il valore con il nome esatto del TR.
    # -------------------------------------------------------------------------
    tr_char_options = _build_tr_char_options(tr_file)
    if tr_char_options:
        calc_offset = len(calculated_cols)
        corrected_count = 0
        for row in matrix_rows:
            for i, col_tuple in enumerate(calculated_cols):
                calc_name = col_tuple[2]   # es. "FAIRING COLOR"
                search_key = col_tuple[3]  # None = bundle, str = caratteristica
                if search_key is None:
                    continue  # Salta i bundle, processa solo le caratteristiche
                col_pos = -(calc_offset - i)
                current_value = row[col_pos]
                if not current_value or current_value == 'No':
                    continue
                aligned = _align_to_tr_option(current_value, calc_name, tr_char_options)
                if aligned != current_value:
                    row[col_pos] = aligned
                    corrected_count += 1
        log.info('[OK] Allineamento TR: %s valori corretti', corrected_count)

    # -------------------------------------------------------------------------
    # FASE 7: Salvataggio matrice in Excel
    # -------------------------------------------------------------------------
    log.info('\\n=== SALVATAGGIO ===')

    with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
        pd.DataFrame(matrix_rows).to_excel(writer, sheet_name='Matrix', index=False, header=False, startrow=3)
        ws = writer.sheets['Matrix']

        for col_idx, (bundle, desc, readable) in enumerate(col_structure, start=1):
            ws.cell(1, col_idx, bundle if bundle else '')
            ws.cell(2, col_idx, desc if desc else '')
            ws.cell(3, col_idx, readable if readable else col_structure[col_idx-1][0])

        last_col = get_column_letter(len(col_structure))
        tab = Table(displayName="MatrixTable_V2", ref=f"A3:{last_col}{len(matrix_rows)+3}")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tab)

    log.info('\\n=== RISULTATO V2 ===')
    log.info('Salvato: %s', output_file)
    log.info('Dimensioni: %s righe × %s colonne', len(matrix_rows), len(col_structure))

    return matrix_rows


# ============================================================================
# ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    excel_path = r'.\Input\V21E\BOM_150V21E.xlsx'

    process_excel_to_matrix_v2(
        excel_path=excel_path,
        mapping_file='.\\Input\\V21E\\Mappatura_V21E.xlsx',
        tr_file='.\\Input\\TR TOTALV21E - Copia.xlsx',
        output_file='.\\Input\\matrix_output_V2.xlsx'
    )


# ============================================================================
# FUNZIONI DA SKUBUNDLE_OPTIMIZED_V2 (DEDUPLICAZIONE CONFIGURAZIONI)
# ============================================================================

# ============================================================================
# HELPER: RIMOZIONE SUFFISSI PANDAS (.1 .2) E CONSOLIDAMENTO DUPLICATI
# ============================================================================

def _strip_and_consolidate(df_in: pd.DataFrame,
                            col_names_all: list,
                            raw_name_row: list) -> pd.DataFrame:
    """
    Rimuove i suffissi .1 .2 aggiunti da pandas per colonne con nome duplicato
    nel file Excel sorgente, poi consolida le coppie con lo stesso nome.

    Logica di consolidamento per colonne Yes/No:
      - 'Yes' prevale su 'No' / NaN (OR logico)
      - Garantisce che un componente associato a una caratteristica non venga
        perso per via di nomi duplicati nella matrice.

    Args:
        df_in:          DataFrame con possibili colonne .1/.2
        col_names_all:  Lista COMPLETA dei nomi pandas di df originale
        raw_name_row:   Lista COMPLETA dei nomi raw (riga Excel) di df originale

    Returns:
        DataFrame con nomi colonna puliti e senza duplicati.
    """
    # --- Costruisci mappa: pandas_name -> raw_name per le colonne presenti ---
    pname_to_raw = {}
    for pname, rname in zip(col_names_all, raw_name_row):
        pname_to_raw[pname] = rname

    # --- Rinomina le colonne di df_in usando i raw names ---
    # re.sub rimuove suffissi .1/.2/... che potrebbero essere rimasti letterali
    # nel file Excel sorgente (non generati da pandas ma scritti tali-quali)
    new_names = [re.sub(r'\.\d+$', '', pname_to_raw.get(c, c)) for c in df_in.columns]

    # Perf: Counter O(N) invece di list.count() in comprehension O(N²)
    from collections import Counter
    _name_counts = Counter(new_names)
    dup_before = {c for c, n in _name_counts.items() if n > 1}
    if not dup_before:
        # Nessun duplicato: rinomina e restituisci
        df_out = df_in.copy()
        df_out.columns = new_names
        return df_out

    log.info('\\n  [FIX .1/.2] Colonne con nome duplicato trovate: %s', dup_before)

    # --- Consolida colonne duplicate ---
    # Iteriamo per posizione per gestire correttamente più occorrenze
    unique_ordered = list(dict.fromkeys(new_names))   # ordine preservato
    consolidated = {}

    for col in unique_ordered:
        positions = [i for i, n in enumerate(new_names) if n == col]
        if len(positions) == 1:
            consolidated[col] = df_in.iloc[:, positions[0]].reset_index(drop=True)
        else:
            # OR logico generalizzato:
            #   - Qualsiasi valore "reale" (diverso da No/nan/vuoto) prevale
            #   - Tra più valori reali, 'Yes' prevale (bundle attivo)
            #   - Se nessun valore reale, resta 'No'
            _inactive = {'no', 'nan', '', 'not'}
            cols_series = [df_in.iloc[:, p].astype(str).str.strip()
                          for p in positions]
            n = len(df_in)
            merged = pd.Series(['No'] * n)
            for s in cols_series:
                for idx in range(n):
                    val = s.iloc[idx]
                    val_low = val.lower()
                    cur_low = merged.iloc[idx].lower()
                    if val_low in _inactive:
                        continue  # valore inattivo, non sovrascrive
                    if cur_low in _inactive:
                        merged.iloc[idx] = val  # primo valore reale trovato
                    elif val_low == 'yes':
                        merged.iloc[idx] = 'Yes'  # Yes vince su altri valori
            consolidated[col] = merged.reset_index(drop=True)
            log.info("    '%s': %s occorrenze -> consolidate (OR)", col, len(positions))

    return pd.DataFrame(consolidated)


# ============================================================================
# COLLAPSE PACK COLOR COLUMNS
# ============================================================================

_PACK_BUNDLES_TARGET = {'touring pack', 'top case pack', 'touring pack rally'}

def _collapse_pack_color_columns(df: pd.DataFrame,
                                  bundle_row: list,
                                  desc_row: list,
                                  raw_name_row: list) -> pd.DataFrame:
    """
    Crea colonne valore collassate per i componenti di Touring Pack e Top Case Pack.

    Per ogni gruppo (Bundle, CT_desc) nei pack target:
      - Color variant (>1 opzione OPPURE "color variant" in desc):
          valore = descrizione colore attivo (raw_name_row)
      - Pack invariant (1 opzione):
          valore = "Yes" se presente

    Regola "Yes" fallback: se nessuna opzione del CT e' attiva per questa riga
    MA il pack e' comunque attivo (almeno una opzione QUALSIASI del pack e' attiva),
    il valore diventa "Yes" anziche' "No". Questo garantisce che le righe SKU
    pack-invariant (es. Central Stand) matchino correttamente le configurazioni
    simulate con touring pack attivo, indipendentemente dal colore panniers.

    Regola "vince color variant": se un CT ha un valore colore specifico, prevale su "Yes".

    Args:
        df:            DataFrame matrice (header=2, TUTTE le colonne incluse)
        bundle_row:    Lista nomi bundle per ogni colonna (riga 0 matrice)
        desc_row:      Lista descrizioni CT per ogni colonna (riga 1 matrice)
        raw_name_row:  Lista nomi readable per ogni colonna (riga 2 matrice,
                       senza suffissi .1/.2 di pandas)

    Returns:
        DataFrame con le colonne collassate (stesso indice di df)
    """
    from collections import defaultdict

    _inactive_vals = {'', 'nan', 'none', 'no', 'not'}

    # Raccoglie colonne per (bundle_norm, ct_desc) E per bundle (per pack-activity check)
    groups = defaultdict(list)          # (bundle_norm, ct_desc) -> [(col_idx, raw_name), ...]
    bundle_all_col_indices = defaultdict(list)  # bundle_norm -> [col_idx, ...]

    for i, (b, d, rn) in enumerate(zip(bundle_row, desc_row, raw_name_row)):
        b_str = str(b).strip()
        b_norm = b_str.lower()
        d_str = str(d).strip()

        if b_norm not in _PACK_BUNDLES_TARGET:
            continue
        if not b_str or b_str in ('nan', 'None', 'none'):
            continue
        if not d_str or d_str in ('nan', 'None', 'none'):
            continue

        groups[(b_norm, d_str)].append((i, str(rn).strip()))
        bundle_all_col_indices[b_norm].append(i)

    if not groups:
        return pd.DataFrame(index=df.index)

    n_rows = len(df)

    # Identifica colonne calcolate per ogni pack: bundle='' E desc='' nel header.
    # Queste sono le colonne finali Yes/No/Not prodotte da gestione_dipendenze_V2
    # e rappresentano la fonte autorevole per stabilire se un pack e' attivo.
    # Usarle al posto delle colonne raw previene falsi positivi quando un CT/CV
    # e' condiviso tra pack incompatibili (es: CT000051/CV000025 in V22E mappa
    # sia su 'Alu Top case pack' che su 'Top Case Pack': la colonna raw e' attiva
    # per entrambi ma la colonna calcolata riflette il discriminante corretto).
    calc_pack_col_indices = {}  # bundle_norm -> col_index_in_df
    for i, (b, d, rn) in enumerate(zip(bundle_row, desc_row, raw_name_row)):
        b_str = str(b).strip()
        d_str = str(d).strip()
        rn_lower = str(rn).strip().lower()
        is_empty_bundle = not b_str or b_str in ('nan', 'None', 'none')
        is_empty_desc   = not d_str or d_str in ('nan', 'None', 'none')
        if is_empty_bundle and is_empty_desc and rn_lower in _PACK_BUNDLES_TARGET:
            calc_pack_col_indices[rn_lower] = i

    # Pre-calcola per ogni (bundle_norm, row_idx) se il pack e' attivo.
    # Usa la colonna calcolata (autorevole) se disponibile, altrimenti
    # deriva l'attivita' dalle colonne raw del mapping (comportamento precedente).
    # Perf: vettorializza valutazione attivita' pack invece di iloc cell-by-cell
    pack_active = {}  # (bundle_norm, row_idx) -> bool
    for b_norm, all_cols in bundle_all_col_indices.items():
        if b_norm in calc_pack_col_indices:
            # Colonna calcolata disponibile: usa Yes/No/Not gia' risolti
            calc_idx = calc_pack_col_indices[b_norm]
            col_vals = df.iloc[:, calc_idx].astype(str).str.strip().str.lower()
            mask = ~col_vals.isin(_inactive_vals)
            for row_idx in range(n_rows):
                pack_active[(b_norm, row_idx)] = bool(mask.iloc[row_idx])
        else:
            valid_cols = [ci for ci in all_cols if ci < df.shape[1]]
            if valid_cols:
                sub = df.iloc[:, valid_cols].astype(str).apply(
                    lambda s: s.str.strip().str.lower()
                )
                notna_mask = df.iloc[:, valid_cols].notna()
                in_inactive = sub.isin(_inactive_vals)
                active_mask = (notna_mask & ~in_inactive).any(axis=1)
            else:
                active_mask = pd.Series(False, index=df.index)
            for row_idx in range(n_rows):
                pack_active[(b_norm, row_idx)] = bool(active_mask.iloc[row_idx])

    new_cols = {}  # col_name -> list of values

    for (bundle_norm, ct_desc), col_list in groups.items():
        n_options = len(col_list)
        is_cv = ('color variant' in ct_desc.lower()) or (n_options > 1)

        col_indices   = [ci for ci, _ in col_list]
        col_raw_names = [rn for _, rn in col_list]

        values = ['No'] * n_rows

        for row_idx in range(n_rows):
            specific_found = None
            for ci, rn in zip(col_indices, col_raw_names):
                if ci >= df.shape[1]:
                    continue
                cell = df.iloc[row_idx, ci]
                if pd.notna(cell) and str(cell).strip().lower() not in _inactive_vals:
                    specific_found = rn if is_cv else 'Yes'
                    break

            if specific_found is not None:
                values[row_idx] = specific_found
            elif pack_active.get((bundle_norm, row_idx), False):
                # Pack attivo ma nessuna opzione specifica di questo CT -> "Yes"
                values[row_idx] = 'Yes'
            # else: pack non attivo -> resta 'No'

        col_name = ct_desc  # nome colonna = descrizione CT

        if col_name in new_cols:
            # Merge: color variant prevale su Yes/No
            existing = new_cols[col_name]
            for i in range(n_rows):
                if existing[i] == 'No' and values[i] != 'No':
                    existing[i] = values[i]
                elif values[i] not in ('No', 'Yes') and existing[i] in ('No', 'Yes'):
                    existing[i] = values[i]  # color prevale su Yes
        else:
            new_cols[col_name] = values

    return pd.DataFrame(new_cols, index=df.index)


# ============================================================================
# FUNZIONE DI RILEVAMENTO COLONNE (cuore del zero-hardcoding)
# ============================================================================

def _rileva_colonne(file_path: str, model_col_names=None, model_prefix: str = 'MSV4'):
    """
    Legge le 3 righe header della matrice V2 e classifica ogni colonna.

    LOGICA DI CLASSIFICAZIONE:
    --------------------------
    1. Prime 5 colonne -> metadati (escluse)
    2. Colonne con bundle='' E desc='' E posizione >= 5 -> colonne calcolate (dal TR)
    3. Colonne rimanenti (posizione >= 5, non calcolate) -> colonne mapping (dal Schema)
    4. Tra le mapping: quelle il cui nome readable corrisponde a model_col_names
       oppure al prefisso model_prefix -> colonne modello

    PERCHÉ bundle='' E desc='' identifica le calcolate:
    In gestione_dipendenze_V2.py, load_calculated_cols_from_tr produce tuple
    ('', '', calc_name, search_key). Il salvataggio Excel scrive '' in riga 1 e
    riga 2 per queste colonne. Le colonne mapping del Schema hanno sempre almeno
    una descrizione caratteristica non vuota in riga 2.

    Args:
        file_path:      Path al file matrix_output_V2.xlsx
        model_col_names: Lista esplicita di nomi colonne modello (es: ['MSV4','MSV4S'])
                         Se None, usa model_prefix per rilevamento automatico
        model_prefix:   Prefisso per rilevamento automatico modelli (default 'MSV4')

    Returns:
        dict con chiavi:
          'col_num_componenti': nome colonna "Numero componenti"
          'col_modello':        lista nomi colonne modello
          'col_calcolate':      lista nomi colonne calcolate
          'col_mapping_altre':  lista nomi altre colonne mapping (non modello, non metadati)
    """
    N_METADATA = 5  # Prime 5 colonne sempre escluse

    # Leggi le 3 righe header senza applicare nessun header
    df_raw = pd.read_excel(file_path, sheet_name='Matrix', header=None, nrows=3)

    bundle_row = df_raw.iloc[0].fillna('').astype(str).str.strip()
    desc_row   = df_raw.iloc[1].fillna('').astype(str).str.strip()
    name_row   = df_raw.iloc[2].fillna('').astype(str).str.strip()

    n_cols = len(name_row)

    # -------------------------------------------------------------------------
    # Classifica ogni colonna
    # -------------------------------------------------------------------------
    is_calcolata = []
    for i in range(n_cols):
        if i < N_METADATA:
            is_calcolata.append(False)
        else:
            # Calcolata = bundle vuoto E desc vuoto (invariante di gestione_dipendenze_V2)
            is_calcolata.append(bundle_row.iloc[i] == '' and desc_row.iloc[i] == '')

    # Leggi il DataFrame con header=2 (riga 3 Excel = nomi colonne)
    df = pd.read_excel(file_path, sheet_name='Matrix', header=2)

    col_names = list(df.columns)          # Nomi effettivi dopo pandas (gestisce duplicati con .1 .2)
    raw_names = list(name_row)            # Nomi raw dalla riga 3 Excel (senza suffisso pandas)

    # -------------------------------------------------------------------------
    # Colonna "Numero componenti" = prima colonna
    # -------------------------------------------------------------------------
    col_num_componenti = col_names[0]

    # -------------------------------------------------------------------------
    # Colonne calcolate
    # -------------------------------------------------------------------------
    col_calcolate = [col_names[i] for i, calc in enumerate(is_calcolata) if calc]

    # -------------------------------------------------------------------------
    # Colonne mapping (non metadata, non calcolate)
    # -------------------------------------------------------------------------
    mapping_indices = [i for i in range(N_METADATA, n_cols) if not is_calcolata[i]]
    col_mapping_tutte = [col_names[i] for i in mapping_indices]
    raw_mapping_names = [raw_names[i] for i in mapping_indices]

    # -------------------------------------------------------------------------
    # Colonne modello: sottoinsieme delle mapping il cui nome readable
    # corrisponde a un codice modello
    # -------------------------------------------------------------------------
    if model_col_names is not None:
        # Lista esplicita fornita dall'utente
        model_names_lower = {m.lower() for m in model_col_names}
        col_modello = [col_names[i]
                       for i in mapping_indices
                       if raw_names[i].lower() in model_names_lower
                       or col_names[i].lower() in model_names_lower]
    else:
        # Rilevamento automatico tramite prefisso
        prefix_lower = model_prefix.lower()
        col_modello = [col_names[i]
                       for i in mapping_indices
                       if raw_names[i].lower().startswith(prefix_lower)
                       or col_names[i].lower().startswith(prefix_lower)]

    col_mapping_altre = [c for c in col_mapping_tutte if c not in col_modello]

    return {
        'df':                  df,
        'col_num_componenti':  col_num_componenti,
        'col_modello':         col_modello,
        'col_calcolate':       col_calcolate,
        'col_mapping_altre':   col_mapping_altre,
        'raw_name_row':        raw_names,
        'bundle_row':          list(bundle_row),
        'desc_row':            list(desc_row),
    }


# ============================================================================
# FUNZIONE PRINCIPALE
# ============================================================================

def crea_tutte_righe_univoche_v2(file_path: str,
                                  output_path: str = None,
                                  model_col_names=None,
                                  model_prefix: str = 'MSV4',
                                  exclude_always_no: bool = True,
                                  write_output: bool = True):
    """
    Crea il file tutte_righe_univoche_V2.xlsx dalla matrice V2.

    DIFFERENZE RISPETTO A V1 (skubundle_OPTIMIZED.py):
    --------------------------------------------------
    - Nessun indice colonna hardcoded (no df.columns[4:7] o df.columns[148:])
    - Nessun nome modello hardcoded (no ['MSV4','MSV4S','MSV4PP'])
    - Rilevamento automatico colonne calcolate dalle righe header Excel
    - Rilevamento automatico colonne modello dal prefisso (default: 'MSV4')
    - Opzione per escludere colonne calcolate sempre a 'No' (inutili per dedup)
    - Foglio di debug "Colonne_Usate" nell'output per tracciabilità

    Args:
        file_path:         Path al file matrix_output_V2.xlsx
        output_path:       Path per tutte_righe_univoche_V2.xlsx
        model_col_names:   Lista esplicita nomi colonne modello (es: ['MSV4','MSV4S'])
                           Se None, usa model_prefix per rilevamento automatico
        model_prefix:      Prefisso per rilevamento automatico modelli (default 'MSV4')
        exclude_always_no: Se True (default), esclude le colonne calcolate che sono
                           sempre 'No' (es: bundle Rally non ancora in Schema)
                           Riduce la dimensione dell'output senza perdere informazioni

    Returns:
        DataFrame con le righe univoche
    """
    log.info('\\n=== SKUBUNDLE V2 — CREAZIONE RIGHE UNIVOCHE ===')
    log.info('Input:  %s', file_path)
    log.info('Output: %s', output_path)

    # -------------------------------------------------------------------------
    # FASE 1: Rilevamento dinamico delle colonne
    # -------------------------------------------------------------------------
    log.info('\\n[1/5] Analisi struttura colonne...')
    info = _rileva_colonne(file_path, model_col_names=model_col_names,
                           model_prefix=model_prefix)

    df                   = info['df']
    col_num_componenti   = info['col_num_componenti']
    col_modello          = info['col_modello']
    col_calcolate        = info['col_calcolate']

    righe_originali = df.shape[0]
    log.info('  Righe originali:        %s', righe_originali)
    log.info('  Colonne totali:         %s', df.shape[1])
    log.info('  Colonne modello (%s):   %s', len(col_modello), col_modello)
    log.info('  Colonne calcolate (%s): %s', len(col_calcolate), col_calcolate)

    # -------------------------------------------------------------------------
    # FASE 2: Filtra colonne calcolate sempre 'No' (opzionale)
    # -------------------------------------------------------------------------
    col_calcolate_usate = []
    col_calcolate_escluse = []

    for col in col_calcolate:
        vals = df[col].fillna('').astype(str)
        unique_vals = set(vals.unique()) - {'', 'nan'}
        if exclude_always_no and unique_vals == {'No'}:
            col_calcolate_escluse.append(col)
        elif exclude_always_no and not unique_vals:
            col_calcolate_escluse.append(col)
        else:
            col_calcolate_usate.append(col)

    if col_calcolate_escluse:
        log.info("\\n  Colonne calcolate escluse (sempre 'No', %s):", len(col_calcolate_escluse))
        for c in col_calcolate_escluse:
            log.info("    - '%s'", c)

    log.info('\\n  Colonne calcolate usate per deduplicazione (%s):', len(col_calcolate_usate))
    for c in col_calcolate_usate:
        log.info("    - '%s'", c)

    # -------------------------------------------------------------------------
    # FASE 3: Selezione colonne rilevanti
    # -------------------------------------------------------------------------
    log.info('\\n[2/5] Selezione colonne...')

    # Seleziona: Numero componenti + colonne modello + colonne calcolate (usate)
    colonne_da_esplodere = col_modello + col_calcolate_usate
    colonne_selezionate  = [col_num_componenti] + colonne_da_esplodere
    df_result = df[colonne_selezionate].copy()

    # -------------------------------------------------------------------------
    # FASE 3.5: Collapse colonne pack color/invariant
    # Per le colonne color-variant (panniers, top case): sovrascrive le colonne
    # pack esistenti ("Touring pack", "Top Case pack") con il valore colore.
    # Per le colonne pack-invariant (Central Stand, Grips): aggiunge come nuove.
    # Deve avvenire PRIMA della dedup per separare configurazioni per colore.
    # -------------------------------------------------------------------------
    log.info('\\n[2b/5] Collapse colonne pack color (Touring Pack, Top Case Pack)...')
    _pack_color_df = _collapse_pack_color_columns(
        df,
        bundle_row=info['bundle_row'],
        desc_row=info['desc_row'],
        raw_name_row=info['raw_name_row'],
    )
    if not _pack_color_df.empty:
        _pack_color_df = _pack_color_df.reset_index(drop=True)
        df_result = df_result.reset_index(drop=True)

        # Mappa CT description -> (lista candidati target_lower, colonna_finale)
        # target: nomi possibili nel matrix_output (gestisce sia old che new matrix).
        # colonna_finale: nome desiderato nell'output (deve coincidere con df_wide/TR).
        #
        # VECCHIO matrix: "Touring pack (color variant)" come colonna calcolata
        # NUOVO  matrix:  "Touring pack"  (dopo rigenerazione BOM con TR aggiornato)
        # Il codice prova entrambi in ordine.
        _COLOR_CT_TO_PACK_COL = {
            # ct_desc_lower              -> (candidati target lower,                       finale TR)
            'panniers covers (color variant)': (('touring pack', 'touring pack (color variant)'), 'Touring pack'),
            'top case cover (color variant)':  (('top case pack',),                               'Top Case Pack'),
            'panniers covers':                 (('touring pack rally',),                          'Touring pack Rally'),
            'valigie laterali':                (('touring pack', 'touring pack (color variant)'), 'Touring pack'),
        }
        _df_result_cols_lower = {c.lower(): c for c in df_result.columns}

        # Le CT color-variant sovrascrivono la colonna pack con il valore colore.
        # "Valigie Laterali" (RS) e' pack-invariant: sovrascrive con "Yes"
        # (nessuna scelta colore -> il matching usera' l'esatto "Yes" == "Yes").
        # Tutti gli altri CT del bundle (Central Stand, Grips, ecc.) vengono
        # ignorati: la loro presenza e' gia' catturata dal valore non-"No" nella
        # colonna pack.
        _ct_cols_ignored = []
        _renames = {}   # { colonna_attuale: colonna_finale }

        for _ct_col in _pack_color_df.columns:
            _ct_lower = _ct_col.lower()
            if _ct_lower in _COLOR_CT_TO_PACK_COL:
                _target_candidates, _final_name = _COLOR_CT_TO_PACK_COL[_ct_lower]
                # Prova i candidati in ordine finche' uno viene trovato
                _target_actual = None
                for _cand in _target_candidates:
                    _target_actual = _df_result_cols_lower.get(_cand, None)
                    if _target_actual is not None:
                        break
                if _target_actual is not None:
                    import numpy as np
                    _new_vals = _pack_color_df[_ct_col].values
                    _old_vals = df_result[_target_actual].astype(str).str.strip().values
                    # Preserva "Not" originale da gestione_dipendenze:
                    # se il colore-collapse darebbe "No" ma la colonna calcolata
                    # aveva "Not" (riga "Not X Pack"), mantieni "Not".
                    _merged = np.where(
                        (_new_vals == 'No') & (_old_vals == 'Not'),
                        'Not',
                        _new_vals
                    )
                    _n_not_preserved = int(((_new_vals == 'No') & (_old_vals == 'Not')).sum())
                    df_result[_target_actual] = _merged
                    log.info("  '%s' -> sovrascritto '%s' (Not preservati: %s)", _ct_col, _target_actual, _n_not_preserved)
                    # Segna rinomina se il nome attuale != nome finale desiderato
                    if _target_actual.lower() != _final_name.lower():
                        _renames[_target_actual] = _final_name
                else:
                    log.warning("  [WARN] Colonna pack per '%s' non trovata (candidati: %s) -> ignorato", _ct_col, _target_candidates)
            else:
                _ct_cols_ignored.append(_ct_col)

        # Rinomina colonne (es. 'Touring pack (color variant)' -> 'Touring pack')
        # per allineare i nomi a df_wide (simulazione)
        if _renames:
            df_result = df_result.rename(columns=_renames)
            for old, new in _renames.items():
                log.info("  Rinominata colonna: '%s' -> '%s'", old, new)

        if _ct_cols_ignored:
            log.info('  CT bundle ignorati (non usati per matching): %s', _ct_cols_ignored)
    else:
        log.info('  Nessuna colonna pack trovata nei pack target.')

    # -------------------------------------------------------------------------
    # FASE 4: Creazione colonna Model combinata
    # -------------------------------------------------------------------------
    log.info('[3/5] Creazione colonna Model...')

    # Una riga è "attiva" per un modello se il valore è non-vuoto e non 'No'
    # (stesso criterio della V1)
    masks = {}
    for model_col in col_modello:
        masks[model_col] = (
            df_result[model_col].notna() &
            (df_result[model_col].astype(str) != '') &
            (df_result[model_col].astype(str) != 'No')
        )

    # Costruzione vettorizzata della stringa Model
    # Usa numpy arrays per performance (identico a V1)
    n_rows = len(df_result)
    mask_arrays = {col: masks[col].values for col in col_modello}

    model_values = []
    for i in range(n_rows):
        parts = [col for col in col_modello if mask_arrays[col][i]]
        model_values.append('+'.join(parts) if parts else 'Nessuno')

    df_result['Model'] = model_values

    # Rimuovi le colonne modello originali (ora ridondanti)
    cols_to_drop = [col for col in col_modello if col in df_result.columns]
    df_result = df_result.drop(columns=cols_to_drop)

    # Sposta 'Model' come seconda colonna (dopo Numero componenti)
    cols_ordered = [col_num_componenti, 'Model'] + \
                   [c for c in df_result.columns if c not in [col_num_componenti, 'Model']]
    df_result = df_result[cols_ordered]

    # -------------------------------------------------------------------------
    # FIX V2: Strip suffissi .1/.2 e consolida colonne duplicate
    # -------------------------------------------------------------------------
    # Deve avvenire DOPO la rimozione delle colonne modello (che potrebbero avere
    # suffissi propri) e PRIMA della deduplicazione (altrimenti righe identiche
    # che differivano solo per la colonna duplicata non vengono unite).
    df_result = _strip_and_consolidate(
        df_result,
        col_names_all=list(df.columns),
        raw_name_row=info['raw_name_row']
    )
    # Aggiorna col_calcolate_usate ai nuovi nomi (senza .1/.2) e rimuovi
    # eventuali duplicati che ora puntano allo stesso nome raw.
    _p2r = dict(zip(list(df.columns), info['raw_name_row']))
    _seen_cal: set = set()
    _cal_new: list = []
    for _c in col_calcolate_usate:
        _new_c = _p2r.get(_c, _c)
        if _new_c not in _seen_cal:
            _seen_cal.add(_new_c)
            _cal_new.append(_new_c)
    col_calcolate_usate = _cal_new
    # Aggiorna anche col_calcolate_escluse (usata nei fogli debug/riepilogo)
    _seen_excl: set = set()
    _excl_new: list = []
    for _c in col_calcolate_escluse:
        _new_c = _p2r.get(_c, _c)
        if _new_c not in _seen_excl:
            _seen_excl.add(_new_c)
            _excl_new.append(_new_c)
    col_calcolate_escluse = _excl_new
    col_num_componenti  = df_result.columns[0]   # sempre prima colonna (invariante)

    # -------------------------------------------------------------------------
    # FASE 5: Deduplicazione
    # -------------------------------------------------------------------------
    log.info('[4/5] Deduplicazione...')
    log.info('  Righe prima  di drop_duplicates: %s', len(df_result))

    df_result = df_result.drop_duplicates()

    n_univoche = len(df_result)
    n_rimossi  = righe_originali - n_univoche
    pct        = (1 - n_univoche / righe_originali) * 100 if righe_originali > 0 else 0

    log.info('  Righe dopo   drop_duplicates:    %s', n_univoche)
    log.info('  Duplicati rimossi:               %s (%.2f%%)', n_rimossi, pct)

    # -------------------------------------------------------------------------
    # FASE 5.5: Espansione righe "(Ducati Red+PP)" nei pack columns
    # La voce combinata "(Ducati Red+PP)" nella matrice rappresenta due ZCOL
    # distinte ("Ducati Red" e "Pikes Peak Livery"). Ogni riga con questo valore
    # viene espansa in due righe separate per garantire il matching 1:1 con la
    # simulazione (che produce "Ducati Red" o "Pikes Peak Livery" come valori).
    # -------------------------------------------------------------------------
    _PACK_COMBINED_EXPANSIONS = [
        ('touring pack',   '(Ducati Red+PP)', ['Ducati Red', 'Pikes Peak Livery']),
        ('top case pack',  '(Ducati Red+PP)', ['Ducati Red', 'Pikes Peak Livery']),
    ]
    _df_res_cols_lower = {c.lower(): c for c in df_result.columns}
    _expansion_done = False
    for _col_lower, _combined_val, _split_vals in _PACK_COMBINED_EXPANSIONS:
        _col_actual = _df_res_cols_lower.get(_col_lower, None)
        if _col_actual is None:
            continue
        _mask = df_result[_col_actual] == _combined_val
        _n_combined = int(_mask.sum())
        if _n_combined == 0:
            continue
        _rows_combined = df_result[_mask].copy()
        df_result = df_result[~_mask].copy()
        _expanded = pd.concat(
            [_rows_combined.assign(**{_col_actual: sv}) for sv in _split_vals],
            ignore_index=True,
        )
        df_result = pd.concat([df_result, _expanded], ignore_index=True)
        log.info("  [%s] %s righe '%s' -> %s righe (split: %s)", _col_actual, _n_combined, _combined_val, _n_combined * len(_split_vals), _split_vals)
        _df_res_cols_lower = {c.lower(): c for c in df_result.columns}  # aggiorna dopo concat
        _expansion_done = True
    if not _expansion_done:
        log.info("  Nessuna riga '(Ducati Red+PP)' da espandere.")

    # Dedup post-espansione (l'espansione potrebbe creare righe gia' presenti)
    if _expansion_done:
        _pre = len(df_result)
        df_result = df_result.drop_duplicates()
        _post = len(df_result)
        if _pre != _post:
            log.info('  Dedup post-espansione: %s -> %s (%s duplicati rimossi)', _pre, _post, _pre - _post)

    # Se write_output=False, ritorna il df grezzo senza ID ne' salvataggio.
    # Usato quando si processano piu' file separatamente prima di concatenare.
    if not write_output:
        return df_result

    n_univoche = len(df_result)
    n_rimossi  = righe_originali - n_univoche
    pct        = (1 - n_univoche / righe_originali) * 100 if righe_originali > 0 else 0

    # Aggiungi ID progressivo come prima colonna
    df_result.insert(0, 'ID', range(1, n_univoche + 1))

    # -------------------------------------------------------------------------
    # FASE 6: Salvataggio
    # -------------------------------------------------------------------------
    log.info('[5/5] Salvataggio...')

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:

        # Foglio principale: righe univoche
        df_result.to_excel(writer, sheet_name='Tutte_Righe_Univoche', index=False)

        # Foglio riepilogo statistiche
        riepilogo = pd.DataFrame({
            'Metrica': [
                'Righe totali originali',
                'Righe univoche',
                'Duplicati rimossi',
                'Riduzione %',
                'Colonne modello',
                'Colonne calcolate usate',
                'Colonne calcolate escluse (sempre No)',
            ],
            'Valore': [
                righe_originali,
                n_univoche,
                n_rimossi,
                f"{pct:.2f}%",
                ', '.join(col_modello),
                ', '.join(col_calcolate_usate),
                ', '.join(col_calcolate_escluse) if col_calcolate_escluse else 'Nessuna',
            ]
        })
        riepilogo.to_excel(writer, sheet_name='Riepilogo', index=False)

        # Foglio debug: dettaglio colonne usate (tracciabilità zero-hardcoding)
        colonne_debug = []
        for col in df_result.columns:
            if col == 'ID':
                tipo = 'ID progressivo'
            elif col == col_num_componenti:
                tipo = 'Metadato'
            elif col == 'Model':
                tipo = f'Modello combinato (da: {", ".join(col_modello)})'
            elif col in col_calcolate_usate:
                tipo = 'Calcolata (dal TR)'
            else:
                tipo = 'Sconosciuto'
            colonne_debug.append({'Colonna': col, 'Tipo': tipo})

        for col in col_calcolate_escluse:
            colonne_debug.append({
                'Colonna': col,
                'Tipo': 'Calcolata — ESCLUSA (sempre No)'
            })

        pd.DataFrame(colonne_debug).to_excel(writer, sheet_name='Colonne_Usate', index=False)

    log.info('\\n=== RISULTATO SKUBUNDLE V2 ===')
    log.info('Salvato: %s', output_path)
    log.info('Dimensioni: %s righe × %s colonne', n_univoche, len(df_result.columns))
    log.info('\\nPrime 3 righe:')
    log.info('%s', df_result.head(3).to_string())

    return df_result



# ============================================================================
# FUNZIONI DA PROCESS_ALL_BOM_COMBINATIONS (ORCHESTRATORE MULTI-BOM)
# ============================================================================

import time

INPUT_DIR = Path(".") / "Input"
OUTPUT_DIR = Path(".") / "Input"

# File per quality checks
PREZZI_FILE = INPUT_DIR / "Prezzi.xlsx"
STATES_FILE = INPUT_DIR / "States.xlsx"

_MAPPATURA_UNIFICATA = INPUT_DIR / "Mappatura_Unificata.xlsx"

# Leggi dinamicamente le cartelle dei modelli
MODEL_DIR = INPUT_DIR / "MODEL"

def get_available_models(model_dir=None):
    """
    Restituisce un dizionario con i modelli disponibili e i loro percorsi.

    Scansiona `model_dir` (default: MODEL_DIR = .\\Input\\MODEL) cercando le
    sottocartelle che contengono una `Mappatura_<nome>.xlsx`. Passare model_dir
    esplicito per scoprire i modelli sotto una input_dir diversa (es. la
    pipeline che usa config.input_dir, potenzialmente assoluta).
    """
    md = Path(model_dir) if model_dir is not None else MODEL_DIR
    models = {}
    if md.exists():
        for model_folder in md.iterdir():
            if model_folder.is_dir():
                model_name = model_folder.name
                mappatura_path = model_folder / f"Mappatura_{model_name}.xlsx"
                if mappatura_path.exists():
                    models[model_name] = {
                        "in_dir": model_folder,
                        "mappatura": mappatura_path
                    }
    return models

# Ottieni i modelli disponibili
AVAILABLE_MODELS = get_available_models()

# Costruisci le combinazioni dinamicamente
COMBINAZIONI = [
    {
        "nome": model_name,
        "in_dir": model_info["in_dir"],
        "mappatura": model_info["mappatura"],
        "tr": INPUT_DIR / "TR TOTALV21E - Copia.xlsx",
    }
    for model_name, model_info in AVAILABLE_MODELS.items()
]

# Costruisci la lista delle mappature originali dinamicamente
MAPPATURE_ORIGINALI = [model_info["mappatura"] for model_info in AVAILABLE_MODELS.values()]


def _normalize_bundle(name) -> str:
    """
    Normalizza il nome di un Bundle: rimuove spazi ridondanti e uniforma le
    maiuscole/minuscole in Title Case.

    Convenzione:
      - Strip leading/trailing whitespace + collasso spazi interni
      - Title Case su ogni parola (Python str.title gestisce anche separatori -/_) 
      - Valori vuoti/NaN -> stringa vuota
    """
    if name is None or (isinstance(name, float) and np.isnan(name)):
        return ''
    s = ' '.join(str(name).split())  # strip + collasso spazi multipli
    if not s or s.lower() in ('nan', 'none'):
        return ''
    return s.title()
def build_unified_mapping(source_files=None, output_path=None) -> pd.DataFrame:
    """
    Legge le mappature originali (source_files), le unisce con dedup/merge
    case-insensitive sui bundle name e salva il risultato in output_path.

    Regole di merge per coppie (Bundle_norm, Codice) presenti in più BOM:
      - Opzioni identiche -> una sola riga (deduplicata)
      - Opzioni diverse, unione ≤ 19 -> riga unica con tutte le opzioni unite
      - Opzioni diverse, unione > 19 -> righe separate

    Args:
        source_files: lista di Path delle mappature originali (default: MAPPATURE_ORIGINALI)
        output_path:  Path dove salvare il file unificato (default: _MAPPATURA_UNIFICATA)
    """
    if source_files is None:
        source_files = MAPPATURE_ORIGINALI
    if output_path is None:
        output_path = _MAPPATURA_UNIFICATA

    files = [str(f) for f in source_files if Path(f).exists()]
    if not files:
        raise FileNotFoundError(f"Nessuna mappatura originale trovata: {source_files}")

    log.info('\\n[MAPPATURA UNIFICATA] Leggo %s file: %s', len(files), [Path(f).name for f in files])

    opz_cols  = [f'Opz{i}'  for i in range(1, 20)]
    desc_cols = [f'Desc{i}' for i in range(1, 20)]
    # IMPORTANTE: le colonne devono essere interleaved (Opz1,Desc1,Opz2,Desc2,...)
    # perché load_mapping_from_file legge coppie consecutive col_idx / col_idx+1
    interleaved_cols = [col for i in range(1, 20) for col in (f'Opz{i}', f'Desc{i}')]
    schema_cols = ['Bundle', 'Codice caratteristica', 'Descrizione caratteristica'] + interleaved_cols

    # -----------------------------------------------------------------
    # 1. Leggi e normalizza
    # -----------------------------------------------------------------
    dfs = []
    for f in files:
        df = pd.read_excel(f, sheet_name='Schema')
        # Assicura che le colonne esistano
        for c in schema_cols:
            if c not in df.columns:
                df[c] = np.nan
        df = df[schema_cols].copy()
        df['Bundle'] = df['Bundle'].apply(_normalize_bundle)
        df['_src'] = Path(f).name
        dfs.append(df)

    all_df = pd.concat(dfs, ignore_index=True)

    # Rimuovi righe prive di Codice caratteristica
    all_df = all_df[all_df['Codice caratteristica'].notna()].copy()
    all_df['Codice caratteristica'] = all_df['Codice caratteristica'].astype(str).str.strip()

    # -----------------------------------------------------------------
    # 2. Dedup / merge per (Bundle_norm_lower, Codice)
    # -----------------------------------------------------------------
    all_df['_bkey'] = all_df['Bundle'].str.lower() + '|||' + all_df['Codice caratteristica'].str.lower()

    result_rows = []
    seen_bkeys = {}  # bkey -> index nella result_rows

    # Perf: to_dict('records') ~10x più veloce di iterrows()
    for row in all_df.to_dict('records'):
        bkey = row['_bkey']
        row_opz  = [row[c] for c in opz_cols]
        row_desc = [row[c] for c in desc_cols]

        def _clean_vals(vals):
            """Filtra i valori vuoti/NaN da una lista di opzioni."""
            return [v for v in vals if pd.notna(v) and str(v).strip() not in ('', 'nan')]

        if bkey not in seen_bkeys:
            seen_bkeys[bkey] = len(result_rows)
            result_rows.append({
                'Bundle': row['Bundle'],
                'Codice caratteristica': row['Codice caratteristica'],
                'Descrizione caratteristica': row['Descrizione caratteristica'],
                **{f'Opz{i+1}':  row_opz[i]  for i in range(19)},
                **{f'Desc{i+1}': row_desc[i] for i in range(19)},
            })
        else:
            # Controlla se si può fare il merge
            idx = seen_bkeys[bkey]
            existing = result_rows[idx]
            ex_opz  = _clean_vals([existing[f'Opz{i+1}']  for i in range(19)])
            ex_desc = _clean_vals([existing[f'Desc{i+1}'] for i in range(19)])
            new_opz  = _clean_vals(row_opz)
            new_desc = _clean_vals(row_desc)

            # Opzioni da aggiungere (non già presenti)
            ex_opz_set = set(str(v) for v in ex_opz)
            to_add_opz  = [v for v in new_opz  if str(v) not in ex_opz_set]
            to_add_desc = [new_desc[new_opz.index(v)] if v in new_opz and new_opz.index(v) < len(new_desc) else ''
                             for v in to_add_opz]

            if not to_add_opz:
                # Identiche -> già presente, salta
                continue

            merged_opz  = ex_opz  + to_add_opz
            merged_desc = ex_desc + to_add_desc

            if len(merged_opz) <= 19:
                # Merge in place
                for i in range(19):
                    existing[f'Opz{i+1}']  = merged_opz[i]  if i < len(merged_opz)  else np.nan
                    existing[f'Desc{i+1}'] = merged_desc[i] if i < len(merged_desc) else np.nan
            else:
                # Non rientra -> riga separata (con nuovo bkey disambiguato)
                new_bkey = bkey + f'__{len(result_rows)}'
                seen_bkeys[new_bkey] = len(result_rows)
                result_rows.append({
                    'Bundle': row['Bundle'],
                    'Codice caratteristica': row['Codice caratteristica'],
                    'Descrizione caratteristica': row['Descrizione caratteristica'],
                    **{f'Opz{i+1}':  row_opz[i]  for i in range(19)},
                    **{f'Desc{i+1}': row_desc[i] for i in range(19)},
                })

    final_df = pd.DataFrame(result_rows, columns=schema_cols)

    log.info('[MAPPATURA UNIFICATA] Righe totali input: %s | Righe unificate: %s', len(all_df), len(final_df))

    # -----------------------------------------------------------------
    # 3. Salva
    # -----------------------------------------------------------------
    if output_path is not None:
        try:
            with pd.ExcelWriter(str(output_path), engine='openpyxl') as writer:
                final_df.to_excel(writer, sheet_name='Schema', index=False)
            log.info('[MAPPATURA UNIFICATA] Salvato: %s', Path(output_path).name)
        except PermissionError:
            raise PermissionError(
                f"Impossibile scrivere '{output_path}': il file e' aperto in Excel. "
                f"Chiudilo e rilancia lo Step 1."
            )

    return final_df


def run_quality_checks() -> tuple:
    """
    Esegue quality checks su input file comuni a tutte le combinazioni.

    Returns:
        (residual_issues, price_issues, ct_cv_issues, tr_char_options, states_set)
    """
    log.info('\\n%s', '='*80)
    log.info('QUALITY CHECKS - VALIDAZIONE INPUT FILE')
    log.info('%s', '='*80)

    residual_issues = []
    price_issues = []
    ct_cv_issues = []
    tr_char_options = {}
    states_set = set()

    # Quality check residual/quantity: rimosso (file quantity e residual.xlsx eliminato)

    # Quality check prezzi
    if PREZZI_FILE.exists():
        try:
            prezzi_df = pd.read_excel(PREZZI_FILE)
            # price_issues = _check_price_quality(prezzi_df, 'Prezzo netto')
            log.info('  [INFO] Price check: %s issues trovati', len(price_issues))
        except Exception as e:
            log.warning('  [WARN] Price check skipped: %s', e)

    # Load states (market-specific filtering)
    if STATES_FILE.exists():
        states_set = _load_states_set(str(STATES_FILE))

    # Build TR char options (per allineamento nomi)
    try:
        tr_char_options = _build_tr_char_options(str(INPUT_DIR / "TR TOTALV21E - Copia.xlsx"))
    except Exception as e:
        log.warning('  [WARN] TR char options: %s', e)

    return residual_issues, price_issues, ct_cv_issues, tr_char_options, states_set


def _cleanup_bom_files(directory) -> None:
    """Elimina i file BOM intermedi stali da una directory.

    Rimuove matrix_output_*.xlsx e tutte_righe_univoche_*.xlsx (esclusi i file
    SA, gestiti da crea_catalogo_sa). Va eseguito all'inizio dello step_1 per
    evitare che file di run precedenti (es. con supermodel diversi) restino e
    falsino il consolidato. Replica _cleanup_safety_stock_files di SolMidTerm-ALL.
    """
    d = Path(directory)
    if not d.exists():
        return
    for pattern in ("matrix_output_*.xlsx", "tutte_righe_univoche_*.xlsx"):
        for f in d.glob(pattern):
            if "_SA_" in f.name:
                continue
            try:
                f.unlink()
                log.info("[CLEANUP] Rimosso: %s", f.name)
            except Exception as e:
                log.warning("[CLEANUP] Impossibile rimuovere %s: %s", f.name, e)


def _expand_exclusions(df: pd.DataFrame, exclusions: dict) -> pd.DataFrame:
    """
    In parole semplici: se una configurazione contiene due optional che non
    possono coesistere, la riga viene divisa in piu' configurazioni valide (una
    per ciascuna alternativa ammessa).

    Per ogni riga con combinazioni incompatibili, espande in N righe valide.

    Esempio: riga con (Touring pack=Yes, Radar Pack=Yes) dove le due
    caratteristiche sono incompatibili → diventa due righe:
      - riga1: Touring pack=Yes, Radar Pack=No
      - riga2: Touring pack=No,  Radar Pack=Yes

    Per N conflitti simultanei in una riga, espansione è combinatoria
    (risolta iterativamente: ogni coppia raddoppia le righe, poi drop_duplicates).

    Port deterministico dello split BOM di SolMidTerm-ALL
    (process_all_bom_combinations._expand_exclusions). Speculare alla risoluzione
    probabilistica del Monte Carlo (montecarlo.find_conflicts): qui enumera
    entrambe le alternative valide a monte, nello step_1.

    Due forme di regola, gestite via selettori (come montecarlo.find_conflicts):
    - gruppo-gruppo (bundle binario, chiave = nome gruppo): split in due
      alternative valide (comportamento storico, invariato);
    - valore-valore ('GRUPPO:VALORE'): la coppia di valori e' inammissibile e
      riguarda gruppi obbligatori (es. sedili), che non si possono disattivare;
      la combinazione non esiste a catalogo, quindi la riga viene RIMOSSA.
    """
    if not exclusions:
        return df

    col_lower_map = {c.lower(): c for c in df.columns}

    def _is_active(val) -> bool:
        """True se l'optional e' selezionato (diverso da No/Not/Nessuno/vuoto)."""
        s = str(val).strip().lower()
        return s not in ('no', 'not', '', 'nan', 'nessuno')

    def _parse_selector(token):
        """Token -> (gruppo, valore_or_None). 'GRUPPO:VALORE' -> valore-livello."""
        token = str(token).strip()
        if ':' in token:
            g, v = token.split(':', 1)
            return (g.strip(), v.strip())
        return (token, None)

    def _sel_match(r, col, val):
        """True se il selettore (col, val) matcha la riga r."""
        cell = r.get(col, 'No')
        if val is None:
            return _is_active(cell)            # bundle: attivo (!= No/Not/...)
        return str(cell).strip() == val        # valore-livello: match esatto

    # Traduce le regole in coppie di selettori risolti sulle colonne del catalogo.
    # Ogni elemento: (col_a, val_a, col_b, val_b); val_* None = gruppo-livello.
    excl_pairs = []
    for opt_a, incompat_list in exclusions.items():
        ga, va = _parse_selector(opt_a)
        col_a = col_lower_map.get(ga.lower())
        if col_a is None:
            continue
        for opt_b in incompat_list:
            gb, vb = _parse_selector(opt_b)
            col_b = col_lower_map.get(gb.lower())
            if col_b is None:
                continue
            excl_pairs.append((col_a, va, col_b, vb))

    if not excl_pairs:
        log.info("  [Esclusioni] Nessuna colonna corrispondente nel catalogo — skip.")
        return df

    log.info("  [Esclusioni] Coppie incompatibili trovate nel catalogo: %s", excl_pairs)

    result_rows = []
    n_split = 0
    n_dropped = 0

    for _, row in df.iterrows():
        rows_in_progress = [row.to_dict()]

        for col_a, va, col_b, vb in excl_pairs:
            new_rows = []
            for r in rows_in_progress:
                if _sel_match(r, col_a, va) and _sel_match(r, col_b, vb):
                    if va is None and vb is None:
                        # gruppo-gruppo (bundle): genera due alternative valide
                        r1 = dict(r); r1[col_b] = 'No'  # mantieni A, disattiva B
                        r2 = dict(r); r2[col_a] = 'No'  # disattiva A, mantieni B
                        new_rows.extend([r1, r2])
                        n_split += 1
                    else:
                        # valore-valore: combinazione inammissibile -> drop riga
                        n_dropped += 1
                else:
                    new_rows.append(r)
            rows_in_progress = new_rows

        result_rows.extend(rows_in_progress)

    log.info("  [Esclusioni] Conflitti bundle risolti (split): %s | righe valore-valore rimosse: %s "
             "| Righe prima: %s | Righe dopo (pre-dedup): %s",
             n_split, n_dropped, len(df), len(result_rows))

    result_df = pd.DataFrame(result_rows, columns=df.columns).drop_duplicates().reset_index(drop=True)
    log.info("  [Esclusioni] Righe dopo dedup: %s", len(result_df))
    return result_df


def process_combination(combo: dict, tr_char_options: dict = None, filter_untraced: bool = True, exclusions: dict = None) -> pd.DataFrame:
    """
    Processa una singola combinazione BOM+Mappatura in un unico step integrato:
      1. Enrichment BOM grezzo → DataFrame arricchito (enrich_bom_version)
      2. Elaborazione DataFrame → matrice (process_excel_to_matrix_v2 con df_input)

    Args:
        combo: dict con chiavi 'nome', 'in_dir', 'mappatura', 'tr'
        exclusions: regole optional incompatibili (parse_exclusions). Se passate,
            le righe con coppie incompatibili vengono espanse via _expand_exclusions.

    Returns:
        DataFrame con le righe univoche della combinazione
    """
    nome          = combo["nome"]
    in_dir        = combo["in_dir"]
    mappatura_path = combo["mappatura"]
    tr_path       = combo["tr"]

    log.info('\\n%s', '='*80)
    log.info('PROCESSAMENTO COMBINAZIONE: %s', nome)
    log.info('%s', '='*80)
    log.info('  In_dir:    %s', in_dir)
    log.info('  Mappatura: %s', mappatura_path.name)
    log.info('  TR:        %s', tr_path.name)

    # -------------------------------------------------------------------------
    # STEP 0: Enrichment BOM grezzo (ex bom_enrich.py)
    # -------------------------------------------------------------------------
    log.info('\\n[0/2] Enrichment BOM grezzo per %s...', nome)
    df_enriched = enrich_bom_version(nome, in_dir, save=True)

    # Quality check CT/CV (usa il df già in memoria, nessuna lettura file extra)
    try:
        mapping_df = load_mapping_from_file(str(mappatura_path))
        ct_cv_issues = []   # decommentare _check_ct_cv_traceability se necessario
    except Exception as e:
        log.warning('  [WARN] CT/CV check skipped per %s: %s', nome, e)
        ct_cv_issues = []

    # -------------------------------------------------------------------------
    # STEP 1: Elaborazione DataFrame arricchito → matrix_output
    # -------------------------------------------------------------------------
    matrix_output_path = OUTPUT_DIR / f"matrix_output_{nome}.xlsx"

    log.info('\\n[1/2] Generazione matrix_output_%s.xlsx...', nome)
    process_excel_to_matrix_v2(
        df_input=df_enriched,           # DataFrame in memoria — nessuna lettura file
        mapping_file=str(mappatura_path),
        tr_file=str(tr_path),
        output_file=str(matrix_output_path),
        states_file=str(STATES_FILE) if STATES_FILE.exists() else None,
        prezzi_file=str(PREZZI_FILE) if PREZZI_FILE.exists() else None,
        filter_untraced=filter_untraced,
        bom_file=f"BOM_150{nome}.xlsx"
    )

    # Verifica che il file sia stato creato
    if not matrix_output_path.exists():
        raise FileNotFoundError(f"Matrix output non creato: {matrix_output_path}")

    file_size = matrix_output_path.stat().st_size
    log.info('[OK] Salvato: %s (%s bytes)', matrix_output_path.name, format(file_size, ","))

    # Attendi che il file sia completamente scritto
    time.sleep(2)

    # -------------------------------------------------------------------------
    # STEP 2: Crea tutte_righe_univoche per questa combinazione
    # -------------------------------------------------------------------------
    righe_univoche_path = OUTPUT_DIR / f"tutte_righe_univoche_{nome}.xlsx"

    log.info('\\n[2/2] Generazione tutte_righe_univoche_%s.xlsx...', nome)

    # Rileva modelli dalla mappatura per pasarli a crea_tutte_righe_univoche_v2
    bom_models_for_univoche = []
    try:
        _mapp_df = pd.read_excel(str(mappatura_path), sheet_name='Schema')
        _model_row = _mapp_df[_mapp_df['Codice caratteristica'].astype(str).str.upper() == 'ZPRODGRP1']
        if _model_row.empty:
            _model_row = _mapp_df.head(1)
        _opz_cols = [f'Opz{i}' for i in range(1, 20)]
        for _c in _opz_cols:
            if _c in _model_row.columns:
                for _v in _model_row[_c].dropna():
                    _s = str(_v).strip()
                    if _s and _s.upper() != 'FINE':  # esclude 'FINE' (marcatore di fine)
                        bom_models_for_univoche.append(_s)
        bom_models_for_univoche = list(dict.fromkeys(bom_models_for_univoche))  # dedup preserva ordine
        log.info('  Modelli rilevati per univoche: %s', bom_models_for_univoche)
    except Exception as e:
        log.warning('  [WARN] Rilevamento modelli per univoche fallito: %s', e)

    df_righe = crea_tutte_righe_univoche_v2(
        file_path=str(matrix_output_path),
        output_path=str(righe_univoche_path),
        model_col_names=bom_models_for_univoche if bom_models_for_univoche else None
    )

    # -------------------------------------------------------------------------
    # POST-PROCESSING: gestione righe Model = "Nessuno"
    #   - BOM multi-modello (es. V21E con MSV4+MSV4S+MSV4PP):
    #       ogni riga Nessuno viene espansa in N righe, una per ogni modello
    #   - BOM mono-modello (es. V22E -> MSV4RI, V24T -> MSV4RS):
    #       "Nessuno" viene sostituito direttamente con l'unico modello BOM
    #
    # I modelli vengono rilevati dinamicamente dalla riga 3 del foglio Matrix,
    # senza hardcoding: funziona anche se in futuro cambiano nomi/numero modelli.
    # -------------------------------------------------------------------------
    if 'Model' in df_righe.columns:
        # Rileva modelli dalla matrice già generata
        bom_models = []
        try:
            df_raw = pd.read_excel(str(matrix_output_path), sheet_name='Matrix',
                                   header=None, nrows=3)
            name_row = df_raw.iloc[2].fillna('').astype(str).str.strip()
            # Legge i valori modello dalla mappatura (riga ZPRODGRP1 o prima riga Bundle=No)
            try:
                _mapp_df = pd.read_excel(str(mappatura_path), sheet_name='Schema')
                _model_row = _mapp_df[_mapp_df['Codice caratteristica'].astype(str).str.upper() == 'ZPRODGRP1']
                if _model_row.empty:
                    _model_row = _mapp_df.head(1)
                _opz_cols = [f'Opz{i}' for i in range(1, 20)]
                _model_vals = set()
                for _c in _opz_cols:
                    if _c in _model_row.columns:
                        for _v in _model_row[_c].dropna():
                            _s = str(_v).strip()
                            if _s:
                                _model_vals.add(_s.lower())
                bom_models = [v for v in name_row if v.lower() in _model_vals and v]
            except Exception:
                # fallback: qualsiasi token non vuoto e non numerico nella riga 3
                bom_models = [v for v in name_row if v and not v.replace('.','').isdigit()]
            log.info('  [Model] Modelli rilevati per %s: %s', nome, bom_models)
        except Exception as e:
            log.warning('  [WARN] Rilevamento modelli BOM fallito: %s', e)

        mask_nessuno = df_righe['Model'] == 'Nessuno'
        n_nessuno = int(mask_nessuno.sum())

        if n_nessuno > 0 and bom_models:
            if len(bom_models) > 1:
                # Multi-modello: split ogni riga Nessuno in len(bom_models) righe
                righe_nessuno = df_righe[mask_nessuno].copy()
                righe_altri   = df_righe[~mask_nessuno].copy()

                esplose = []
                for model in bom_models:
                    blocco = righe_nessuno.copy()
                    blocco['Model'] = model
                    esplose.append(blocco)

                df_esplose = pd.concat(esplose, ignore_index=True)
                df_righe   = pd.concat([righe_altri, df_esplose], ignore_index=True)

                # Rinumera ID dopo l'esplosione
                if 'ID' in df_righe.columns:
                    df_righe['ID'] = range(1, len(df_righe) + 1)

                log.info("  [Model] %s righe 'Nessuno' esplose in %s (%s modelli: %s)", n_nessuno, len(df_esplose), len(bom_models), ', '.join(bom_models))
            else:
                # Mono-modello: sostituzione diretta
                df_righe.loc[mask_nessuno, 'Model'] = bom_models[0]
                log.info("  [Model] %s righe 'Nessuno' sostituiti con '%s'", n_nessuno, bom_models[0])
        elif n_nessuno > 0:
            log.warning("  [WARN] %s righe 'Nessuno' rimaste: nessun modello rilevato dalla matrice", n_nessuno)

    # -------------------------------------------------------------------------
    # POST-PROCESSING: Espansione combinazioni incompatibili (esclusioni.xlsx)
    # Righe con optional A + optional B incompatibili → N righe valide.
    # Deve avvenire DOPO l'espansione Model (Nessuno→modelli reali) perché
    # le righe esplose potrebbero introdurre nuovi conflitti.
    # -------------------------------------------------------------------------
    if exclusions:
        log.info('\\n[3/2] Espansione combinazioni incompatibili per %s...', nome)
        df_righe = _expand_exclusions(df_righe, exclusions)

    # Risalva il file con la colonna Model corretta (sovrascrive la versione
    # prodotta da crea_tutte_righe_univoche_v2 che aveva ancora i Nessuno)
    with pd.ExcelWriter(str(righe_univoche_path), engine='openpyxl') as writer:
        df_righe.to_excel(writer, sheet_name='Tutte_Righe_Univoche', index=False)
    log.info('[OK] Salvato (Model corretti + esclusioni espanse): %s', righe_univoche_path)

    # Aggiungi colonna di versione per tracciabilità nel DataFrame in memoria
    # (non viene risalvata nel file singolo; compare solo nel consolidato V2)
    df_righe.insert(1, 'Versione', nome)

    return df_righe


def consolidate_results(results: dict) -> pd.DataFrame:
    """
    Consolida tutti i risultati in un unico DataFrame con deduplicazione e tracciabilità.

    Args:
        results: dict {nome_versione: DataFrame}

    Returns:
        DataFrame consolidato con colonna 'Versione_BOM' tracciata
    """
    log.info('\\n%s', '='*80)
    log.info('CONSOLIDAMENTO RISULTATI CON DEDUPLICAZIONE')
    log.info('%s', '='*80)

    dfs = []
    for nome, df in results.items():
        log.info('  %s: %s righe univoche', nome, len(df))
        dfs.append(df)

    # Concatena mantenendo tutte le righe
    df_all = pd.concat(dfs, ignore_index=True, sort=False)

    # ─────────────────────────────────────────────────────────────────────────────
    # FILL NaN → 'No' sulle colonne caratteristiche (optional/pack/seat/etc.)
    # Serve a rendere invarianti le righe rispetto alle caratteristiche assenti
    # nella propria BOM: una riga V24T senza FAIRING COLOR diventa equivalente
    # a una riga che ha esplicitamente 'No' per quella caratteristica.
    # Questo garantisce che il matching con l'output della simulazione funzioni
    # correttamente anche tra versioni con colonne optional diverse.
    # ─────────────────────────────────────────────────────────────────────────────
    _structural = {'ID', 'ID_Globale', 'Versione', 'Versioni_BOM',
                   'Model', 'Numero componenti'}
    _char_cols = [c for c in df_all.columns if c not in _structural]
    df_all[_char_cols] = df_all[_char_cols].fillna('No')

    righe_pre_dedup = len(df_all)
    log.info('\\n  Righe PRIMA deduplicazione: %s', righe_pre_dedup)

    # ─────────────────────────────────────────────────────────────────────────────
    # DEDUPLICAZIONE INTRA-VERSIONE: rimuove solo duplicati all'interno della
    # stessa BOM (es. 13 righe V21E identiche tra loro).
    # NON deduplica mai tra versioni diverse: una riga V24T con Model=MSV4RS e
    # stessa configurazione di una riga V21E con Model=MSV4 è considerata
    # semanticamente distinta e viene mantenuta in entrambe.
    # ─────────────────────────────────────────────────────────────────────────────

    # Colonne di identità INCLUSA 'Versione' → dedup solo intra-BOM
    id_cols_intra = [col for col in df_all.columns if col not in ['ID', 'ID_Globale']]

    # Prima eseguiamo la dedup intra-versione
    df_all['Versioni_BOM'] = df_all['Versione']   # ogni riga appartiene a una sola versione
    df_consolidated = df_all.drop_duplicates(subset=id_cols_intra, keep='first').copy()

    righe_post_dedup = len(df_consolidated)
    duplicati_rimossi = righe_pre_dedup - righe_post_dedup

    log.info('  Righe DOPO deduplicazione:  %s', righe_post_dedup)
    log.info('  Duplicati intra-BOM rimossi: %s (%.1f%%)', duplicati_rimossi, (duplicati_rimossi/righe_pre_dedup*100))
    for ver in df_consolidated['Versione'].dropna().unique():
        n = (df_consolidated['Versione'] == ver).sum()
        log.info('    %s: %s righe', ver, n)

    # Renumera gli ID
    df_consolidated.insert(0, 'ID_Globale', range(1, len(df_consolidated) + 1))

    # Riordina colonne: ID_Globale, Versioni_BOM (tracciabilità), Versione, Model, resto
    cols_order = ['ID_Globale', 'Versioni_BOM', 'Versione', 'Model']
    cols_order = [c for c in cols_order if c in df_consolidated.columns]
    cols_remaining = [c for c in df_consolidated.columns if c not in cols_order]
    df_consolidated = df_consolidated[cols_order + cols_remaining]

    # Statistiche per tracciabilità
    log.info('\\n  TOTALE CONSOLIDATO: %s righe univoche', righe_post_dedup)

    # Mostra quante righe sono comuni/uniche per versione
    log.info('\\n  Distribuzione Versioni BOM:')
    versioni_dist = df_consolidated['Versioni_BOM'].value_counts()
    for versioni, count in versioni_dist.items():
        log.info('    %s: %s righe', versioni, count)

    return df_consolidated
